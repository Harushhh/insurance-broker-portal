from django.contrib import messages
from django.contrib.auth import login as auth_login
from django.contrib.auth.models import User, Group
from django.core.exceptions import PermissionDenied
from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.template.loader import render_to_string
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, F, Count, Sum, Case, When, Value, CharField
from django.db.models.functions import Coalesce
from django.core.mail import send_mail
from django.core.paginator import Paginator
from django.core.cache import cache
from django.conf import settings
from django.db import transaction
from django import forms
import csv
import json
from urllib.parse import urlencode
import logging
import os
import re
import hashlib
import hmac
import time
import tempfile
import threading
import ast
import mimetypes
import secrets

from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from collections import defaultdict
from openpyxl import Workbook

# --- DRF & SWAGGER IMPORTS ---
from rest_framework import generics
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework_api_key.permissions import HasAPIKey
from .serializers import RateMasterSerializer
# -----------------------------

from . import sso

from .models import (
    RTOMaster, MakeModelMaster, PincodeMaster, RateMaster, YesNoNAMaster,
    ProductMaster, SubProductMaster, PolicyTypeMaster,
    FuelTypeMaster, MakeModelClassMaster,
    RateGroup, AuditLog, GridDocument, UserProfile,
    ExtractionField, FieldSynonym, PolicyDocumentUpload, PolicyMISRecord,
    LockedPolicy, SupportTicket, MISFile, MappingConfiguration,
    HealthRateMaster, SpecialRateRequest, MISFailedRow,
)

# Import our Gemini AI utility and background logic engines
from .utils import extract_data_with_gemini
from .forms import (
    ExtractionFieldForm, MISUploadForm, MappingConfigurationForm,
    MgBgRateRequestForm, BgRateRequestForm,
)
from .tasks import process_mis_mapping_task, process_policy_document_task
from .health_grid_utils import (
    parse_number as parse_health_number,
    parse_percent as parse_health_percent,
    parse_grid_date as parse_health_date,
    upsert_health_rate_row,
)

logger = logging.getLogger("security")

# =========================================================
# PAGE-LEVEL ACCESS GROUPS
# Enforced by page_access_required() in insurance/urls.py — each entry here
# is one Django Group, toggled per-user from the Access Control checkboxes
# in user_management() below.
# =========================================================
PAGE_GROUPS = [
    "Can_View_Dashboard",
    "Can_View_Analysis",
    "Can_View_Motor_Payout_Rates",
    "Can_Upload_CSV",
    "Can_View_RTO_Dashboard",
    "Can_View_Make_Model_Dashboard",
    "Can_View_Audit_Log",
    "Can_View_Grid_Management",
    "Can_View_Alias_Management",
    "Can_View_Tickets",
    "Can_View_Policy_Locker",
    "Can_View_Locked_Policies",
    "Can_View_Motor_Points_Logs",
    "Can_View_AI_Rulebook",
    "Can_View_OCR_Pipeline",
    "Can_View_MIS_Register",
    "Can_View_MIS_Payout_Engine",
    "Can_View_Rate_Master_Health",
    "Can_View_Health_Rate_Master",
    "Can_View_Pincode_Dashboard",
    "Can_View_Health_Payout_Rates",
    "Can_View_Special_Rates",
    "Can_Review_Special_Rates",
    "Can_View_Life_Payout_Grid",
]


def _life_payout_grid_handoff(request, role):
    """
    There's no password on the Life Payout Grid app -- instead we sign a
    short-lived token (this user's username + role + a 2-minute expiry)
    with a secret shared between the two apps, and hand it off in the URL.
    The Next.js app verifies the signature and, if valid, trusts the
    embedded role, since this portal already checked the relevant
    permission before minting it (see the two callers below).
    """
    secret = settings.LIFE_PAYOUT_GRID_AUTH_SECRET
    if not secret:
        return HttpResponse(
            "LIFE_PAYOUT_GRID_AUTH_SECRET is not configured on this portal -- "
            "the Life Payout Grid app can't be reached until it is.",
            status=500,
        )
    expiry = int(time.time()) + 120
    payload = f"{request.user.username}:{expiry}:{role}"
    signature = hmac.new(secret.encode(), payload.encode(), hashlib.sha256).hexdigest()
    token = f"{payload}.{signature}"
    url = f"{settings.LIFE_PAYOUT_GRID_URL}/api/admin/portal-login?{urlencode({'token': token})}"
    return redirect(url)


def life_payout_grid_redirect(request):
    """Entry point for the sidebar's "Life Payout Grid" link -- requires Can_View_Life_Payout_Grid (page_access_required in urls.py, so ADMIN also gets in)."""
    return _life_payout_grid_handoff(request, "viewer")


def life_payout_grid_admin_redirect(request):
    """Entry point for the sidebar's "Update Payout Rates" link -- requires SUPER_ADMIN specifically (super_admin_required in urls.py), not just ADMIN."""
    return _life_payout_grid_handoff(request, "admin")


# =========================================================
# RATE CHECKER (unified Motor / Health / Life entry point)
# Sidebar link consolidating what used to be three separate items --
# Motor Policy Locker, Health Quote Simulator, Life Payout Grid -- behind
# one "Rate Checker" entry. The underlying pages, URLs, and permissions
# (Can_View_Policy_Locker / Can_View_Health_Payout_Rates /
# Can_View_Life_Payout_Grid) are unchanged; this is just a landing view
# that sends the user to whichever one they can access, plus a shared tab
# bar (insurance/templates/partials/rate_checker_tabs.html) included at the
# top of policy_lock_checker.html / health_payout_rates.html for switching
# between them. Life stays a plain link to the existing cross-origin
# handoff (_life_payout_grid_handoff above) -- it's a separate app on a
# different domain, so it can't be rendered inline the way Motor/Health can.
# =========================================================
RATE_CHECKER_TABS = [
    ("Can_View_Policy_Locker", "policy_lock_checker"),
    ("Can_View_Health_Payout_Rates", "health_payout_rates"),
    ("Can_View_Life_Payout_Grid", "life_payout_grid_redirect"),
]

# The single "Rate Checker" checkbox in the /user-management/ Page
# Permissions grid stands in for these three real page-access groups (see
# user_management() below). It shows checked only when a user already holds
# all three -- so saving without deliberately re-checking it revokes a
# partial subset rather than silently granting the rest.
RATE_CHECKER_GROUPS = {group_name for group_name, _ in RATE_CHECKER_TABS}


def rate_checker_entry(request):
    """
    Entry point for the sidebar's "Rate Checker" link -- redirects to
    whichever of Motor/Health/Life the user can access, in that priority
    order. Wrapped with plain login_required in urls.py (like home_dashboard);
    not page_access_required, since that only checks a single group and this
    is an any-of-three check, so it raises PermissionDenied itself instead --
    same style as staff_required/super_admin_required.
    """
    is_admin = request.user.groups.filter(name="ADMIN").exists()
    for group_name, url_name in RATE_CHECKER_TABS:
        if is_admin or request.user.groups.filter(name=group_name).exists():
            return redirect(url_name)
    raise PermissionDenied("You don't have access to any Rate Checker tool.")


# =========================================================
# INBOUND PARTNER PORTAL SSO HANDOFF
# The inbound mirror of the outbound handoff above: lets an
# already-authenticated user on an external partner portal (e.g.
# ArhamSecure's partner.arhamsecure.com) land here already logged in,
# scoped to a fixed subset of pages. See insurance/sso.py for the ticket
# signing/verification logic.
# =========================================================

# Pages an ArhamSecure-partner-portal handoff is allowed to grant -- exactly
# the three pages unified under "Rate Checker" (see RATE_CHECKER_TABS
# above), never ADMIN/SUPER_ADMIN or anything outside this set. If a second
# partner ever needs a different set, split this into a per-partner mapping
# instead of widening this one.
PARTNER_ARHAMSECURE_ALLOWED_GROUPS = {
    "Can_View_Policy_Locker",
    "Can_View_Health_Payout_Rates",
    "Can_View_Life_Payout_Grid",
}

# requested_pages may name this bundle instead of listing the 3 groups
# individually -- expanded to its members in IssueSSOTicketAPIView.post
# before the allow-list intersection below, so the partner only ever needs
# to ask for "Rate_Checker" once access to the unified page is wanted.
SSO_SCOPE_BUNDLES = {
    "Rate_Checker": PARTNER_ARHAMSECURE_ALLOWED_GROUPS,
}

# landing_page value the partner may request -> URL name to redirect to.
# Deliberately a fixed allow-list, never a raw URL/path taken from the
# request -- an open redirect here would defeat the point of scoping access.
SSO_ALLOWED_LANDING_PAGES = {
    "rate_checker": "rate_checker",
    "policy_lock_checker": "policy_lock_checker",
    "health_payout_rates": "health_payout_rates",
}


class IssueSSOTicketAPIView(APIView):
    """
    Server-to-server endpoint for the partner portal's backend: authenticated
    with a rest_framework_api_key key (never called from a browser), it
    JIT-provisions/updates the target user and mints a short-lived,
    single-use ticket for the browser-redirect leg (see sso_consume_view).
    """
    permission_classes = [HasAPIKey]

    def post(self, request):
        if not settings.PARTNER_SSO_TICKET_SECRET:
            return Response(
                {"error": "PARTNER_SSO_TICKET_SECRET is not configured on this portal."},
                status=500,
            )
        email = (request.data.get("email") or "").strip().lower()
        if not email:
            return Response({"error": "email is required"}, status=400)
        full_name = (request.data.get("full_name") or "").strip()
        requested_pages = set(request.data.get("requested_pages") or [])
        landing_page = request.data.get("landing_page") or "rate_checker"

        # Expand any bundle name (e.g. "Rate_Checker") into its member
        # groups before clamping -- this is purely a convenience for the
        # caller, the allow-list intersection right after is what actually
        # decides the grant either way.
        for bundle_name, member_groups in SSO_SCOPE_BUNDLES.items():
            if bundle_name in requested_pages:
                requested_pages |= member_groups

        # Server-side clamp: this app decides the actual grant, it never
        # trusts the caller's requested list wholesale.
        granted = sorted(requested_pages & PARTNER_ARHAMSECURE_ALLOWED_GROUPS)

        user, created = User.objects.get_or_create(
            username=email,
            defaults={"email": email},
        )
        if created:
            # No usable password -- this account can only ever be entered
            # via a freshly minted ticket, never the normal login form, so
            # it adds no credential-stuffing / password-guessing surface.
            user.set_unusable_password()
            if full_name:
                first, _, last = full_name.partition(" ")
                user.first_name, user.last_name = first, last
            user.is_active = True
            user.save()
            UserProfile.objects.get_or_create(user=user)
        elif not user.is_active:
            # An existing but deactivated account (e.g. an internal signup
            # pending approval) must not be silently reactivated by a
            # partner handoff.
            return Response({"error": "account is not active"}, status=403)

        user.groups.set(Group.objects.filter(name__in=granted))

        jti = secrets.token_urlsafe(16)
        ticket = sso.mint_ticket(user, landing_page, jti)
        AuditLog.objects.create(
            user=user,
            action="sso_ticket_issued",
            details=f"partner=arhamsecure pages={granted} landing={landing_page}",
        )
        logger.info("SSO ticket issued for %s, pages=%s", email, granted)

        redirect_url = f"{request.build_absolute_uri('/sso/consume/')}?{urlencode({'ticket': ticket})}"
        return Response({"redirect_url": redirect_url})


def sso_consume_view(request):
    """
    Public landing endpoint for the partner-portal redirect. Deliberately
    NOT wrapped in login_required -- that's the entire point of this route.
    A missing/expired/replayed/invalid ticket just falls back to the normal
    login page, same as any other unauthenticated visitor.
    """
    result = sso.verify_and_consume_ticket(request.GET.get("ticket", ""))
    if not result:
        logger.warning(
            "Rejected SSO ticket from %s",
            request.META.get("HTTP_X_FORWARDED_FOR", request.META.get("REMOTE_ADDR", "unknown")),
        )
        return redirect("login")

    user, landing_page = result
    auth_login(request, user)
    # Rotate the session key, same as a normal password login
    # (ThrottledLoginView.form_valid in auth_views.py), to prevent session
    # fixation.
    request.session.cycle_key()
    AuditLog.objects.create(user=user, action="sso_ticket_consumed", details=f"landing={landing_page}")
    logger.info("SSO login consumed for %s", user.username)

    return redirect(SSO_ALLOWED_LANDING_PAGES.get(landing_page, "home"))

# =========================================================
# NA CONFIGURATION & HELPERS
# =========================================================
NA_MAKE_MODEL_MAP = {
    "TW": ["Scooter", "Bike"],
    "Private Car": ["Car"],
    "GCV 4W": ["Truck", "Pick Up Van", "Dumper Tipper", "Trailer", "Tanker", "Goods Carrying Tractor"],
    "GCV 3W": ["Goods Carrying Rickshaw"],
    "PCV 3W": ["Passenger Carrying Rickshaw"],
    "PCV 4W": ["Taxi", "School Bus", "Other Bus"],
    "MISCD": ["MISCD-Tractor", "MISCD-Others"],
}

# Bulletproof case-insensitive dictionary matcher
def get_translated_make_model(product_name):
    if not product_name:
        return None
    # Standardize map keys to uppercase and strip spaces
    normalized_map = {str(k).strip().upper(): v for k, v in NA_MAKE_MODEL_MAP.items()}
    # Standardize the incoming database string
    clean_product = str(product_name).strip().upper()
    
    if clean_product in normalized_map:
        return ", ".join(normalized_map[clean_product])
    return None

def get_dynamic_make_model_class_list(product_id):
    """
    Returns the appropriate make/model classes for the selected product.
    Instead of just renaming 'NA', it explicitly filters the list to show valid individual classes.
    """
    all_classes = list(MakeModelClassMaster.objects.all().order_by("name"))
    
    if not product_id or not str(product_id).isdigit():
        return all_classes

    prod_obj = ProductMaster.objects.filter(id=int(product_id)).first()
    if not prod_obj:
        return all_classes

    product_name = prod_obj.name.strip()

    # If the product exists in our map, filter the dropdown to only show those classes + NA
    if product_name in NA_MAKE_MODEL_MAP:
        valid_names = [c.lower() for c in NA_MAKE_MODEL_MAP[product_name]]
        valid_names.append("na")  # Always include NA as a fallback

        filtered_classes = []
        for m in all_classes:
            if m.name.strip().lower() in valid_names:
                filtered_classes.append(m)
        return filtered_classes

    return all_classes

# =========================================================
# UTILITY FUNCTIONS
# =========================================================
def strict_match_in_cluster(search_term, cluster_string):
    if not cluster_string:
        return False
    term = str(search_term).strip().upper()
    items = [x.strip().upper() for x in str(cluster_string).split(",")]
    if term in items:
        return True
    pattern = rf"(?<![A-Z0-9]){re.escape(term)}(?![A-Z0-9])"
    for item in items:
        if re.search(pattern, item):
            return True
    return False

def parse_date(value):
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%d/%m/%Y").date()
    except Exception:
        return None

def parse_yes_no_na(value):
    def _get_or_create_code(code_str):
        obj, _ = YesNoNAMaster.objects.get_or_create(code=code_str)
        return obj

    if not value:
        return _get_or_create_code("NA")
    
    v_clean = str(value).strip().lower()
    if v_clean in ["yes", "y", "true", "1"]:
        return _get_or_create_code("YES")
    elif v_clean in ["no", "n", "false", "0"]:
        return _get_or_create_code("NO")
    else:
        return _get_or_create_code("NA")

GROUP_FIELDS = [
    "new_vehicle_makes", "insurer_vertical", "insurance_company", "product", "sub_product",
    "policy_type", "vehicle_age_min", "vehicle_age_max", "make_model_class", "fuel_type",
    "pi_od_rate", "pi_tp_rate", "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5",
    "pi_net_rate", "pi_flat_amount", "pi_vli", "pi_type",
    "tariff_min", "tariff_max", "is_ncb", "is_cpa", "cc_min", "cc_max",
    "is_zd", "from_date", "to_date", "sc_min", "sc_max", "add_tnc",
]

def normalize(v):
    if v is None:
        return ""
    return str(v).strip().lower()

def build_key_hash(cleaned_dict):
    parts = []
    for f in GROUP_FIELDS:
        v = cleaned_dict.get(f)
        if hasattr(v, "pk"):
            v = v.pk
        if hasattr(v, "strftime"):
            v = v.strftime("%Y-%m-%d")
        parts.append(normalize(v))
    key_text = "|".join(parts)
    key_hash = hashlib.sha256(key_text.encode("utf-8")).hexdigest()
    return key_hash, key_text

def unique_join(values):
    seen = set()
    out = []
    for v in values:
        if not v:
            continue
        s = str(v).strip()
        if not s:
            continue
        k = s.lower()
        if k in seen:
            continue
        seen.add(k)
        out.append(s)
    return ", ".join(out)

def split_csv_values(values):
    out = []
    for v in values:
        if not v:
            continue
        for part in str(v).split(","):
            part = part.strip()
            if part:
                out.append(part)
    return out

def apply_range_filter(qs, field_min, field_max, range_val):
    if not range_val:
        return qs
    if "-" in range_val:
        parts = range_val.split("-")
        val_min = parts[0].strip()
        val_max = parts[1].strip()
        if val_min:
            qs = qs.filter(**{field_min: val_min})
        if val_max:
            qs = qs.filter(**{field_max: val_max})
    else:
        qs = qs.filter(**{field_min: range_val})
    return qs

def get_na_class_list_for_product(product_id: str):
    if product_id and str(product_id).isdigit():
        p = ProductMaster.objects.filter(id=int(product_id)).first()
        if p:
            return NA_MAKE_MODEL_MAP.get(p.name, [])
        return []
    combined = []
    for vals in NA_MAKE_MODEL_MAP.values():
        combined.extend(vals)
    return combined

def apply_make_model_filter(qs, product_id: str, make_model_class: str):
    """
    Unified filter for Make Model Classes applied across all rate views.
    Handles fallbacks correctly so that 'NA' databases hits are included when a specific sub-class is searched.
    """
    if not make_model_class:
        return qs

    clean_mmc = str(make_model_class).strip()
    
    if clean_mmc.isdigit():
        # If passed an ID (e.g. for "Scooter"), fetch exact ID OR fallback to 'NA' if mapped
        mmc_obj = MakeModelClassMaster.objects.filter(id=int(clean_mmc)).first()
        if mmc_obj:
            mmc_name = mmc_obj.name.strip().upper()
            if mmc_name == "NA":
                qs = qs.filter(make_model_class__name__iexact="NA")
            else:
                is_na_equivalent = any(mmc_name.lower() in [c.lower() for c in mapped_classes] for mapped_classes in NA_MAKE_MODEL_MAP.values())
                if is_na_equivalent:
                    qs = qs.filter(Q(make_model_class_id=clean_mmc) | Q(make_model_class__name__iexact="NA"))
                else:
                    qs = qs.filter(make_model_class_id=clean_mmc)
        else:
            qs = qs.filter(make_model_class_id=clean_mmc)

    elif clean_mmc.upper() == "NA":
        qs = qs.filter(make_model_class__name__iexact="NA")
    else:
        is_na_equivalent = False
        for mapped_classes in NA_MAKE_MODEL_MAP.values():
            if clean_mmc.lower() in [c.lower() for c in mapped_classes]:
                is_na_equivalent = True
                break
        
        if is_na_equivalent:
            qs = qs.filter(Q(make_model_class__name__iexact=clean_mmc) | Q(make_model_class__name__iexact="NA"))
        else:
            qs = qs.filter(make_model_class__name__iexact=clean_mmc)
            
    return qs


# json.dumps() alone does not escape "</script>" — safe if only rendered as
# JSON, unsafe the moment a template marks it |safe to embed straight into an
# inline <script> block (as motor_payout_rates.html does). This is the same
# escaping django.utils.html.json_script applies internally.
_JSON_SCRIPT_ESCAPES = {ord(">"): "\\u003E", ord("<"): "\\u003C", ord("&"): "\\u0026"}


def safe_json_for_script(value):
    return json.dumps(value).translate(_JSON_SCRIPT_ESCAPES)


def get_make_mapping_context():
    all_makes_objs = list(MakeModelMaster.objects.all())
    all_individual_makes = set()

    for obj in all_makes_objs:
        if obj.make_model_cluster:
            for item in str(obj.make_model_cluster).split(","):
                item = item.strip()
                if item:
                    all_individual_makes.add(item)

    all_individual_makes = sorted(list(all_individual_makes))
    class_to_makes = defaultdict(set)

    # Built once and reused below instead of rescanning the full
    # MakeModelMaster table for every (rate row x make) pair — that was an
    # O(n*m) scan repeated on every request to this and the payout/lock-checker
    # views that call this function.
    makes_by_name = {
        obj.make_model_name.strip(): obj
        for obj in all_makes_objs
        if obj.make_model_name
    }

    # .distinct() matters a lot in practice here: many thousands of RateMaster
    # rows (differing only in numeric rate fields) commonly share the exact
    # same (make_model_class, new_vehicle_makes) pair, so without it this
    # loop was repeating identical work per duplicate row. .order_by() with
    # no arguments is required alongside it — RateMaster's default
    # Meta.ordering is "-id", and Postgres requires ORDER BY columns to
    # appear in a SELECT DISTINCT's column list, so without clearing it
    # Django silently pulls the (always-unique) id into the comparison and
    # .distinct() ends up deduplicating nothing at all.
    rate_makes = RateMaster.objects.exclude(make_model_class__isnull=True).exclude(
        new_vehicle_makes__isnull=True
    ).exclude(new_vehicle_makes="").order_by().values_list(
        "make_model_class_id", "new_vehicle_makes"
    ).distinct()

    for mmc_id, makes_str in rate_makes:
        rate_groups = [m.strip() for m in makes_str.split(",")]
        for rg in rate_groups:
            if not rg:
                continue
            obj = makes_by_name.get(rg)
            if obj and obj.make_model_cluster:
                for item in str(obj.make_model_cluster).split(","):
                    item = item.strip()
                    if item:
                        class_to_makes[str(mmc_id)].add(item)

    class_makes_mapping = {k: sorted(list(v)) for k, v in class_to_makes.items()}
    return safe_json_for_script(all_individual_makes), safe_json_for_script(class_makes_mapping), all_individual_makes

# =========================================================
# AI EXTRACTION HELPERS
# =========================================================
def safe_decimal(value):
    if value in [None, ""]:
        return None
    cleaned = str(value).strip().replace(",", "")
    cleaned = re.sub(r"[^\d.\-]", "", cleaned)
    if not cleaned:
        return None
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        return None

def safe_date_multi(value):
    if not value:
        return None
    value = str(value).strip()
    formats = [
        "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d",
        "%d/%m/%y", "%d-%m-%y", "%d %b %Y", "%d %B %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(value, fmt).date()
        except Exception:
            pass
    return None

def process_policy_document(document_obj):
    try:
        document_obj.status = PolicyDocumentUpload.STATUS_PROCESSING
        document_obj.error_message = ""
        document_obj.save(update_fields=["status", "error_message"])

        # .path isn't available on non-filesystem storage backends, and the
        # Gemini SDK wants a real local path — so pull the file down to a
        # temp file first rather than handing it a storage-backed stream.
        suffix = os.path.splitext(document_obj.uploaded_file.name)[1]
        with document_obj.uploaded_file.open("rb") as src, \
                tempfile.NamedTemporaryFile(suffix=suffix, delete=False) as tmp:
            tmp.write(src.read())
            tmp_path = tmp.name
        try:
            mapped_data = extract_data_with_gemini(tmp_path)
        finally:
            os.remove(tmp_path)

        document_obj.extracted_text = "Extracted directly via Gemini Multimodal API."
        document_obj.extraction_method = "Gemini 2.5 Flash"
        document_obj.parsed_json = mapped_data
        document_obj.status = PolicyDocumentUpload.STATUS_COMPLETED
        document_obj.processed_at = timezone.now()
        document_obj.save()

        PolicyMISRecord.objects.update_or_create(
            source_document=document_obj,
            defaults={
                "insurer_name": mapped_data.get("insurance_company") or mapped_data.get("insurer_name"),
                "policy_number": mapped_data.get("policy_number"),
                "insured_name": mapped_data.get("insured_name"),
                "vehicle_registration_number": mapped_data.get("vehicle_registration_number"),
                "vehicle_make": mapped_data.get("vehicle_make"),
                "vehicle_model": mapped_data.get("vehicle_model"),
                "vehicle_make_model": mapped_data.get("vehicle_make_model"),
                "engine_number": mapped_data.get("engine_number"),
                "chassis_number": mapped_data.get("chassis_number"),
                "fuel_type": mapped_data.get("fuel_type"),
                "cubic_capacity_cc": mapped_data.get("cubic_capacity_cc"),
                "gross_premium": safe_decimal(mapped_data.get("gross_premium")),
                "net_premium": safe_decimal(mapped_data.get("net_premium")),
                "tax_amount": safe_decimal(mapped_data.get("tax_amount")),
                "policy_start_date": safe_date_multi(mapped_data.get("policy_start_date")),
                "policy_end_date": safe_date_multi(mapped_data.get("policy_end_date")),
                "issue_date": safe_date_multi(mapped_data.get("issue_date")),
                "rto_location": mapped_data.get("rto_location"),
                "raw_ai_json": mapped_data,
                "confidence_notes": "Extracted via Gemini 2.5 Flash with Synonyms injection.",
                "ai_model_name": "gemini-2.5-flash",
            }
        )

    except Exception as e:
        print(f"\n❌ AI EXTRACTION ERROR: {str(e)}\n")
        document_obj.status = PolicyDocumentUpload.STATUS_FAILED
        document_obj.error_message = str(e)
        document_obj.processed_at = timezone.now()
        document_obj.save(update_fields=["status", "error_message", "processed_at"])

# =========================================================
# FORMS
# =========================================================
class PolicyDocumentUploadForm(forms.ModelForm):
    class Meta:
        model = PolicyDocumentUpload
        fields = ["uploaded_file"]
        widgets = {
            "uploaded_file": forms.FileInput(attrs={"accept": ".pdf,.png,.jpg,.jpeg"})
        }

    def clean_uploaded_file(self):
        f = self.cleaned_data.get("uploaded_file")
        if not f:
            raise forms.ValidationError("Please select a file.")
        ext = f.name.split(".")[-1].lower() if "." in f.name else ""
        if ext not in ["pdf", "png", "jpg", "jpeg"]:
            raise forms.ValidationError("Only PDF, PNG, JPG, JPEG files are allowed.")
        return f

class RTOForm(forms.ModelForm):
    class Meta:
        model = RTOMaster
        fields = ["rto_name", "rto_cluster"]

class MakeModelForm(forms.ModelForm):
    class Meta:
        model = MakeModelMaster
        fields = ["make_model_name", "make_model_cluster"]

class PincodeForm(forms.ModelForm):
    class Meta:
        model = PincodeMaster
        fields = ["pincode_zone", "pincode_cluster"]

# RTOMaster/MakeModelMaster back the new_rto_list/new_vehicle_makes multi-select
# choices on RateForm, rebuilt on every GET and POST of the (frequently used)
# Bulk Edit Rate Group page. Both tables are near-static reference data, so a
# short cache avoids re-querying them on every request; invalidated explicitly
# wherever these tables are written to (see api_upload_chunk) rather than
# relying solely on the TTL, so new master rows show up immediately.
RTO_MAKE_CHOICES_CACHE_KEY = "rate_form_rto_make_choices_v1"
RTO_MAKE_CHOICES_CACHE_TTL = 600  # seconds


def get_rto_and_make_choices():
    cached = cache.get(RTO_MAKE_CHOICES_CACHE_KEY)
    if cached is not None:
        return cached
    data = {
        "rtos": list(RTOMaster.objects.values_list("rto_name", flat=True).order_by("rto_name")),
        "makes": list(MakeModelMaster.objects.values_list("make_model_name", flat=True).order_by("make_model_name")),
    }
    cache.set(RTO_MAKE_CHOICES_CACHE_KEY, data, RTO_MAKE_CHOICES_CACHE_TTL)
    return data


class RateForm(forms.ModelForm):
    product = forms.ModelChoiceField(
        queryset=ProductMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    sub_product = forms.ModelChoiceField(
        queryset=SubProductMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    policy_type = forms.ModelChoiceField(
        queryset=PolicyTypeMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    fuel_type = forms.ModelChoiceField(
        queryset=FuelTypeMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    make_model_class = forms.ModelChoiceField(
        queryset=MakeModelClassMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    is_ncb = forms.ModelChoiceField(
        queryset=YesNoNAMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    is_cpa = forms.ModelChoiceField(
        queryset=YesNoNAMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    is_zd = forms.ModelChoiceField(
        queryset=YesNoNAMaster.objects.none(),
        required=False,
        widget=forms.Select(attrs={"class": "select2-single"})
    )
    new_rto_list = forms.MultipleChoiceField(
        required=False,
        widget=forms.SelectMultiple(attrs={"class": "select2-multiple"})
    )
    new_vehicle_makes = forms.MultipleChoiceField(
        required=False,
        widget=forms.SelectMultiple(attrs={"class": "select2-multiple"})
    )

    class Meta:
        model = RateMaster
        exclude = ["group", "created_at"]

    def __init__(self, *args, **kwargs):
        super(RateForm, self).__init__(*args, **kwargs)

        self.fields["product"].queryset = ProductMaster.objects.all()
        self.fields["sub_product"].queryset = SubProductMaster.objects.all()
        self.fields["policy_type"].queryset = PolicyTypeMaster.objects.all()
        self.fields["fuel_type"].queryset = FuelTypeMaster.objects.all()
        self.fields["make_model_class"].queryset = MakeModelClassMaster.objects.all()
        self.fields["is_ncb"].queryset = YesNoNAMaster.objects.all()
        self.fields["is_cpa"].queryset = YesNoNAMaster.objects.all()
        self.fields["is_zd"].queryset = YesNoNAMaster.objects.all()

        # insurance_company/insurer_vertical are plain CharField/TextInput on
        # this form (never declared as ChoiceField above), so `.choices` here
        # was previously computed — via a DISTINCT scan of the ~100k-row
        # RateMaster table, on every single GET and POST — and then silently
        # discarded: TextInput ignores it for rendering, and CharField.clean()
        # never consults it either. Intentionally not rebuilding it.

        rto_make_choices = get_rto_and_make_choices()

        makes = list(rto_make_choices["makes"])
        initial_makes = kwargs.get("initial", {}).get("new_vehicle_makes", [])
        instance_makes = [x.strip() for x in self.instance.new_vehicle_makes.split(",")] if self.instance and self.instance.new_vehicle_makes else []
        all_makes = set(instance_makes + initial_makes)
        for m in all_makes:
            if m and m not in makes:
                makes.append(m)
        self.fields["new_vehicle_makes"].choices = [(m, m) for m in makes]

        rtos = list(rto_make_choices["rtos"])
        initial_rtos = kwargs.get("initial", {}).get("new_rto_list", [])
        instance_rtos = [x.strip() for x in self.instance.new_rto_list.split(",")] if self.instance and self.instance.new_rto_list else []
        all_existing_rtos = set(instance_rtos + initial_rtos)
        for r in all_existing_rtos:
            if r and r not in rtos:
                rtos.append(r)
        self.fields["new_rto_list"].choices = [(r, r) for r in rtos]

    def clean_new_rto_list(self):
        data = self.cleaned_data.get("new_rto_list")
        return ", ".join(data) if data else ""

    def clean_new_vehicle_makes(self):
        data = self.cleaned_data.get("new_vehicle_makes")
        return ", ".join(data) if data else ""

# =========================================================
# UNIFIED HOME DASHBOARD
# =========================================================
def home_dashboard(request):
    is_admin_user = True
    qs = PolicyMISRecord.objects.select_related("source_document", "source_document__uploaded_by")

    total_cases = qs.count()
    total_premium = qs.aggregate(total=Sum('gross_premium'))['total'] or 0

    product_mix_query = qs.annotate(
        category=Case(
            When(
                Q(vehicle_registration_number__isnull=False) & ~Q(vehicle_registration_number=""), 
                then=Value('Motor')
            ),
            When(
                Q(vehicle_make__isnull=False) & ~Q(vehicle_make=""), 
                then=Value('Motor')
            ),
            default=Value('Health'),
            output_field=CharField(),
        )
    ).values('category').annotate(count=Count('id'))

    product_mix = {item['category']: item['count'] for item in product_mix_query}
    motor_count = product_mix.get('Motor', 0)
    health_count = product_mix.get('Health', 0)

    recent_activity = qs.order_by('-created_at')[:5]

    my_tickets_qs = SupportTicket.objects.filter(
        user=request.user,
        created_at__gte=timezone.now() - timedelta(hours=72)
    ).order_by('-created_at')
    my_open_ticket_count = my_tickets_qs.exclude(status="CLOSED").count()
    my_tickets = my_tickets_qs[:20]

    my_locked_qs = LockedPolicy.objects.filter(locked_by=request.user, status="LOCKED").order_by('-locked_at')
    my_locked_count = my_locked_qs.count()
    my_locked_policies = my_locked_qs[:20]

    context = {
        'is_admin_user': is_admin_user,
        'total_cases': total_cases,
        'total_premium': total_premium,
        'motor_count': motor_count,
        'health_count': health_count,
        'recent_activity': recent_activity,
        'my_tickets': my_tickets,
        'my_open_ticket_count': my_open_ticket_count,
        'my_locked_policies': my_locked_policies,
        'my_locked_count': my_locked_count,
        'my_notification_count': my_open_ticket_count + my_locked_count,
    }
    return render(request, 'insurance/home.html', context)

# ---------------------------------------------------------
# NEW ENTERPRISE STREAMING CSV UPLOAD & API CHUNKING LOGIC
# ---------------------------------------------------------

def import_data_view(request):
    """
    Renders the beautiful SaaS upload interface. 
    The heavy lifting is handled via JS (PapaParse) to bypass Server RAM and timeouts.
    """
    return render(request, "upload.html")


def api_upload_chunk(request):
    """
    Receives JSON chunks from the PapaParse frontend uploader.
    Uses bulk_create to efficiently insert data while mapping your specific schemas.
    """
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            rows = data.get('rows', [])
            target_table = data.get('target_table', 'rate_master')
            
            if not rows:
                return JsonResponse({'status': 'error', 'message': 'No data provided'}, status=400)

            inserted = 0

            with transaction.atomic():
                if target_table == 'rto_master':
                    for row in rows:
                        rto_name = (row.get("rto_name") or "").strip()
                        rto_cluster = (row.get("rto_cluster") or "").strip()
                        if not rto_name:
                            raise ValueError("rto_name is blank")
                        RTOMaster.objects.update_or_create(
                            rto_name=rto_name,
                            defaults={"rto_cluster": rto_cluster or None}
                        )
                        inserted += 1
                    cache.delete(RTO_MAKE_CHOICES_CACHE_KEY)

                elif target_table == 'make_model_master':
                    for row in rows:
                        make_model_name = (row.get("make_model_name") or "").strip()
                        make_model_cluster = (row.get("make_model_cluster") or "").strip()
                        if not make_model_name:
                            raise ValueError("make_model_name is blank")
                        MakeModelMaster.objects.update_or_create(
                            make_model_name=make_model_name,
                            defaults={"make_model_cluster": make_model_cluster or None}
                        )
                        inserted += 1
                    cache.delete(RTO_MAKE_CHOICES_CACHE_KEY)

                elif target_table == 'pincode_master':
                    for row in rows:
                        pincode_zone = (row.get("pincode_zone") or "").strip()
                        pincode_cluster = (row.get("pincode_cluster") or "").strip()
                        if not pincode_zone:
                            raise ValueError("pincode_zone is blank")
                        PincodeMaster.objects.update_or_create(
                            pincode_zone=pincode_zone,
                            defaults={"pincode_cluster": pincode_cluster or None}
                        )
                        inserted += 1

                elif target_table == 'rate_master':
                    valid_rtos = {str(rto).lower() for rto in RTOMaster.objects.values_list("rto_name", flat=True) if rto}
                    valid_makes = {str(make).lower() for make in MakeModelMaster.objects.values_list("make_model_name", flat=True) if make}

                    products_map = {p.name.lower(): p for p in ProductMaster.objects.all()}
                    sub_products_map = {sp.name.lower(): sp for sp in SubProductMaster.objects.all()}
                    policy_types_map = {pt.name.lower(): pt for pt in PolicyTypeMaster.objects.all()}
                    fuel_types_map = {ft.name.lower(): ft for ft in FuelTypeMaster.objects.all()}
                    mmc_classes_map = {m.name.lower(): m for m in MakeModelClassMaster.objects.all()}
                    
                    ynn_map = {y.code.lower(): y for y in YesNoNAMaster.objects.all()}
                    for code in ["YES", "NO", "NA"]:
                        if code.lower() not in ynn_map:
                            ynn_map[code.lower()] = YesNoNAMaster.objects.create(code=code)

                    existing_groups = {g.key_hash: g for g in RateGroup.objects.all()}

                    def get_ynn(val):
                        if not val: return ynn_map["na"]
                        v = str(val).strip().lower()
                        if v in ["yes", "y", "true", "1"]: return ynn_map["yes"]
                        if v in ["no", "n", "false", "0"]: return ynn_map["no"]
                        return ynn_map["na"]

                    def get_master(val, mapping_dict, ModelClass):
                        if not val: return None
                        v = str(val).strip()
                        if not v: return None
                        if v.isdigit() and ModelClass != ProductMaster:
                            obj = ModelClass.objects.filter(id=int(v)).first()
                            if obj:
                                mapping_dict[obj.name.lower()] = obj
                                return obj
                        v_low = v.lower()
                        if v_low in mapping_dict:
                            return mapping_dict[v_low]
                        obj = ModelClass.objects.create(name=v)
                        mapping_dict[v_low] = obj
                        return obj

                    instances_to_create = []

                    for row in rows:
                        raw_rtos = row.get("new_rto_list") or ""
                        rto_items = [x.strip() for x in raw_rtos.split(",") if x.strip()]
                        for rto in rto_items:
                            if rto.lower() not in valid_rtos:
                                raise ValueError(f"RTO '{rto}' does not exist in RTOMaster. Please add it first.")

                        raw_makes = row.get("new_vehicle_makes") or ""
                        make_items = [x.strip() for x in raw_makes.split(",") if x.strip()]
                        for make in make_items:
                            if make.lower() not in valid_makes:
                                raise ValueError(f"Vehicle Make '{make}' does not exist in MakeModelMaster. Please add it first.")

                        product_val = str(row.get("product", "")).strip()
                        product_obj = None
                        if product_val:
                            if product_val.isdigit():
                                product_obj = ProductMaster.objects.filter(id=int(product_val)).first()
                            else:
                                if product_val.lower() in products_map:
                                    product_obj = products_map[product_val.lower()]
                                else:
                                    product_obj = ProductMaster.objects.create(name=product_val)
                                    products_map[product_val.lower()] = product_obj

                        sub_product_obj = get_master(row.get("sub_product"), sub_products_map, SubProductMaster)
                        policy_type_obj = get_master(row.get("policy_type"), policy_types_map, PolicyTypeMaster)
                        fuel_type_obj = get_master(row.get("fuel_type"), fuel_types_map, FuelTypeMaster)
                        mmc_obj = get_master(row.get("make_model_class"), mmc_classes_map, MakeModelClassMaster)
                        
                        is_ncb_obj = get_ynn(row.get("is_ncb"))
                        is_cpa_obj = get_ynn(row.get("is_cpa"))
                        is_zd_obj = get_ynn(row.get("is_zd"))

                        cleaned = {
                            "new_vehicle_makes": row.get("new_vehicle_makes") or None,
                            "insurer_vertical": row.get("insurer_vertical") or None,
                            "insurance_company": str(row.get("insurance_company", "")).strip(),
                            "product": product_obj,
                            "sub_product": sub_product_obj,
                            "policy_type": policy_type_obj,
                            "vehicle_age_min": float(row.get("vehicle_age_min") or 0),
                            "vehicle_age_max": float(row.get("vehicle_age_max") or 0),
                            "make_model_class": mmc_obj,
                            "fuel_type": fuel_type_obj,
                            "pi_od_rate": float(row.get("pi_od_rate") or 0),
                            "pi_tp_rate": float(row.get("pi_tp_rate") or 0),
                            "pi_tp_2": float(row.get("pi_tp_2") or 0),
                            "pi_tp_3": float(row.get("pi_tp_3") or 0),
                            "pi_tp_4": float(row.get("pi_tp_4") or 0),
                            "pi_tp_5": float(row.get("pi_tp_5") or 0),
                            "pi_net_rate": float(row.get("pi_net_rate") or 0),
                            "pi_flat_amount": float(row.get("pi_flat_amount") or 0),
                            "pi_vli": float(row.get("pi_vli") or 0),
                            "pi_type": row.get("pi_type") or None,
                            "tariff_min": float(row.get("tariff_min") or 0),
                            "tariff_max": float(row.get("tariff_max") or 0),
                            "is_ncb": is_ncb_obj,
                            "is_cpa": is_cpa_obj,
                            "is_zd": is_zd_obj,
                            "cc_min": float(row.get("cc_min") or 0),
                            "cc_max": float(row.get("cc_max") or 0),
                            "from_date": parse_date(row.get("from_date")),
                            "to_date": parse_date(row.get("to_date")),
                            "sc_min": float(row.get("sc_min") or 0),
                            "sc_max": float(row.get("sc_max") or 0),
                            "user_id": int(float(row.get("user_id") or 0)) if row.get("user_id") else None,
                            "veh_use": row.get("veh_use") or None,
                            "add_tnc": row.get("add_tnc") or None,
                            "remarks": row.get("remarks") or None,
                            "po_type": row.get("po_type") or None,
                            "po_od_rate": float(row.get("po_od_rate") or 0),
                            "po_tp_rate": float(row.get("po_tp_rate") or 0),
                            "po_net_rate": float(row.get("po_net_rate") or 0),
                            "po_flat_amount": float(row.get("po_flat_amount") or 0),
                        }

                        key_hash, key_text = build_key_hash(cleaned)
                        if key_hash in existing_groups:
                            group_obj = existing_groups[key_hash]
                        else:
                            group_obj = RateGroup.objects.create(key_hash=key_hash, key_text=key_text)
                            existing_groups[key_hash] = group_obj

                        instances_to_create.append(
                            RateMaster(
                                group=group_obj,
                                status="INACTIVE",
                                is_deleted="NO",
                                new_vehicle_makes=cleaned["new_vehicle_makes"],
                                new_rto_list=row.get("new_rto_list") or None,
                                insurer_vertical=cleaned["insurer_vertical"],
                                insurance_company=cleaned["insurance_company"],
                                product=product_obj,
                                sub_product=sub_product_obj,
                                policy_type=policy_type_obj,
                                fuel_type=fuel_type_obj,
                                make_model_class=mmc_obj,
                                vehicle_age_min=cleaned["vehicle_age_min"],
                                vehicle_age_max=cleaned["vehicle_age_max"],
                                pi_od_rate=cleaned["pi_od_rate"],
                                pi_tp_rate=cleaned["pi_tp_rate"],
                                pi_tp_2=cleaned["pi_tp_2"],
                                pi_tp_3=cleaned["pi_tp_3"],
                                pi_tp_4=cleaned["pi_tp_4"],
                                pi_tp_5=cleaned["pi_tp_5"],
                                pi_net_rate=cleaned["pi_net_rate"],
                                pi_flat_amount=cleaned["pi_flat_amount"],
                                pi_vli=cleaned["pi_vli"],
                                pi_type=cleaned["pi_type"],
                                tariff_min=cleaned["tariff_min"],
                                tariff_max=cleaned["tariff_max"],
                                is_ncb=is_ncb_obj,
                                is_cpa=is_cpa_obj,
                                is_zd=is_zd_obj,
                                cc_min=cleaned["cc_min"],
                                cc_max=cleaned["cc_max"],
                                from_date=cleaned["from_date"],
                                to_date=cleaned["to_date"],
                                user_id=cleaned["user_id"],
                                sc_min=cleaned["sc_min"],
                                sc_max=cleaned["sc_max"],
                                veh_use=cleaned["veh_use"],
                                add_tnc=cleaned["add_tnc"],
                                remarks=cleaned["remarks"],
                                po_type=cleaned["po_type"],
                                po_od_rate=cleaned["po_od_rate"],
                                po_tp_rate=cleaned["po_tp_rate"],
                                po_net_rate=cleaned["po_net_rate"],
                                po_flat_amount=cleaned["po_flat_amount"],
                            )
                        )

                    RateMaster.objects.bulk_create(instances_to_create, ignore_conflicts=True)
                    inserted = len(instances_to_create)

                elif target_table == 'health_rate_master':
                    for row in rows:
                        insurance_company = (row.get("insurance_company") or "").strip()
                        if not insurance_company:
                            raise ValueError("insurance_company is blank")

                        cleaned = {
                            "insurance_company": insurance_company,
                            "product_name": (row.get("product_name") or "").strip() or None,
                            "policy_category": (row.get("policy_category") or "").strip() or None,
                            "plan_names": (row.get("plan_names") or "").strip() or None,
                            "business_type": (row.get("business_type") or "").strip() or None,
                            "min_deductible": parse_health_number(row.get("min_deductible")),
                            "max_deductible": parse_health_number(row.get("max_deductible")),
                            "min_sum_insured": parse_health_number(row.get("min_sum_insured")),
                            "max_sum_insured": parse_health_number(row.get("max_sum_insured")),
                            "min_age": parse_health_number(row.get("min_age")),
                            "max_age": parse_health_number(row.get("max_age")),
                            "pincode_zone": (row.get("pincode_zone") or "").strip() or None,
                            "from_date": parse_health_date(row.get("from_date")),
                            "to_date": parse_health_date(row.get("to_date")),
                            "payin_rate": parse_health_percent(row.get("payin_rate")),
                            "one_year_rate": parse_health_percent(row.get("one_year_rate")),
                            "multi_year_2_rate": parse_health_percent(row.get("multi_year_2_rate")),
                            "multi_year_3_rate": parse_health_percent(row.get("multi_year_3_rate")),
                            "multi_year_4_rate": parse_health_percent(row.get("multi_year_4_rate")),
                            "multi_year_5_rate": parse_health_percent(row.get("multi_year_5_rate")),
                        }

                        # status/is_deleted/remarks are optional columns — only
                        # touch them on update if this row actually provided a
                        # value, so a re-upload that omits them doesn't clobber
                        # manual edits made afterwards in the UI (upsert_health_
                        # rate_row only defaults them ACTIVE/NO on creation).
                        status_val = (row.get("status") or "").strip().upper()
                        if status_val:
                            if status_val not in dict(HealthRateMaster.STATUS_CHOICES):
                                raise ValueError(f"'{status_val}' is not a valid Status (ACTIVE or INACTIVE).")
                            cleaned["status"] = status_val
                        is_deleted_val = (row.get("is_deleted") or "").strip().upper()
                        if is_deleted_val:
                            if is_deleted_val not in dict(HealthRateMaster.IS_DELETED_CHOICES):
                                raise ValueError(f"'{is_deleted_val}' is not a valid Is Deleted value (YES or NO).")
                            cleaned["is_deleted"] = is_deleted_val
                        remarks_val = (row.get("remarks") or "").strip()
                        if remarks_val:
                            cleaned["remarks"] = remarks_val

                        upsert_health_rate_row(cleaned)
                        inserted += 1

                else:
                    return JsonResponse({'status': 'error', 'message': 'Invalid table selected'}, status=400)

            return JsonResponse({'status': 'success', 'inserted': inserted})

        except Exception as e:
            print(f"\n[UPLOAD ERROR] {str(e)}\n")
            return JsonResponse({'status': 'error', 'message': str(e)}, status=500)

    return JsonResponse({'status': 'error', 'message': 'Method not allowed'}, status=405)

# -------------------------
# Dashboard (GROUPED view)
# -------------------------
DASHBOARD_BATCH_SIZE = 50


def dashboard(request):
    qs = RateMaster.objects.select_related(
        "group", "product", "sub_product", "policy_type", "fuel_type",
        "make_model_class", "is_ncb", "is_cpa", "is_zd"
    ).all()

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    is_deleted_filter = (request.GET.get("is_deleted") or "").strip().upper()
    created_date = (request.GET.get("created_date") or "").strip()
    date_range = (request.GET.get("date_range") or "").strip()

    insurance_company = (request.GET.get("insurance_company") or "").strip()
    product = (request.GET.get("product") or "").strip()
    fuel = (request.GET.get("fuel") or "").strip()
    sub_product = (request.GET.get("sub_product") or "").strip()
    make_model_class = (request.GET.get("make_model_class") or "").strip()
    rto_code = (request.GET.get("rto_code") or "").strip()
    make_model_code = (request.GET.get("make_model_code") or "").strip()
    age_range = (request.GET.get("age_range") or "").strip()
    cc_range = (request.GET.get("cc_range") or "").strip()
    sc_range = (request.GET.get("sc_range") or "").strip()
    is_zd = (request.GET.get("is_zd") or "").strip()
    is_ncb = (request.GET.get("is_ncb") or "").strip()
    is_cpa = (request.GET.get("is_cpa") or "").strip()

    filter_count = sum(1 for v in [
        q, status_filter, is_deleted_filter, created_date, date_range,
        insurance_company, product, fuel, sub_product, make_model_class,
        rto_code, make_model_code, age_range, cc_range, sc_range,
        is_zd, is_ncb, is_cpa,
    ] if v)

    if filter_count < 2:
        # No filter selected yet -- skip the expensive full-table query/scan
        # entirely and just render the filter form with an empty result set.
        return render(request, "dashboard.html", {
            "data": [],
            "page_obj": None,
            "elided_page_range": [],
            "field_names": [],
            "bulk_update_fields": ALLOWED_BULK_UPDATE_FIELDS,
            "total": 0,
            "active_count": 0,
            "inactive_count": 0,
            "insurance_company_list": RateMaster.objects.exclude(insurance_company="").values_list(
                "insurance_company", flat=True
            ).distinct().order_by("insurance_company"),
            "product_list": ProductMaster.objects.all().order_by("name"),
            "fuel_list": FuelTypeMaster.objects.all().order_by("name"),
            "sub_product_list": SubProductMaster.objects.all().order_by("name"),
            "make_model_class_list": get_dynamic_make_model_class_list(product),
            "yes_no_na_list": YesNoNAMaster.objects.all().order_by("code"),
            "make_class_mapping_json": json.dumps(NA_MAKE_MODEL_MAP),
            "all_make_classes_json": json.dumps(list(MakeModelClassMaster.objects.exclude(name__iexact="NA").values('id', 'name'))),
            "selected": {
                "q": q,
                "status": status_filter,
                "is_deleted": is_deleted_filter,
                "created_date": created_date,
                "insurance_company": insurance_company,
                "product": product,
                "fuel": fuel,
                "sub_product": sub_product,
                "make_model_class": make_model_class,
                "date_range": date_range,
                "rto_code": rto_code,
                "make_model_code": make_model_code,
                "age_range": age_range,
                "cc_range": cc_range,
                "sc_range": sc_range,
                "is_zd": is_zd,
                "is_ncb": is_ncb,
                "is_cpa": is_cpa,
            },
            "no_filter_selected": True,
        })

    if q:
        group_ids = [int(gid.strip()) for gid in q.split(",") if gid.strip().isdigit()]
        if group_ids:
            qs = qs.filter(Q(group_id__in=group_ids) | Q(group_id__isnull=True, id__in=group_ids))
        else:
            qs = qs.none()

    if status_filter:
        qs = qs.filter(status=status_filter)

    if is_deleted_filter:
        qs = qs.filter(is_deleted=is_deleted_filter)
    if created_date:
        qs = qs.filter(created_at__date=created_date)

    if date_range:
        dates = date_range.split(" - ")
        if len(dates) == 2:
            d_from = dates[0].strip()
            d_to = dates[1].strip()
            if d_from and d_to:
                qs = qs.filter(
                    (Q(from_date__lte=d_to) | Q(from_date__isnull=True)) &
                    (Q(to_date__gte=d_from) | Q(to_date__isnull=True))
                )

    if insurance_company:
        qs = qs.filter(insurance_company=insurance_company)
    if product:
        qs = qs.filter(product_id=product)
    if fuel:
        qs = qs.filter(fuel_type_id=fuel)
    if sub_product:
        qs = qs.filter(sub_product_id=sub_product)

    matching_rto_names = []
    if rto_code:
        for rto_record in RTOMaster.objects.filter(rto_cluster__icontains=rto_code):
            if strict_match_in_cluster(rto_code, rto_record.rto_cluster):
                matching_rto_names.append(rto_record.rto_name.strip().upper())
        if matching_rto_names:
            q_rto = Q()
            for rto_name in matching_rto_names:
                q_rto |= Q(new_rto_list__icontains=rto_name)
            qs = qs.filter(q_rto)
        else:
            qs = qs.none()

    matching_make_names = []
    if make_model_code:
        for make_record in MakeModelMaster.objects.filter(make_model_cluster__icontains=make_model_code):
            if strict_match_in_cluster(make_model_code, make_record.make_model_cluster):
                matching_make_names.append(make_record.make_model_name.strip().upper())
        if matching_make_names:
            q_make = Q()
            for make_name in matching_make_names:
                q_make |= Q(new_vehicle_makes__icontains=make_name)
            qs = qs.filter(q_make)
        else:
            qs = qs.none()

    qs = apply_make_model_filter(qs, product, make_model_class)
    qs = apply_range_filter(qs, "vehicle_age_min", "vehicle_age_max", age_range)
    qs = apply_range_filter(qs, "cc_min", "cc_max", cc_range)
    qs = apply_range_filter(qs, "sc_min", "sc_max", sc_range)

    if is_zd:
        qs = qs.filter(is_zd__code=is_zd)
    if is_ncb:
        qs = qs.filter(is_ncb__code=is_ncb)
    if is_cpa:
        qs = qs.filter(is_cpa__code=is_cpa)

    active_count = qs.filter(status="ACTIVE").count()
    inactive_count = qs.filter(status="INACTIVE").count()

    qs = qs.order_by("-id")

    matching_gids_set = set()
    ordered_gids = []

    for row in qs.iterator(chunk_size=2000):
        if rto_code and matching_rto_names:
            if not row.new_rto_list:
                continue
            row_rtos = [x.strip().upper() for x in row.new_rto_list.split(",")]
            if not any(x in matching_rto_names for x in row_rtos):
                continue

        if make_model_code and matching_make_names:
            if not row.new_vehicle_makes:
                continue
            row_makes = [x.strip().upper() for x in row.new_vehicle_makes.split(",")]
            if not any(x in matching_make_names for x in row_makes):
                continue

        gid = row.group_id if row.group_id is not None else row.id
        if gid not in matching_gids_set:
            matching_gids_set.add(gid)
            ordered_gids.append(gid)

    paginator = Paginator(ordered_gids, DASHBOARD_BATCH_SIZE)
    try:
        page_number = int(request.GET.get("page") or 1)
    except ValueError:
        page_number = 1
    page_obj = paginator.get_page(page_number)
    page_gids = list(page_obj.object_list)
    elided_page_range = list(paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1))

    buckets = defaultdict(list)
    if page_gids:
        full_group_qs = RateMaster.objects.select_related(
            "group", "product", "sub_product", "policy_type", "fuel_type",
            "make_model_class", "is_ncb", "is_cpa", "is_zd"
        ).filter(Q(group_id__in=page_gids) | Q(id__in=page_gids))

        for row in full_group_qs.iterator(chunk_size=2000):
            gid = row.group_id if row.group_id is not None else row.id
            buckets[gid].append(row)

    grouped_rows = []
    for gid in page_gids:
        rows = buckets.get(gid, [])
        if not rows:
            continue

        first = rows[0]
        all_rtos = split_csv_values([r.new_rto_list for r in rows])
        all_fuels = [r.fuel_type.name for r in rows if r.fuel_type]

        first.display_group_id = gid
        first.display_rto_list = unique_join(all_rtos)
        first.display_fuel_types = unique_join(all_fuels)

        mmc_name = first.make_model_class.name.strip().upper() if first.make_model_class else "NA"
        
        if mmc_name in ["NA", ""]:
            prod_name = first.product.name if first.product else ""
            translated_name = get_translated_make_model(prod_name)
            if translated_name:
                first.make_model_class = MakeModelClassMaster(name=translated_name)
                first.display_make_model_class = translated_name
            else:
                first.display_make_model_class = "NA"
        else:
            first.display_make_model_class = first.make_model_class.name

        age_min_str = f"{first.vehicle_age_min:.2f}" if first.vehicle_age_min is not None else ""
        age_max_str = f"{first.vehicle_age_max:.2f}" if first.vehicle_age_max is not None else ""
        first.age_range = f"{age_min_str} - {age_max_str}" if (age_min_str or age_max_str) else ""

        cc_min_str = f"{first.cc_min:.2f}" if first.cc_min is not None else ""
        cc_max_str = f"{first.cc_max:.2f}" if first.cc_max is not None else ""
        first.cc_range = f"{cc_min_str} - {cc_max_str}" if (cc_min_str or cc_max_str) else ""

        sc_min_str = f"{first.sc_min:.2f}" if first.sc_min is not None else ""
        sc_max_str = f"{first.sc_max:.2f}" if first.sc_max is not None else ""
        first.sc_range = f"{sc_min_str} - {sc_max_str}" if (sc_min_str or sc_max_str) else ""

        grouped_rows.append(first)

    field_names = [
        "display_group_id", "new_vehicle_makes", "display_rto_list",
        "insurer_vertical", "insurance_company", "product", "sub_product", "policy_type",
        "display_fuel_types", "display_make_model_class", "age_range",
        "pi_od_rate", "pi_tp_rate", "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5",
        "pi_net_rate", "pi_flat_amount", "pi_vli", "pi_type", "tariff_min", "tariff_max",
        "is_ncb", "is_cpa", "is_zd", "cc_range", "from_date", "to_date", "sc_range",
        "user_id", "veh_use", "remarks", "add_tnc",
        "po_type", "po_od_rate", "po_tp_rate", "po_net_rate", "po_flat_amount",
        "status", "is_deleted"
    ]

    insurance_company_list = RateMaster.objects.exclude(insurance_company="").values_list(
        "insurance_company", flat=True
    ).distinct().order_by("insurance_company")
    product_list = ProductMaster.objects.all().order_by("name")
    fuel_list = FuelTypeMaster.objects.all().order_by("name")
    sub_product_list = SubProductMaster.objects.all().order_by("name")
    yes_no_na_list = YesNoNAMaster.objects.all().order_by("code")

    return render(request, "dashboard.html", {
        "data": grouped_rows,
        "page_obj": page_obj,
        "elided_page_range": elided_page_range,
        "field_names": field_names,
        "bulk_update_fields": ALLOWED_BULK_UPDATE_FIELDS,
        "total": paginator.count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "insurance_company_list": insurance_company_list,
        "product_list": product_list,
        "fuel_list": fuel_list,
        "sub_product_list": sub_product_list,
        "make_model_class_list": get_dynamic_make_model_class_list(product),
        "yes_no_na_list": yes_no_na_list,
        
        "make_class_mapping_json": json.dumps(NA_MAKE_MODEL_MAP),
        "all_make_classes_json": json.dumps(list(MakeModelClassMaster.objects.exclude(name__iexact="NA").values('id', 'name'))),
        
        "selected": {
            "q": q,
            "status": status_filter,
            "is_deleted": is_deleted_filter,
            "created_date": created_date,
            "insurance_company": insurance_company,
            "product": product,
            "fuel": fuel,
            "sub_product": sub_product,
            "make_model_class": make_model_class,
            "date_range": date_range,
            "rto_code": rto_code,
            "make_model_code": make_model_code,
            "age_range": age_range,
            "cc_range": cc_range,
            "sc_range": sc_range,
            "is_zd": is_zd,
            "is_ncb": is_ncb,
            "is_cpa": is_cpa,
        }
    })

# -------------------------
# Export UNGROUPED to CSV (Perfect Match 1:1 Schema Update)
# -------------------------
def export_rates_xlsx(request):
    qs = RateMaster.objects.select_related(
        "product", "sub_product", "policy_type", "fuel_type", "make_model_class", "is_ncb", "is_cpa", "is_zd"
    ).all()

    q = (request.GET.get("q") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    is_deleted_filter = (request.GET.get("is_deleted") or "").strip().upper()
    created_date = (request.GET.get("created_date") or "").strip()
    date_range = (request.GET.get("date_range") or "").strip()

    insurance_company = (request.GET.get("insurance_company") or "").strip()
    product = (request.GET.get("product") or "").strip()
    fuel = (request.GET.get("fuel") or "").strip()
    sub_product = (request.GET.get("sub_product") or "").strip()
    make_model_class = (request.GET.get("make_model_class") or "").strip()
    rto_code = (request.GET.get("rto_code") or "").strip()
    make_model_code = (request.GET.get("make_model_code") or "").strip()
    age_range = (request.GET.get("age_range") or "").strip()
    cc_range = (request.GET.get("cc_range") or "").strip()
    sc_range = (request.GET.get("sc_range") or "").strip()
    is_zd = (request.GET.get("is_zd") or "").strip()
    is_ncb = (request.GET.get("is_ncb") or "").strip()
    is_cpa = (request.GET.get("is_cpa") or "").strip()

    if q:
        group_ids = [int(gid.strip()) for gid in q.split(",") if gid.strip().isdigit()]
        if group_ids:
            qs = qs.filter(Q(group_id__in=group_ids) | Q(group_id__isnull=True, id__in=group_ids))
        else:
            qs = qs.none()

    if status_filter:
        qs = qs.filter(status=status_filter)
    if is_deleted_filter:
        qs = qs.filter(is_deleted=is_deleted_filter)
    if created_date:
        qs = qs.filter(created_at__date=created_date)

    if date_range:
        dates = date_range.split(" - ")
        if len(dates) == 2:
            d_from = dates[0].strip()
            d_to = dates[1].strip()
            if d_from and d_to:
                qs = qs.filter(
                    (Q(from_date__lte=d_to) | Q(from_date__isnull=True)) &
                    (Q(to_date__gte=d_from) | Q(to_date__isnull=True))
                )

    if insurance_company:
        qs = qs.filter(insurance_company=insurance_company)
    if product:
        qs = qs.filter(product_id=product)
    if fuel:
        qs = qs.filter(fuel_type_id=fuel)
    if sub_product:
        qs = qs.filter(sub_product_id=sub_product)

    matching_rto_names = []
    if rto_code:
        for rto_record in RTOMaster.objects.filter(rto_cluster__icontains=rto_code):
            if strict_match_in_cluster(rto_code, rto_record.rto_cluster):
                matching_rto_names.append(rto_record.rto_name.strip().upper())
        if matching_rto_names:
            q_rto = Q()
            for rto_name in matching_rto_names:
                q_rto |= Q(new_rto_list__icontains=rto_name)
            qs = qs.filter(q_rto)
        else:
            qs = qs.none()

    matching_make_names = []
    if make_model_code:
        for make_record in MakeModelMaster.objects.filter(make_model_cluster__icontains=make_model_code):
            if strict_match_in_cluster(make_model_code, make_record.make_model_cluster):
                matching_make_names.append(make_record.make_model_name.strip().upper())
        if matching_make_names:
            q_make = Q()
            for make_name in matching_make_names:
                q_make |= Q(new_vehicle_makes__icontains=make_name)
            qs = qs.filter(q_make)
        else:
            qs = qs.none()

    qs = apply_make_model_filter(qs, product, make_model_class)
    qs = apply_range_filter(qs, "vehicle_age_min", "vehicle_age_max", age_range)
    qs = apply_range_filter(qs, "cc_min", "cc_max", cc_range)
    qs = apply_range_filter(qs, "sc_min", "sc_max", sc_range)

    if is_zd:
        qs = qs.filter(is_zd__code=is_zd)
    if is_ncb:
        qs = qs.filter(is_ncb__code=is_ncb)
    if is_cpa:
        qs = qs.filter(is_cpa__code=is_cpa)

    qs = qs.order_by("-id")

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="rates_export.csv"'

    writer = csv.writer(response)
    
    # Header aligned perfectly with the keys processed during `api_upload_chunk`
    writer.writerow([
        "id", "group_id", "status", "is_deleted", "insurance_company", "insurer_vertical", 
        "product", "sub_product", "policy_type", "fuel_type", "make_model_class",
        "new_vehicle_makes", "new_rto_list", "vehicle_age_min", "vehicle_age_max", 
        "cc_min", "cc_max", "sc_min", "sc_max", "pi_od_rate", "pi_tp_rate", 
        "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5", "pi_net_rate", "pi_flat_amount", 
        "pi_vli", "pi_type", "tariff_min", "tariff_max", "is_ncb", "is_cpa", "is_zd", 
        "from_date", "to_date", "user_id", "veh_use", "add_tnc", "remarks", 
        "po_type", "po_od_rate", "po_tp_rate", "po_net_rate", "po_flat_amount"
    ])

    for r in qs.iterator(chunk_size=2000):
        writer.writerow([
            r.id,
            r.group_id,
            r.status,
            r.is_deleted,
            r.insurance_company,
            r.insurer_vertical,
            r.product.name if r.product else "",
            r.sub_product.name if r.sub_product else "",
            r.policy_type.name if r.policy_type else "",
            r.fuel_type.name if r.fuel_type else "",
            r.make_model_class.name if r.make_model_class else "",
            r.new_vehicle_makes,
            r.new_rto_list,
            r.vehicle_age_min,
            r.vehicle_age_max,
            r.cc_min,
            r.cc_max,
            r.sc_min,
            r.sc_max,
            r.pi_od_rate,
            r.pi_tp_rate,
            r.pi_tp_2,
            r.pi_tp_3,
            r.pi_tp_4,
            r.pi_tp_5,
            r.pi_net_rate,
            r.pi_flat_amount,
            r.pi_vli,
            r.pi_type,
            r.tariff_min,
            r.tariff_max,
            r.is_ncb.code if r.is_ncb else "",
            r.is_cpa.code if r.is_cpa else "",
            r.is_zd.code if r.is_zd else "",
            r.from_date,
            r.to_date,
            r.user_id,
            r.veh_use,
            r.add_tnc,
            r.remarks,
            r.po_type,
            r.po_od_rate,
            r.po_tp_rate,
            r.po_net_rate,
            r.po_flat_amount
        ])

    return response

# -------------------------
# Alias / Field Management Configurator
# -------------------------
def field_configurator(request):
    if request.method == 'POST':
        form = ExtractionFieldForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('field_configurator')
    else:
        form = ExtractionFieldForm()

    fields = ExtractionField.objects.prefetch_related('synonyms').order_by('category', 'order_index')

    return render(request, 'insurance/configurator.html', {
        'fields': fields,
        'form': form
    })

def delete_field(request, pk):
    if request.method == 'POST':
        field = get_object_or_404(ExtractionField, pk=pk)
        field.delete()
    return redirect('field_configurator')

# -------------------------
# Edit Field (Configurations & Synonyms)
# -------------------------
def edit_field(request, pk):
    field_obj = get_object_or_404(ExtractionField, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'update_field':
            form = ExtractionFieldForm(request.POST, instance=field_obj)
            if form.is_valid():
                form.save()
                return redirect('edit_field', pk=pk)
        
        elif action == 'add_synonym':
            synonym_text = request.POST.get('synonym_text', '').strip()
            if synonym_text:
                FieldSynonym.objects.get_or_create(extraction_field=field_obj, synonym_text=synonym_text)
            return redirect('edit_field', pk=pk)
            
        elif action == 'delete_synonym':
            syn_id = request.POST.get('synonym_id')
            FieldSynonym.objects.filter(id=syn_id).delete()
            return redirect('edit_field', pk=pk)
            
    else:
        form = ExtractionFieldForm(instance=field_obj)
        
    return render(request, 'insurance/edit_field.html', {
        'form': form,
        'field_obj': field_obj
    })

# -------------------------
# Upload & Extract PDF / Image
# -------------------------
def upload_extract_pdf(request):
    msg = ""
    error = ""
    upload_form = PolicyDocumentUploadForm()

    if request.method == "POST":
        upload_form = PolicyDocumentUploadForm(request.POST, request.FILES)
        if upload_form.is_valid():
            uploaded_file = request.FILES['uploaded_file']
            
            if PolicyDocumentUpload.objects.filter(original_filename=uploaded_file.name).exists():
                error = f"⚠️ Duplicate File: A document named '{uploaded_file.name}' has already been processed."
            else:
                try:
                    doc_obj = upload_form.save(commit=False)
                    doc_obj.original_filename = doc_obj.uploaded_file.name
                    doc_obj.uploaded_by = request.user
                    doc_obj.mime_type = getattr(doc_obj.uploaded_file, "content_type", "") or ""
                    doc_obj.status = PolicyDocumentUpload.STATUS_PENDING
                    doc_obj.save()

                    process_policy_document_task.delay(doc_obj.id)
                    return redirect("upload_extract_pdf")

                except Exception as e:
                    error = f"⚠️ {str(e)}"
        else:
            error = "⚠️ Please correct the upload form."

    documents = PolicyDocumentUpload.objects.order_by("-created_at")[:20]
    latest_document = documents[0] if documents else None

    if latest_document and latest_document.status == PolicyDocumentUpload.STATUS_COMPLETED:
        msg = f"✅ Last processed document: {latest_document.original_filename}"

    return render(request, "upload_extract_pdf.html", {
        "upload_form": upload_form,
        "documents": documents,
        "latest_document": latest_document,
        "msg": msg,
        "error": error,
    })

# -------------------------
# My MIS
# -------------------------
def my_mis(request):
    qs = PolicyMISRecord.objects.select_related("source_document").order_by("-created_at")

    policy_number = (request.GET.get("policy_number") or "").strip()
    insured_name = (request.GET.get("insured_name") or "").strip()
    insurer_name = (request.GET.get("insurer_name") or "").strip()

    if policy_number:
        qs = qs.filter(policy_number__icontains=policy_number)
    if insured_name:
        qs = qs.filter(insured_name__icontains=insured_name)
    if insurer_name:
        qs = qs.filter(insurer_name__icontains=insurer_name)

    stats = qs.aggregate(
        total_processed=Count('id'),
        total_premium=Sum('gross_premium')
    )

    return render(request, "my_mis.html", {
        "records": qs[:200],
        "stats": stats,
        "selected": {
            "policy_number": policy_number,
            "insured_name": insured_name,
            "insurer_name": insurer_name,
        }
    })

# -------------------------
# User Management
# -------------------------
# Editable Teams-hierarchy fields on UserProfile (excludes user/contact_number/designation,
# which already have dedicated inputs in the edit modal).
HIERARCHY_FIELDS = [
    "vertical_path", "team", "team_id", "user_type", "emp_id", "code", "role",
    "branch_code", "branch_name",
    "rm_code", "rm_name", "tc_code", "tc_name", "csc_code", "csc_name",
    "posp_code", "posp_name",
    "agent_type",
    "pan", "bank_account", "bank_name", "ifsc",
    "membership_id", "qc_verticals",
    # NOTE: user_id_code is deliberately excluded — it's system-assigned
    # (UserProfile.save() auto-fills it) and unique, not admin-editable.
]
HIERARCHY_FLAGS = ["is_qc", "is_plvc", "personal_qc"]
HIERARCHY_FIELD_LABELS = [
    ("vertical_path", "Vertical Path"), ("team", "Team"), ("team_id", "Team ID"),
    ("user_type", "User Type"), ("emp_id", "Emp ID"), ("code", "Code"), ("role", "Role"),
    ("branch_code", "Branch Code"), ("branch_name", "Branch Name"),
    ("rm_code", "RM Code"), ("rm_name", "RM Name"),
    ("tc_code", "TC Code"), ("tc_name", "TC Name"),
    ("csc_code", "CSC Code"), ("csc_name", "CSC Name"),
    ("posp_code", "Ref/POSP Code"), ("posp_name", "Ref/POSP Name"),
    ("agent_type", "Agent Type"),
    ("pan", "PAN"), ("bank_account", "Bank Account"), ("bank_name", "Bank Name"), ("ifsc", "IFSC"),
    ("membership_id", "Membership ID"), ("qc_verticals", "QC Verticals"),
]


def user_management(request):
    Group.objects.get_or_create(name="ADMIN")
    # Deliberately not offered as a checkbox below -- grant it via Django's
    # own /admin/ site. See super_admin_required in urls.py.
    Group.objects.get_or_create(name="SUPER_ADMIN")
    for group_name in PAGE_GROUPS:
        Group.objects.get_or_create(name=group_name)

    msg = ""
    error = ""

    if request.method == "POST":
        action = request.POST.get("action")
        user_id = request.POST.get("user_id")

        if action == "create_user":
            uname = request.POST.get("new_username", "").strip()
            fname = request.POST.get("full_name", "").strip()
            uemail = request.POST.get("email", "").strip()
            ucontact = request.POST.get("contact_number", "").strip()
            upass = secrets.token_urlsafe(9)  # random one-time password, not a shared constant

            if uname:
                if User.objects.filter(username=uname).exists():
                    error = f"⚠️ User '{uname}' already exists."
                elif uemail and User.objects.filter(email__iexact=uemail).exists():
                    error = f"⚠️ An account with email '{uemail}' already exists."
                else:
                    new_user = User.objects.create_user(username=uname, email=uemail, password=upass)
                    if fname:
                        parts = fname.split(" ", 1)
                        new_user.first_name = parts[0]
                        if len(parts) > 1:
                            new_user.last_name = parts[1]
                    new_user.save()

                    UserProfile.objects.get_or_create(user=new_user, defaults={"contact_number": ucontact})
                    msg = f"✅ User '{uname}' created successfully."

        elif action == "update_user_access":
            u = User.objects.get(id=user_id)
            u.groups.remove(Group.objects.get(name="ADMIN"))
            for pg in PAGE_GROUPS:
                u.groups.remove(Group.objects.get(name=pg))

            selected_pages = set(request.POST.getlist(f"pages_{user_id}"))
            if "Rate_Checker" in selected_pages:
                selected_pages.discard("Rate_Checker")
                selected_pages |= RATE_CHECKER_GROUPS
            for pg in selected_pages:
                u.groups.add(Group.objects.get(name=pg))

            profile, _ = UserProfile.objects.get_or_create(user=u)
            profile.contact_number = request.POST.get("contact_number", "")
            profile.designation = request.POST.get("designation", "")

            for field in HIERARCHY_FIELDS:
                profile.__dict__[field] = request.POST.get(field, "").strip() or None
            for flag in HIERARCHY_FLAGS:
                setattr(profile, flag, request.POST.get(flag) == "on")

            reports_to_id = request.POST.get("reports_to") or None
            profile.reports_to_id = reports_to_id if reports_to_id else None

            new_email = request.POST.get("email", "").strip()
            if new_email and User.objects.filter(email__iexact=new_email).exclude(id=u.id).exists():
                error = f"⚠️ An account with email '{new_email}' already exists."
            else:
                try:
                    profile.full_clean(exclude=["user"])
                    profile.save()
                    u.email = new_email
                    u.save()
                    msg = f"✅ Profile and access for '{u.username}' updated."
                except forms.ValidationError as e:
                    error = "⚠️ " + " ".join(m for msgs in e.message_dict.values() for m in msgs)

        elif action == "approve_user" and user_id:
            u = User.objects.get(id=user_id)
            u.is_active = True
            u.save()
            msg = f"✅ '{u.username}' approved — they can now log in."

        elif action == "reject_user" and user_id:
            u = User.objects.get(id=user_id)
            uname = u.username
            u.delete()
            msg = f"🗑️ Signup request from '{uname}' rejected and removed."

        elif action == "reset_password" and user_id:
            u = User.objects.get(id=user_id)
            default_password = secrets.token_urlsafe(9)  # random one-time password, not a shared constant
            u.set_password(default_password)
            u.save()
            logger.info("Password reset by admin for user '%s'", u.username)
            msg = f"✅ Password for '{u.username}' reset to '{default_password}'. Share it with them directly."

        elif action == "make_admin" and user_id:
            u = User.objects.get(id=user_id)
            u.groups.clear()
            u.groups.add(Group.objects.get(name="ADMIN"))
            msg = f"✅ User '{u.username}' promoted to Full Admin."

        elif action == "delete_user" and user_id:
            u = User.objects.get(id=user_id)
            if u.is_superuser:
                error = "⚠️ You cannot delete a Super Admin account."
            else:
                uname = u.username
                u.delete()
                msg = f"🗑️ User '{uname}' has been completely removed from the system."

    all_users = User.objects.select_related("profile").all().order_by("-is_superuser", "username")
    users = [u for u in all_users if u.is_active]
    pending_users = [u for u in all_users if not u.is_active]

    def profile_dict(u):
        profile = getattr(u, "profile", None)
        d = {
            "id": u.id,
            "username": u.username,
            "full_name": u.get_full_name(),
            "email": u.email,
            "contact_number": profile.contact_number if profile else "",
            "designation": profile.designation if profile else "",
            "is_superuser": u.is_superuser,
            "is_admin": u.groups.filter(name="ADMIN").exists(),
            # Filtered to PAGE_GROUPS -- otherwise this leaks any group the
            # user happens to belong to, including ones that no longer mean
            # anything (like the retired Can_Manage_Life_Payout_Grid) or
            # aren't meant to show here at all (ADMIN, SUPER_ADMIN).
            "pages": list(u.groups.filter(name__in=PAGE_GROUPS).values_list("name", flat=True)),
            "user_id_code": profile.user_id_code if profile else "",
        }
        for field in HIERARCHY_FIELDS:
            d[field] = getattr(profile, field, "") or ""
        for flag in HIERARCHY_FLAGS:
            d[flag] = bool(getattr(profile, flag, False))
        d["reports_to"] = profile.reports_to_id if profile else None
        return d

    user_rows = [profile_dict(u) for u in users]
    pending_rows = [
        {
            "id": u.id,
            "username": u.username,
            "full_name": u.get_full_name(),
            "email": u.email,
            "contact_number": u.profile.contact_number if hasattr(u, "profile") else "",
            "designation": u.profile.designation if hasattr(u, "profile") else "",
            "date_joined": u.date_joined,
        }
        for u in pending_users
    ]

    manager_choices = [{"id": u.id, "username": u.username} for u in users]

    return render(request, "user_management.html", {
        "users": user_rows,
        "pending_users": pending_rows,
        "manager_choices": manager_choices,
        "page_groups": PAGE_GROUPS,
        "hierarchy_field_labels": HIERARCHY_FIELD_LABELS,
        "msg": msg,
        "error": error
    })

# -------------------------
# RTO DASHBOARD
# -------------------------
def rto_dashboard(request):
    qs = RTOMaster.objects.all().order_by("rto_name")

    rto_names = request.GET.getlist("rto_name")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if rto_names and "" not in rto_names:
        qs = qs.filter(rto_name__in=rto_names)
    if cluster_q:
        qs = qs.filter(rto_cluster__icontains=cluster_q)

    rto_name_list = RTOMaster.objects.values_list("rto_name", flat=True).distinct().order_by("rto_name")

    return render(request, "rto_dashboard.html", {
        "data": qs,
        "total": qs.count(),
        "rto_name_list": rto_name_list,
        "selected": {
            "rto_names": rto_names,
            "cluster_q": cluster_q
        },
        "is_admin": True
    })

def export_rto_xlsx(request):
    qs = RTOMaster.objects.all().order_by("rto_name")

    rto_names = request.GET.getlist("rto_name")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if rto_names and "" not in rto_names:
        qs = qs.filter(rto_name__in=rto_names)
    if cluster_q:
        qs = qs.filter(rto_cluster__icontains=cluster_q)

    wb = Workbook()
    ws = wb.active
    ws.title = "RTO Master"
    ws.append(["ID", "RTO NAME", "RTO CLUSTER"])

    for r in qs:
        ws.append([r.id, r.rto_name, r.rto_cluster or ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="rto_master.xlsx"'
    wb.save(response)
    return response

# -------------------------
# MAKE MODEL DASHBOARD
# -------------------------
def make_model_dashboard(request):
    qs = MakeModelMaster.objects.all().order_by("make_model_name")

    make_model_names = request.GET.getlist("make_model_name")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if make_model_names and "" not in make_model_names:
        qs = qs.filter(make_model_name__in=make_model_names)
    if cluster_q:
        qs = qs.filter(make_model_cluster__icontains=cluster_q)

    make_model_name_list = MakeModelMaster.objects.values_list(
        "make_model_name", flat=True
    ).distinct().order_by("make_model_name")

    return render(request, "make_model_dashboard.html", {
        "data": qs,
        "total": qs.count(),
        "make_model_name_list": make_model_name_list,
        "selected": {
            "make_model_names": make_model_names,
            "cluster_q": cluster_q
        },
        "is_admin": True
    })

def export_make_model_xlsx(request):
    qs = MakeModelMaster.objects.all().order_by("make_model_name")

    make_model_names = request.GET.getlist("make_model_name")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if make_model_names and "" not in make_model_names:
        qs = qs.filter(make_model_name__in=make_model_names)
    if cluster_q:
        qs = qs.filter(make_model_cluster__icontains=cluster_q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Make Model Master"
    ws.append(["ID", "MAKE MODEL NAME", "MAKE MODEL CLUSTER"])

    for r in qs:
        ws.append([r.id, r.make_model_name, r.make_model_cluster or ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="make_model_master.xlsx"'
    wb.save(response)
    return response

# -------------------------
# PINCODE DASHBOARD (Health's equivalent of RTO Dashboard)
# -------------------------
def _sync_pincode_zones_from_health_data():
    """Ensures every pincode_zone value actually used in HealthRateMaster has
    a PincodeMaster row (blank cluster if new), so the dashboard always shows
    every zone that needs its real pincode list filled in — rather than
    silently missing one after a Health grid re-import introduces a new zone."""
    existing = set(PincodeMaster.objects.values_list("pincode_zone", flat=True))
    used = set(
        HealthRateMaster.objects.exclude(pincode_zone__isnull=True)
        .exclude(pincode_zone="")
        .values_list("pincode_zone", flat=True)
        .distinct()
    )
    for zone in used - existing:
        PincodeMaster.objects.get_or_create(pincode_zone=zone)


def pincode_dashboard(request):
    _sync_pincode_zones_from_health_data()

    qs = PincodeMaster.objects.all().order_by("pincode_zone")

    zone_names = request.GET.getlist("pincode_zone")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if zone_names and "" not in zone_names:
        qs = qs.filter(pincode_zone__in=zone_names)
    if cluster_q:
        qs = qs.filter(pincode_cluster__icontains=cluster_q)

    zone_name_list = PincodeMaster.objects.values_list("pincode_zone", flat=True).distinct().order_by("pincode_zone")

    return render(request, "pincode_dashboard.html", {
        "data": qs,
        "total": qs.count(),
        "zone_name_list": zone_name_list,
        "selected": {
            "zone_names": zone_names,
            "cluster_q": cluster_q
        },
        "is_admin": True
    })

def export_pincode_xlsx(request):
    _sync_pincode_zones_from_health_data()

    qs = PincodeMaster.objects.all().order_by("pincode_zone")

    zone_names = request.GET.getlist("pincode_zone")
    cluster_q = (request.GET.get("cluster_q") or "").strip()

    if zone_names and "" not in zone_names:
        qs = qs.filter(pincode_zone__in=zone_names)
    if cluster_q:
        qs = qs.filter(pincode_cluster__icontains=cluster_q)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pincode Master"
    # Header row matches PincodeMaster field names — lets this export be
    # edited and re-uploaded via Import Data's "Pincode Master (Clusters CSV)"
    # option without renaming any columns.
    ws.append(["id", "pincode_zone", "pincode_cluster"])

    for r in qs:
        ws.append([r.id, r.pincode_zone, r.pincode_cluster or ""])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="pincode_master.xlsx"'
    wb.save(response)
    return response

# -------------------------
# Edit Master Tables
# -------------------------
def edit_rto(request, pk):
    obj = RTOMaster.objects.get(id=pk)
    if request.method == "POST":
        form = RTOForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("rto_dashboard")
    else:
        form = RTOForm(instance=obj)

    return render(request, "edit_master.html", {
        "form": form,
        "title": "Edit RTO Record",
        "back_url": "rto_dashboard"
    })

def edit_make_model(request, pk):
    obj = MakeModelMaster.objects.get(id=pk)
    if request.method == "POST":
        form = MakeModelForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("make_model_dashboard")
    else:
        form = MakeModelForm(instance=obj)

    return render(request, "edit_master.html", {
        "form": form,
        "title": "Edit Make/Model Record",
        "back_url": "make_model_dashboard"
    })

def edit_pincode(request, pk):
    obj = PincodeMaster.objects.get(id=pk)
    if request.method == "POST":
        form = PincodeForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("pincode_dashboard")
    else:
        form = PincodeForm(instance=obj)

    return render(request, "edit_master.html", {
        "form": form,
        "title": "Edit Pincode Zone Record",
        "back_url": "pincode_dashboard"
    })

# -------------------------
# Edit Rate Form
# -------------------------
def edit_rate(request, group_id):
    # group_id here is overloaded: it's either a real RateGroup id (edit the
    # whole group), or — when a caller couldn't resolve a group (e.g. some
    # "Fix Mapping" links on Rate Master Health) — a raw RateMaster.id for a
    # single ungrouped row. These two id spaces are independent auto-increment
    # sequences, so they collide: as of this writing 1,898 RateGroup ids also
    # happen to equal the primary key of some unrelated RateMaster row
    # belonging to a *different* group. Q(group_id=X) | Q(id=X) would match
    # BOTH in that case and silently fold a stranger row into this edit/bulk
    # update. Only fall back to the raw-id interpretation when no real group
    # matches, so the two interpretations can never mix.
    records = RateMaster.objects.filter(group_id=group_id)
    if not records.exists():
        records = RateMaster.objects.filter(id=group_id)
    if not records.exists():
        return HttpResponse("Record not found.", status=404)

    first_record = records.first()
    record_count = records.count()

    all_rtos = split_csv_values([r.new_rto_list for r in records])
    unique_rto_list = list(set([r for r in all_rtos if r]))
    rto_display = unique_join(all_rtos)

    all_makes = split_csv_values([r.new_vehicle_makes for r in records])
    unique_makes_list = list(set([m for m in all_makes if m]))

    initial_data = {
        "new_rto_list": unique_rto_list,
        "new_vehicle_makes": unique_makes_list,
        "insurance_company": first_record.insurance_company,
        "insurer_vertical": first_record.insurer_vertical,
        "status": first_record.status,
        "is_deleted": first_record.is_deleted,
        "product": first_record.product_id,
        "sub_product": first_record.sub_product_id,
        "policy_type": first_record.policy_type_id,
        "fuel_type": first_record.fuel_type_id,
        "make_model_class": first_record.make_model_class_id,
        "is_ncb": first_record.is_ncb_id,
        "is_cpa": first_record.is_cpa_id,
        "is_zd": first_record.is_zd_id,
    }

    if request.method == "POST":
        # initial_data must be passed here too, not just on GET — RateForm.__init__
        # builds the new_rto_list/new_vehicle_makes MultipleChoiceField choices from
        # kwargs["initial"] plus the single `instance` record. Without it, POST-time
        # choices only reflect first_record's own values, so submitting the full
        # group-wide selection the page actually shows fails validation on every
        # value that isn't also on first_record specifically — silently discarding
        # the whole edit.
        form = RateForm(request.POST, instance=first_record, initial=initial_data)
        if form.is_valid():
            # Only write fields the user actually changed from what the form
            # showed them — not the full cleaned_data for every field on the
            # model. This matters most for new_rto_list/new_vehicle_makes:
            # their *displayed* value is the union of every record's own
            # value across the whole group (see unique_rto_list/unique_makes_list
            # above), which is correct for showing "what this group currently
            # covers" but is NOT a value that's safe to blast onto every
            # record — new_rto_list in particular is deliberately excluded
            # from GROUP_FIELDS because individual records in a group
            # routinely carry different single RTOs. Writing the full
            # cleaned_data unconditionally (the old behavior) meant editing
            # any single field — e.g. just a rate — silently overwrote every
            # record's own RTO/vehicle-makes with that group-wide union,
            # destroying the per-record distinction. form.changed_data is
            # Django's built-in "differs from what was shown" check, so a
            # field is only written when the user genuinely touched it.
            update_data = {field: form.cleaned_data[field] for field in form.changed_data}
            if update_data:
                records.update(**update_data)
                AuditLog.objects.create(
                    user=request.user,
                    action="MANUAL EDIT",
                    details=(
                        f"Edited Group/Record ID {group_id} via form. Updated {record_count} rows. "
                        f"Changed fields: {', '.join(sorted(update_data.keys()))}."
                    )
                )
            return redirect("dashboard")
    else:
        form = RateForm(instance=first_record, initial=initial_data)

    return render(request, "edit_rate.html", {
        "form": form,
        "count": records.count(),
        "group_id": group_id,
        "rto_display": rto_display
    })

# -------------------------
# BULK UPDATE RATES
# -------------------------
# Only these RateMaster fields may be touched via bulk update — update_field
# comes straight from the POST body, so anything not on this list (id,
# group_id, created_at, ...) must be rejected rather than passed through to
# records.update(**{field_name: ...}). Ordered (not a set) because this same
# tuple also drives the "Field to Change" dropdown on the dashboard, so its
# order is the order options appear in — keep the two in sync by construction
# rather than maintaining a second field list in the view/template.
ALLOWED_BULK_UPDATE_FIELDS = (
    "new_vehicle_makes", "product", "sub_product", "policy_type", "fuel_type",
    "make_model_class", "is_ncb", "is_cpa", "is_zd", "status", "is_deleted",
    "vehicle_age_min", "vehicle_age_max", "cc_min", "cc_max", "user_id",
    "pi_od_rate", "pi_tp_rate", "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5",
    "pi_net_rate", "pi_flat_amount", "pi_vli", "tariff_min", "tariff_max",
    "sc_min", "sc_max", "from_date", "to_date",
    "po_od_rate", "po_tp_rate", "po_net_rate", "po_flat_amount",
    "insurance_company", "insurer_vertical", "pi_type", "po_type", "veh_use",
    "add_tnc", "remarks",
)


def bulk_update_rates(request):
    if request.method == "POST":
        # 1. Safely extract group IDs regardless of how JS structured the payload
        raw_groups = request.POST.getlist("selected_groups")
        if not raw_groups:
            raw_groups = request.POST.getlist("selected_groups[]")
        if not raw_groups:
            raw_groups = [request.POST.get("selected_groups", "")]

        group_ids = []
        for g in raw_groups:
            for part in str(g).split(","):
                part = part.strip()
                if part.isdigit():
                    group_ids.append(int(part))

        field_name = request.POST.get("update_field")
        new_value = request.POST.get("update_value", "").strip()

        if not group_ids or not field_name:
            return redirect("dashboard")

        if field_name not in ALLOWED_BULK_UPDATE_FIELDS:
            messages.error(
                request,
                f"Bulk update rejected: '{field_name}' is not an editable field. No rows were changed."
            )
            return redirect("dashboard")

        # 2. Fetch all exact records mapped to the selected rows.
        # Same group_id/id ambiguity as edit_rate (see its comment): resolve
        # each selected id against group_id first, and only treat it as a raw
        # RateMaster.id if no group has that id, so a group id that happens to
        # collide with an unrelated row's primary key can't pull that row in.
        ids_matching_a_group = set(
            RateMaster.objects.filter(group_id__in=group_ids).values_list("group_id", flat=True).distinct()
        )
        leftover_ids = [gid for gid in group_ids if gid not in ids_matching_a_group]
        records = RateMaster.objects.filter(Q(group_id__in=group_ids) | Q(id__in=leftover_ids))
        record_count = records.count()

        # 3. Safely cast the incoming text value into the correct Python/Django Database Type
        parsed_value = new_value
        if field_name == "product":
            parsed_value, _ = ProductMaster.objects.get_or_create(name=new_value) if new_value else (None, False)
        elif field_name == "sub_product":
            parsed_value, _ = SubProductMaster.objects.get_or_create(name=new_value) if new_value else (None, False)
        elif field_name == "policy_type":
            parsed_value, _ = PolicyTypeMaster.objects.get_or_create(name=new_value) if new_value else (None, False)
        elif field_name == "fuel_type":
            parsed_value, _ = FuelTypeMaster.objects.get_or_create(name=new_value) if new_value else (None, False)
        elif field_name == "make_model_class":
            parsed_value, _ = MakeModelClassMaster.objects.get_or_create(name=new_value) if new_value else (None, False)
        elif field_name in ["is_ncb", "is_cpa", "is_zd"]:
            parsed_value = parse_yes_no_na(new_value)
        elif field_name == "status":
            parsed_value = str(new_value).strip().upper()
            valid_statuses = dict(RateMaster.STATUS_CHOICES)
            if parsed_value not in valid_statuses:
                messages.error(
                    request,
                    f"Bulk update rejected: '{new_value}' is not a valid Status "
                    f"(must be one of: {', '.join(valid_statuses)}). No rows were changed."
                )
                return redirect("dashboard")
        elif field_name == "is_deleted":
            parsed_value = str(new_value).strip().upper()
            valid_deleted = dict(RateMaster.IS_DELETED_CHOICES)
            if parsed_value not in valid_deleted:
                messages.error(
                    request,
                    f"Bulk update rejected: '{new_value}' is not a valid Is Deleted value "
                    f"(must be one of: {', '.join(valid_deleted)}). No rows were changed."
                )
                return redirect("dashboard")
        elif field_name == "user_id":
            parsed_value = int(float(new_value)) if new_value else None
        elif field_name in [
            "vehicle_age_min", "vehicle_age_max", "cc_min", "cc_max",
            "pi_od_rate", "pi_tp_rate", "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5",
            "pi_net_rate", "pi_flat_amount", "pi_vli", "tariff_min", "tariff_max",
            "sc_min", "sc_max", "po_od_rate", "po_tp_rate", "po_net_rate", "po_flat_amount"
        ]:
            parsed_value = float(new_value) if new_value else None

        # 4. Perform ultra-fast vectorized update to the DB
        records.update(**{field_name: parsed_value})

        # 5. Log the action
        AuditLog.objects.create(
            user=request.user,
            action="BULK UPDATE",
            details=f"Updated {record_count} rows. Changed '{field_name}' to '{new_value}'."
        )

    return redirect("dashboard")

# -------------------------
# RATE MASTER HEALTH
# -------------------------
# Surfaces individual MIS policy rows the mapping engine
# (mapping_engine.process_mis_mapping) could not resolve to a single payout
# rate — ⚠️ MULTIPLE MATCHES, ❌ NO MATCH, and FAILED - BAD DATA. These are
# stored directly in MISFailedRow (written straight from df_final when a
# file finishes processing — see process_mis_mapping), so this view is a
# normal indexed queryset + select_related + Paginator, not something that
# re-parses any Excel/CSV output on every page view.

MIS_HEALTH_BATCH_SIZE = 50

# Fields shown in the Policy Details column's collapsed summary before "View
# Details" is clicked. payload's keys are whatever column headers the
# uploaded MIS file actually had (see mapping_engine._extract_failed_rows_from_df
# — it keeps the source file's own header text), so these are matched as
# case-insensitive substrings rather than exact keys.
POLICY_DETAILS_SUMMARY_TARGETS = ["vehicle make", "model", "policy type", "sub product"]
POLICY_DETAILS_SUMMARY_LIMIT = 3


def _split_payload_summary(payload):
    """
    Pick up to POLICY_DETAILS_SUMMARY_LIMIT of a MISFailedRow's most
    identifying payload fields for the table row's collapsed view; everything
    else is returned separately for the "View Details" expansion. Falls back
    to the first few payload entries if none of the target fields are present
    (e.g. a Health row with no vehicle columns), so the summary is never blank
    while payload itself has data.
    """
    items = [(k, v) for k, v in (payload or {}).items() if v not in (None, "")]

    summary = []
    used_keys = set()
    for target in POLICY_DETAILS_SUMMARY_TARGETS:
        if len(summary) >= POLICY_DETAILS_SUMMARY_LIMIT:
            break
        for k, v in items:
            if k in used_keys:
                continue
            if target in k.strip().lower():
                summary.append((k, v))
                used_keys.add(k)
                break

    if not summary:
        summary = items[:POLICY_DETAILS_SUMMARY_LIMIT]
        used_keys = {k for k, _ in summary}

    rest = [(k, v) for k, v in items if k not in used_keys]
    return summary, rest


def _sync_unsynced_mis_files():
    """
    Backfill for MISFile rows that completed before MISFailedRow existed (or
    whose direct write at process time failed) — cheap indexed no-op query
    in steady state once every completed file has been synced once.
    """
    unsynced = (
        MISFile.objects.filter(status="COMPLETED", health_synced_at__isnull=True)
        .exclude(processed_file="")
        .exclude(processed_file__isnull=True)
    )
    if not unsynced.exists():
        return
    from .mapping_engine import sync_failed_rows_for_file
    for mis_file in unsynced:
        try:
            sync_failed_rows_for_file(mis_file)
        except Exception:
            logger.exception("Could not sync MIS failure rows for MISFile %s", mis_file.id)


# The full set of PI (premium/tariff) rate fields on RateMaster — same
# grouping used elsewhere for PI-side rate handling (see RATE_FIELDS et al).
# pi_type says which of these should carry the rate; the rest must be 0/blank.
PI_RATE_FIELDS = [
    "pi_od_rate", "pi_tp_rate", "pi_tp_2", "pi_tp_3", "pi_tp_4", "pi_tp_5",
    "pi_net_rate", "pi_flat_amount", "pi_vli",
]
PI_TYPE_RULES = [
    ("On Net", ["pi_net_rate"]),
    ("On OD", ["pi_od_rate"]),
    ("On TP", ["pi_tp_rate"]),
    ("On OD and TP", ["pi_od_rate", "pi_tp_rate"]),
]
PI_TYPE_RULES_MAP = dict(PI_TYPE_RULES)


def _rate_master_pi_type_violations_qs(label):
    """
    RateMaster rows whose pi_type is `label` but that have a nonzero value in
    a PI rate field other than the one(s) that pi_type says should hold the
    rate. NULL/0 fields are "unset" and never violate — only a genuinely
    populated field outside the allowed set counts as an error.
    """
    allowed_fields = PI_TYPE_RULES_MAP.get(label)
    if allowed_fields is None:
        return RateMaster.objects.none()
    other_fields_q = Q()
    for field in PI_RATE_FIELDS:
        if field in allowed_fields:
            continue
        other_fields_q |= ~Q(**{field: 0})
    return RateMaster.objects.filter(is_deleted="NO", pi_type=label).filter(other_fields_q)


def _rate_master_pi_type_error_counts():
    return [
        {"label": label, "count": _rate_master_pi_type_violations_qs(label).count()}
        for label, _ in PI_TYPE_RULES
    ]


# Range-validity rules — a RateMaster row is only valid when each "min" bound
# is strictly less than its paired "max" bound. A bound that's NULL can't be
# compared, so — same "unset never violates" convention as the pi_type rules
# above — a row is only flagged once both bounds in the pair are populated.

# STP (Standard Third Party) tariff_max is statutorily fixed rather than a
# free-form upper bound, so its range rule is checked against this constant
# *in addition to* the general tariff_min < tariff_max rule, not instead of it.
STP_SUB_PRODUCT = "STP"
STP_TARIFF_MAX = 100.01


def _rate_master_min_max_violations_qs(min_field, max_field):
    """RateMaster rows where min_field/max_field are both set but min isn't strictly less than max."""
    lookup = {
        f"{min_field}__isnull": False,
        f"{max_field}__isnull": False,
        f"{min_field}__gte": F(max_field),
    }
    return RateMaster.objects.filter(is_deleted="NO", **lookup)


def _rate_master_tariff_violations_qs():
    """
    RateMaster rows that break the tariff_min < tariff_max rule, or — for STP
    sub-products — don't have tariff_max pinned to exactly STP_TARIFF_MAX.
    """
    general_bad = Q(tariff_min__isnull=False, tariff_max__isnull=False, tariff_min__gte=F("tariff_max"))
    stp_bad = (
        Q(sub_product__name__iexact=STP_SUB_PRODUCT, tariff_max__isnull=False)
        & ~Q(tariff_max=STP_TARIFF_MAX)
    )
    return RateMaster.objects.filter(is_deleted="NO").filter(general_bad | stp_bad)


# One entry per Range Validation Errors card. "columns" are the extra
# (field_path, header) pairs the Affected Rows drill-down shows beyond the
# common Group ID / Insurer columns, so the user can see exactly which values
# broke the rule — field_path is anything .values() accepts, including a
# relation lookup like sub_product__name.
RANGE_ERROR_RULES = [
    {
        "label": "Invalid Vehicle Age Range",
        "violations_fn": lambda: _rate_master_min_max_violations_qs("vehicle_age_min", "vehicle_age_max"),
        "columns": [("vehicle_age_min", "Vehicle Age Min"), ("vehicle_age_max", "Vehicle Age Max")],
    },
    {
        "label": "Invalid CC Range",
        "violations_fn": lambda: _rate_master_min_max_violations_qs("cc_min", "cc_max"),
        "columns": [("cc_min", "CC Min"), ("cc_max", "CC Max")],
    },
    {
        "label": "Invalid Seating Capacity Range",
        "violations_fn": lambda: _rate_master_min_max_violations_qs("sc_min", "sc_max"),
        "columns": [("sc_min", "SC Min"), ("sc_max", "SC Max")],
    },
    {
        "label": "Invalid Tariff Range",
        "violations_fn": _rate_master_tariff_violations_qs,
        "columns": [
            ("sub_product__name", "Sub Product"),
            ("tariff_min", "Tariff Min"),
            ("tariff_max", "Tariff Max"),
        ],
    },
]
RANGE_ERROR_RULES_MAP = {rule["label"]: rule for rule in RANGE_ERROR_RULES}


def _rate_master_range_error_counts():
    return [{"label": rule["label"], "count": rule["violations_fn"]().count()} for rule in RANGE_ERROR_RULES]


def _rate_master_equality_violations_qs(min_field, max_field):
    """RateMaster rows where min_field/max_field are both set and exactly equal."""
    lookup = {
        f"{min_field}__isnull": False,
        f"{max_field}__isnull": False,
        min_field: F(max_field),
    }
    return RateMaster.objects.filter(is_deleted="NO", **lookup)


# One entry per Equality Errors card — same shape as RANGE_ERROR_RULES, but
# flagging the narrower "min was typed in as exactly equal to max" mistake
# rather than "min isn't strictly less than max". Every row these catch is
# therefore already counted by the matching Range Validation Errors card too
# (equality is a subset of "not strictly less than") — this is a deliberately
# more specific lens on the same data, not a separate class of bad row.
EQUALITY_ERROR_RULES = [
    {
        "label": "Vehicle Age Equality",
        "violations_fn": lambda: _rate_master_equality_violations_qs("vehicle_age_min", "vehicle_age_max"),
        "columns": [("vehicle_age_min", "Vehicle Age Min"), ("vehicle_age_max", "Vehicle Age Max")],
    },
    {
        "label": "CC Equality",
        "violations_fn": lambda: _rate_master_equality_violations_qs("cc_min", "cc_max"),
        "columns": [("cc_min", "CC Min"), ("cc_max", "CC Max")],
    },
    {
        "label": "SC Equality",
        "violations_fn": lambda: _rate_master_equality_violations_qs("sc_min", "sc_max"),
        "columns": [("sc_min", "SC Min"), ("sc_max", "SC Max")],
    },
    {
        "label": "Tariff Equality",
        "violations_fn": lambda: _rate_master_equality_violations_qs("tariff_min", "tariff_max"),
        "columns": [("tariff_min", "Tariff Min"), ("tariff_max", "Tariff Max")],
    },
    {
        "label": "Validity Date Equality",
        "violations_fn": lambda: _rate_master_equality_violations_qs("from_date", "to_date"),
        "columns": [("from_date", "From Date"), ("to_date", "To Date")],
    },
]
EQUALITY_ERROR_RULES_MAP = {rule["label"]: rule for rule in EQUALITY_ERROR_RULES}


def _rate_master_equality_error_counts():
    return [{"label": rule["label"], "count": rule["violations_fn"]().count()} for rule in EQUALITY_ERROR_RULES]


def _rate_master_grid_summary_qs():
    """
    Pivot of uploaded RateMaster grids across the same six fields the
    dashboard uses to describe a grid: product, sub_product,
    insurance_company, from_date, to_date, status. grid_count is a DISTINCT
    count of group_id (falling back to id for the rare row with no group
    assigned), not a row count, so one grid that exploded into hundreds of
    make/RTO/CC rate lines still only counts once.
    """
    return (
        RateMaster.objects
        .annotate(grid_key=Coalesce("group_id", "id"))
        .values("product__name", "sub_product__name", "insurance_company", "from_date", "to_date", "status")
        .annotate(grid_count=Count("grid_key", distinct=True))
        .order_by("-grid_count", "insurance_company", "product__name", "sub_product__name")
    )


def rate_master_health(request):
    _sync_unsynced_mis_files()

    status_key = (request.GET.get("failure_reason") or "").strip()
    insurer = (request.GET.get("insurer") or "").strip()
    date_from = (request.GET.get("date_from") or "").strip()
    date_to = (request.GET.get("date_to") or "").strip()

    qs = MISFailedRow.objects.select_related("mis_file").order_by("-mis_file__created_at", "-id")

    if status_key:
        qs = qs.filter(status_key=status_key)
    if insurer:
        qs = qs.filter(insurer=insurer)
    if date_from:
        try:
            from_date = datetime.strptime(date_from, "%Y-%m-%d").date()
            qs = qs.filter(mis_file__created_at__date__gte=from_date)
        except ValueError:
            pass
    if date_to:
        try:
            to_date = datetime.strptime(date_to, "%Y-%m-%d").date()
            qs = qs.filter(mis_file__created_at__date__lte=to_date)
        except ValueError:
            pass

    total_all = MISFailedRow.objects.count()
    status_counts = dict(
        MISFailedRow.objects.values("status_key").annotate(n=Count("id")).values_list("status_key", "n")
    )

    paginator = Paginator(qs, MIS_HEALTH_BATCH_SIZE)
    try:
        page_number = int(request.GET.get("page") or 1)
    except ValueError:
        page_number = 1
    page_obj = paginator.get_page(page_number)
    elided_page_range = list(paginator.get_elided_page_range(page_obj.number, on_each_side=1, on_ends=1))

    # select_related already joined mis_file — this only touches the current
    # page's rows (<= MIS_HEALTH_BATCH_SIZE), not the full filtered set.
    for r in page_obj:
        r.file_name = (
            r.mis_file.uploaded_file.name.rsplit("/", 1)[-1]
            if r.mis_file.uploaded_file else f"File #{r.mis_file_id}"
        )
        r.payload_summary, r.payload_rest = _split_payload_summary(r.payload)

    insurer_list = list(
        MISFailedRow.objects.exclude(insurer__isnull=True).exclude(insurer="")
        .order_by().values_list("insurer", flat=True).distinct()
    )
    insurer_list.sort()

    status_types = [c for c in MISFailedRow.STATUS_CHOICES if c[0] != "OTHER"]

    pi_type_error_counts = _rate_master_pi_type_error_counts()
    range_error_counts = _rate_master_range_error_counts()
    equality_error_counts = _rate_master_equality_error_counts()

    # Grid Summary pivot — materialized once so the paginator doesn't re-run
    # the aggregate query, same pattern dashboard() uses for ordered_gids.
    grid_summary_rows = list(_rate_master_grid_summary_qs())
    grid_summary_total_combinations = len(grid_summary_rows)
    grid_summary_total_grids = sum(r["grid_count"] for r in grid_summary_rows)

    grid_summary_paginator = Paginator(grid_summary_rows, MIS_HEALTH_BATCH_SIZE)
    try:
        grid_summary_page_number = int(request.GET.get("grid_page") or 1)
    except ValueError:
        grid_summary_page_number = 1
    grid_summary_page_obj = grid_summary_paginator.get_page(grid_summary_page_number)
    grid_summary_elided_range = list(
        grid_summary_paginator.get_elided_page_range(grid_summary_page_obj.number, on_each_side=1, on_ends=1)
    )

    # Drill-down: clicking an Error Dashboard card lists the distinct Rate
    # Master groups behind that rule's violating rows, so the user can jump
    # straight to Fix Mapping instead of hunting through the full table.
    pi_type_error = (request.GET.get("pi_type_error") or "").strip()
    error_group_page_obj = None
    error_group_elided_range = None
    error_group_ungrouped_count = 0
    if pi_type_error in PI_TYPE_RULES_MAP:
        violations_qs = _rate_master_pi_type_violations_qs(pi_type_error)
        error_group_ungrouped_count = violations_qs.filter(group_id__isnull=True).count()
        grouped_qs = (
            violations_qs.exclude(group_id__isnull=True)
            .values("group_id", "insurance_company")
            .annotate(row_count=Count("id"))
            .order_by("-row_count", "group_id")
        )
        error_paginator = Paginator(grouped_qs, MIS_HEALTH_BATCH_SIZE)
        try:
            error_page_number = int(request.GET.get("err_page") or 1)
        except ValueError:
            error_page_number = 1
        error_group_page_obj = error_paginator.get_page(error_page_number)
        error_group_elided_range = list(
            error_paginator.get_elided_page_range(error_group_page_obj.number, on_each_side=1, on_ends=1)
        )

    # Drill-down: clicking a Range Validation Errors card lists the individual
    # violating rows (not grouped counts, unlike the pi_type drill-down above)
    # so the user can see the actual min/max values that broke the rule.
    range_error = (request.GET.get("range_error") or "").strip()
    selected_range_rule = RANGE_ERROR_RULES_MAP.get(range_error)
    range_rows_page_obj = None
    range_rows_elided_range = None
    if selected_range_rule:
        value_fields = ["id", "group_id", "insurance_company"] + [f for f, _ in selected_range_rule["columns"]]
        range_rows_qs = selected_range_rule["violations_fn"]().values(*value_fields).order_by("-id")
        range_paginator = Paginator(range_rows_qs, MIS_HEALTH_BATCH_SIZE)
        try:
            range_page_number = int(request.GET.get("range_page") or 1)
        except ValueError:
            range_page_number = 1
        range_rows_page_obj = range_paginator.get_page(range_page_number)
        range_rows_elided_range = list(
            range_paginator.get_elided_page_range(range_rows_page_obj.number, on_each_side=1, on_ends=1)
        )

    # Drill-down: clicking an Equality Errors card lists the individual
    # violating rows, the same row-level shape as the Range Validation
    # drill-down above — equality is just a narrower rule, checked separately.
    equality_error = (request.GET.get("equality_error") or "").strip()
    selected_equality_rule = EQUALITY_ERROR_RULES_MAP.get(equality_error)
    equality_rows_page_obj = None
    equality_rows_elided_range = None
    if selected_equality_rule:
        value_fields = ["id", "group_id", "insurance_company"] + [f for f, _ in selected_equality_rule["columns"]]
        equality_rows_qs = selected_equality_rule["violations_fn"]().values(*value_fields).order_by("-id")
        equality_paginator = Paginator(equality_rows_qs, MIS_HEALTH_BATCH_SIZE)
        try:
            equality_page_number = int(request.GET.get("equality_page") or 1)
        except ValueError:
            equality_page_number = 1
        equality_rows_page_obj = equality_paginator.get_page(equality_page_number)
        equality_rows_elided_range = list(
            equality_paginator.get_elided_page_range(equality_rows_page_obj.number, on_each_side=1, on_ends=1)
        )

    return render(request, "rate_master_health.html", {
        "page_obj": page_obj,
        "elided_page_range": elided_page_range,
        "total": paginator.count,
        "total_all": total_all,
        "status_types": status_types,
        "status_counts": status_counts,
        "insurer_list": insurer_list,
        "pi_type_error_counts": pi_type_error_counts,
        "pi_type_error_total": sum(r["count"] for r in pi_type_error_counts),
        "selected_pi_type_error": pi_type_error if pi_type_error in PI_TYPE_RULES_MAP else "",
        "range_error_counts": range_error_counts,
        "range_error_total": sum(r["count"] for r in range_error_counts),
        "selected_range_error": range_error if selected_range_rule else "",
        "selected_range_rule": selected_range_rule,
        "range_rows_page_obj": range_rows_page_obj,
        "range_rows_elided_range": range_rows_elided_range,
        "equality_error_counts": equality_error_counts,
        "equality_error_total": sum(r["count"] for r in equality_error_counts),
        "selected_equality_error": equality_error if selected_equality_rule else "",
        "selected_equality_rule": selected_equality_rule,
        "equality_rows_page_obj": equality_rows_page_obj,
        "equality_rows_elided_range": equality_rows_elided_range,
        "error_group_page_obj": error_group_page_obj,
        "error_group_elided_range": error_group_elided_range,
        "error_group_ungrouped_count": error_group_ungrouped_count,
        "grid_summary_page_obj": grid_summary_page_obj,
        "grid_summary_elided_range": grid_summary_elided_range,
        "grid_summary_total_combinations": grid_summary_total_combinations,
        "grid_summary_total_grids": grid_summary_total_grids,
        "selected": {
            "failure_reason": status_key,
            "insurer": insurer,
            "date_from": date_from,
            "date_to": date_to,
        },
    })


def export_grid_summary_xlsx(request):
    """
    Same pivot as the Grid Summary tab above -- always the full aggregate
    (all combinations), regardless of which page the on-screen table happens
    to be showing, since grid_page only controls that table's pagination.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Grid Summary"
    ws.append(["product", "sub_product", "insurance_company", "from_date", "to_date", "status", "grid_count"])

    for r in _rate_master_grid_summary_qs():
        ws.append([
            r["product__name"] or "",
            r["sub_product__name"] or "",
            r["insurance_company"],
            r["from_date"].strftime("%Y-%m-%d") if r["from_date"] else "",
            r["to_date"].strftime("%Y-%m-%d") if r["to_date"] else "",
            r["status"],
            r["grid_count"],
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="grid_summary_export.xlsx"'
    wb.save(response)
    return response

# -------------------------
# HEALTH RATE MASTER
# -------------------------
# HealthRateMaster rows aren't exploded/grouped behind a separate group
# table like RateMaster is — one imported grid row is one editable rate
# rule, so list/edit/bulk-update all operate directly on HealthRateMaster ids.
HEALTH_RATE_BATCH_SIZE = 50

HEALTH_RATE_FIELD_NAMES = [
    "insurance_company", "product_name", "policy_category", "business_type",
    "plan_names", "min_sum_insured", "max_sum_insured", "min_deductible", "max_deductible",
    "min_age", "max_age", "pincode_zone", "payin_rate",
    "one_year_rate", "multi_year_2_rate", "multi_year_3_rate", "multi_year_4_rate", "multi_year_5_rate",
    "from_date", "to_date", "remarks", "status", "is_deleted",
]

# Only these HealthRateMaster fields may be touched via bulk update — same
# safety rule as RateMaster's ALLOWED_BULK_UPDATE_FIELDS: update_field comes
# straight from the POST body, so anything not on this list (id, source_row_hash,
# created_at, updated_at) must be rejected rather than passed to .update().
ALLOWED_HEALTH_BULK_UPDATE_FIELDS = {
    "insurance_company", "product_name", "policy_category", "business_type", "plan_names",
    "min_deductible", "max_deductible", "min_sum_insured", "max_sum_insured",
    "min_age", "max_age", "pincode_zone",
    "payin_rate", "one_year_rate", "multi_year_2_rate", "multi_year_3_rate",
    "multi_year_4_rate", "multi_year_5_rate",
    "from_date", "to_date", "status", "is_deleted", "remarks",
}
HEALTH_FLOAT_FIELDS = {
    "min_deductible", "max_deductible", "min_sum_insured", "max_sum_insured",
    "min_age", "max_age", "payin_rate", "one_year_rate",
    "multi_year_2_rate", "multi_year_3_rate", "multi_year_4_rate", "multi_year_5_rate",
}


class HealthRateForm(forms.ModelForm):
    class Meta:
        model = HealthRateMaster
        exclude = ["created_at", "updated_at", "source_row_hash"]
        widgets = {
            "plan_names": forms.Textarea(attrs={"rows": 4}),
            "remarks": forms.Textarea(attrs={"rows": 3}),
            "from_date": forms.DateInput(attrs={"type": "date"}),
            "to_date": forms.DateInput(attrs={"type": "date"}),
        }


def _build_health_rate_queryset(request):
    """Filter-building logic shared by health_rate_master (list) and export_health_rates_xlsx."""
    qs = HealthRateMaster.objects.all()

    q = (request.GET.get("q") or "").strip()
    insurance_company = (request.GET.get("insurance_company") or "").strip()
    product_name = (request.GET.get("product_name") or "").strip()
    policy_category = (request.GET.get("policy_category") or "").strip()
    business_type = (request.GET.get("business_type") or "").strip()
    pincode_zone = (request.GET.get("pincode_zone") or "").strip()
    pincode = (request.GET.get("pincode") or "").strip()
    status_filter = (request.GET.get("status") or "").strip()
    is_deleted_filter = (request.GET.get("is_deleted") or "").strip().upper()
    date_range = (request.GET.get("date_range") or "").strip()
    sum_insured_range = (request.GET.get("sum_insured_range") or "").strip()
    age_range = (request.GET.get("age_range") or "").strip()
    deductible_range = (request.GET.get("deductible_range") or "").strip()

    if q:
        # Supports plain text search (insurer/plan name) plus an ID lookup --
        # the table displays each row's id as "h-<id>" (see health_rate_master.html),
        # so accept that same "h-" prefix here (and a comma-separated list of
        # them), alongside a bare numeric id. Only switches into ID-list mode
        # when EVERY comma-separated token looks like an id -- otherwise q may
        # be a genuine text search that happens to contain a comma, so it
        # falls through to the plain insurer/plan-name match untouched.
        id_tokens = []
        for part in q.split(","):
            part = part.strip()
            if not part:
                continue
            if part[:2].lower() == "h-":
                part = part[2:]
            if part.isdigit():
                id_tokens.append(int(part))
            else:
                id_tokens = None
                break

        text_match = Q(insurance_company__icontains=q) | Q(plan_names__icontains=q)
        if id_tokens:
            text_match |= Q(pk__in=id_tokens)
        qs = qs.filter(text_match)
    if insurance_company:
        qs = qs.filter(insurance_company=insurance_company)
    if product_name:
        qs = qs.filter(product_name=product_name)
    if policy_category:
        qs = qs.filter(policy_category=policy_category)
    if business_type:
        qs = qs.filter(business_type=business_type)
    if pincode_zone:
        qs = qs.filter(pincode_zone=pincode_zone)

    # Raw pincode search: resolve through PincodeMaster's cluster to the
    # zone(s) it belongs to, same pattern as rto_code resolving through
    # RTOMaster.rto_cluster to a group name on Motor's dashboard(). A pincode
    # with no known zone mapping yet matches nothing — it means the zone's
    # real pincode list hasn't been filled in on the Pincode Dashboard.
    if pincode:
        matching_zone_names = []
        potential_zones = PincodeMaster.objects.filter(pincode_cluster__icontains=pincode)
        for zone_record in potential_zones:
            if strict_match_in_cluster(pincode, zone_record.pincode_cluster):
                matching_zone_names.append(zone_record.pincode_zone.strip())

        if matching_zone_names:
            qs = qs.filter(pincode_zone__in=matching_zone_names)
        else:
            qs = qs.none()

    if status_filter:
        qs = qs.filter(status=status_filter)
    if is_deleted_filter:
        qs = qs.filter(is_deleted=is_deleted_filter)

    if date_range:
        dates = date_range.split(" - ")
        if len(dates) == 2:
            d_from = dates[0].strip()
            d_to = dates[1].strip()
            if d_from and d_to:
                qs = qs.filter(
                    (Q(from_date__lte=d_to) | Q(from_date__isnull=True)) &
                    (Q(to_date__gte=d_from) | Q(to_date__isnull=True))
                )

    qs = apply_range_filter(qs, "min_sum_insured", "max_sum_insured", sum_insured_range)
    qs = apply_range_filter(qs, "min_age", "max_age", age_range)
    qs = apply_range_filter(qs, "min_deductible", "max_deductible", deductible_range)

    selected = {
        "q": q,
        "insurance_company": insurance_company,
        "product_name": product_name,
        "policy_category": policy_category,
        "business_type": business_type,
        "pincode_zone": pincode_zone,
        "pincode": pincode,
        "status": status_filter,
        "is_deleted": is_deleted_filter,
        "date_range": date_range,
        "sum_insured_range": sum_insured_range,
        "age_range": age_range,
        "deductible_range": deductible_range,
    }
    return qs, selected


def health_rate_master(request):
    qs, selected = _build_health_rate_queryset(request)

    active_count = qs.filter(status="ACTIVE").count()
    inactive_count = qs.filter(status="INACTIVE").count()
    total_count = qs.count()

    qs = qs.order_by("-id")

    try:
        page_number = int(request.GET.get("page") or 1)
    except ValueError:
        page_number = 1
    paginator = Paginator(qs, HEALTH_RATE_BATCH_SIZE)
    page_obj = paginator.get_page(page_number)

    insurer_list = HealthRateMaster.objects.exclude(insurance_company="").values_list(
        "insurance_company", flat=True
    ).distinct().order_by("insurance_company")
    product_list = HealthRateMaster.objects.exclude(product_name__isnull=True).exclude(
        product_name=""
    ).values_list("product_name", flat=True).distinct().order_by("product_name")
    category_list = HealthRateMaster.objects.exclude(policy_category__isnull=True).exclude(
        policy_category=""
    ).values_list("policy_category", flat=True).distinct().order_by("policy_category")
    business_type_list = HealthRateMaster.objects.exclude(business_type__isnull=True).exclude(
        business_type=""
    ).values_list("business_type", flat=True).distinct().order_by("business_type")
    zone_list = HealthRateMaster.objects.exclude(pincode_zone__isnull=True).exclude(
        pincode_zone=""
    ).values_list("pincode_zone", flat=True).distinct().order_by("pincode_zone")

    return render(request, "health_rate_master.html", {
        "page_obj": page_obj,
        "field_names": HEALTH_RATE_FIELD_NAMES,
        "total": total_count,
        "active_count": active_count,
        "inactive_count": inactive_count,
        "insurer_list": insurer_list,
        "product_list": product_list,
        "category_list": category_list,
        "business_type_list": business_type_list,
        "zone_list": zone_list,
        "selected": selected,
    })


def health_rate_master_edit(request, pk):
    record = get_object_or_404(HealthRateMaster, pk=pk)

    if request.method == "POST":
        form = HealthRateForm(request.POST, instance=record)
        if form.is_valid():
            form.save()
            AuditLog.objects.create(
                user=request.user,
                action="HEALTH RATE EDIT",
                details=f"Edited HealthRateMaster #{record.pk} via form."
            )
            return redirect("health_rate_master")
    else:
        form = HealthRateForm(instance=record)

    return render(request, "edit_health_rate.html", {
        "form": form,
        "record": record,
    })


def health_rate_master_bulk_update(request):
    if request.method == "POST":
        raw_ids = request.POST.getlist("selected_rows")
        if not raw_ids:
            raw_ids = [request.POST.get("selected_rows", "")]

        ids = []
        for raw in raw_ids:
            for part in str(raw).split(","):
                part = part.strip()
                if part.isdigit():
                    ids.append(int(part))

        field_name = request.POST.get("update_field")
        new_value = request.POST.get("update_value", "").strip()

        if not ids or not field_name:
            return redirect("health_rate_master")

        if field_name not in ALLOWED_HEALTH_BULK_UPDATE_FIELDS:
            messages.error(
                request,
                f"Bulk update rejected: '{field_name}' is not an editable field. No rows were changed."
            )
            return redirect("health_rate_master")

        records = HealthRateMaster.objects.filter(id__in=ids)
        record_count = records.count()

        parsed_value = new_value
        if field_name == "status":
            parsed_value = new_value.strip().upper()
            valid_statuses = dict(HealthRateMaster.STATUS_CHOICES)
            if parsed_value not in valid_statuses:
                messages.error(
                    request,
                    f"Bulk update rejected: '{new_value}' is not a valid Status "
                    f"(must be one of: {', '.join(valid_statuses)}). No rows were changed."
                )
                return redirect("health_rate_master")
        elif field_name == "is_deleted":
            parsed_value = new_value.strip().upper()
            valid_deleted = dict(HealthRateMaster.IS_DELETED_CHOICES)
            if parsed_value not in valid_deleted:
                messages.error(
                    request,
                    f"Bulk update rejected: '{new_value}' is not a valid Is Deleted value "
                    f"(must be one of: {', '.join(valid_deleted)}). No rows were changed."
                )
                return redirect("health_rate_master")
        elif field_name in HEALTH_FLOAT_FIELDS:
            try:
                parsed_value = float(new_value) if new_value else None
            except ValueError:
                messages.error(
                    request,
                    f"Bulk update rejected: '{new_value}' is not a valid number for '{field_name}'. No rows were changed."
                )
                return redirect("health_rate_master")

        records.update(**{field_name: parsed_value})

        AuditLog.objects.create(
            user=request.user,
            action="HEALTH BULK UPDATE",
            details=f"Updated {record_count} Health rate rows. Changed '{field_name}' to '{new_value}'."
        )

    return redirect("health_rate_master")


def export_health_rates_xlsx(request):
    qs, _ = _build_health_rate_queryset(request)
    qs = qs.order_by("-id")

    wb = Workbook()
    ws = wb.active
    ws.title = "Health Rate Master"

    # Header row matches HealthRateMaster field names exactly (not friendly
    # Title Case) — same convention as Motor's export_rates_xlsx, and it means
    # this export can be edited and re-uploaded via Import Data's "Health
    # Rate Master (Rates CSV)" option without renaming any columns.
    ws.append([
        "id", "insurance_company", "product_name", "policy_category", "business_type",
        "plan_names", "min_deductible", "max_deductible", "min_sum_insured", "max_sum_insured",
        "min_age", "max_age", "pincode_zone", "payin_rate", "one_year_rate",
        "multi_year_2_rate", "multi_year_3_rate", "multi_year_4_rate", "multi_year_5_rate",
        "from_date", "to_date", "status", "is_deleted", "remarks",
    ])

    for r in qs.iterator(chunk_size=2000):
        ws.append([
            r.id, r.insurance_company, r.product_name, r.policy_category, r.business_type,
            r.plan_names, r.min_deductible, r.max_deductible, r.min_sum_insured, r.max_sum_insured,
            r.min_age, r.max_age, r.pincode_zone, r.payin_rate, r.one_year_rate,
            r.multi_year_2_rate, r.multi_year_3_rate, r.multi_year_4_rate, r.multi_year_5_rate,
            r.from_date.strftime("%Y-%m-%d") if r.from_date else "",
            r.to_date.strftime("%Y-%m-%d") if r.to_date else "",
            r.status, r.is_deleted, r.remarks or "",
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="health_rates_export.xlsx"'
    wb.save(response)
    return response

# -------------------------
# HEALTH PAYOUT RATES CHECKER (Health's equivalent of Motor's Quote Simulator)
# -------------------------
HEALTH_PAYOUT_MAX_RESULTS = 200

# Standard Sum Insured ladder (in rupees) offered on the "Insurance Cover"
# dropdown. These are conventional Indian health-insurance cover tiers, not
# values read from the data — a customer picks from a known ladder, and the
# query checks which rate rules' [min_sum_insured, max_sum_insured] band
# contains that amount.
HEALTH_SUM_INSURED_OPTIONS = [
    (500000, "5 Lakhs"), (750000, "7.5 Lakhs"), (1000000, "10 Lakhs"),
    (1500000, "15 Lakhs"), (2000000, "20 Lakhs"), (2500000, "25 Lakhs"),
    ("above_25L", "Above 25 Lakhs"),
]
HEALTH_ABOVE_25L_THRESHOLD = 2500000

# Which HealthRateMaster rate column the "N Yr Policy" radio choice reads.
HEALTH_POLICY_TERM_FIELD = {
    "1": "one_year_rate",
    "2": "multi_year_2_rate",
    "3": "multi_year_3_rate",
    "4": "multi_year_4_rate",
    "5": "multi_year_5_rate",
}


def _build_health_payout_queryset(request, target_date, rate_field):
    """Eligibility-style matching (containment on age/sum-insured/date,
    same as Motor's cc/sc/vehicle-age matching in _build_motor_payout_queryset)
    — deliberately different from Health Rate Master's admin filters, which
    match a rule's own band exactly rather than checking whether it covers
    a given customer's numbers."""
    qs = HealthRateMaster.objects.exclude(is_deleted="YES").filter(status="ACTIVE")

    product_name = (request.GET.get("product_name") or "").strip()
    business_type = (request.GET.get("business_type") or "").strip()
    policy_category = (request.GET.get("policy_category") or "").strip()
    sum_insured = (request.GET.get("sum_insured") or "").strip()
    age = (request.GET.get("age") or "").strip()
    pincode = (request.GET.get("pincode") or "").strip()
    insurance_companies = [c for c in request.GET.getlist("insurance_company") if c]

    if target_date:
        qs = qs.filter(
            (Q(from_date__lte=target_date) | Q(from_date__isnull=True)) &
            (Q(to_date__gte=target_date) | Q(to_date__isnull=True))
        )

    if product_name:
        qs = qs.filter(product_name=product_name)
    if business_type:
        qs = qs.filter(business_type=business_type)
    if policy_category:
        qs = qs.filter(policy_category=policy_category)
    if insurance_companies:
        qs = qs.filter(insurance_company__in=insurance_companies)

    if sum_insured == "above_25L":
        # Open-ended option — not a single point to contain, but "does this
        # rule's cover extend past the threshold at all" (a range-overlap
        # check, not point-containment like the fixed Lakh amounts below).
        qs = qs.filter(
            Q(max_sum_insured__gt=HEALTH_ABOVE_25L_THRESHOLD) | Q(max_sum_insured__isnull=True)
        )
    elif sum_insured:
        try:
            si_val = float(sum_insured)
            qs = qs.filter(
                (Q(min_sum_insured__lte=si_val) | Q(min_sum_insured__isnull=True)) &
                (Q(max_sum_insured__gte=si_val) | Q(max_sum_insured__isnull=True))
            )
        except ValueError:
            pass

    if age:
        try:
            age_val = float(age)
            qs = qs.filter(
                (Q(min_age__lte=age_val) | Q(min_age__isnull=True)) &
                (Q(max_age__gte=age_val) | Q(max_age__isnull=True))
            )
        except ValueError:
            pass

    if pincode:
        matching_zone_names = []
        potential_zones = PincodeMaster.objects.filter(pincode_cluster__icontains=pincode)
        for zone_record in potential_zones:
            if strict_match_in_cluster(pincode, zone_record.pincode_cluster):
                matching_zone_names.append(zone_record.pincode_zone.strip())
        if matching_zone_names:
            qs = qs.filter(pincode_zone__in=matching_zone_names)
        else:
            # Lenient fallback, same pattern as rto_code in Motor's quote
            # checker: maybe the user typed the zone name (or "All") directly
            # instead of an actual pincode, or that zone's cluster hasn't been
            # populated yet on the Pincode Dashboard.
            qs = qs.filter(pincode_zone__icontains=pincode)

    qs = qs.order_by(F(rate_field).desc(nulls_last=True), F("payin_rate").desc(nulls_last=True), "-id")

    return qs


def _run_health_payout_search(request):
    """
    Core eligibility search shared by the health_payout_rates HTML page
    and HealthPayoutRatesAPIView. Returns (results, total_found, has_more,
    has_searched, selected).
    """
    has_searched = bool(request.GET)

    if has_searched:
        _log_health_points_search(request)

    today_str = datetime.today().strftime("%Y-%m-%d")
    target_date = (request.GET.get("target_date") or today_str).strip()
    policy_term = (request.GET.get("policy_term") or "1").strip()
    if policy_term not in HEALTH_POLICY_TERM_FIELD:
        policy_term = "1"
    rate_field = HEALTH_POLICY_TERM_FIELD[policy_term]

    results = []
    total_found = 0
    has_more = False

    if has_searched:
        qs = _build_health_payout_queryset(request, target_date, rate_field)
        total_found = qs.count()
        has_more = total_found > HEALTH_PAYOUT_MAX_RESULTS

        for row in qs[:HEALTH_PAYOUT_MAX_RESULTS]:
            term_rate = getattr(row, rate_field)
            row.applicable_rate = term_rate if term_rate is not None else row.payin_rate
            results.append(row)

    selected = {
        "product_name": (request.GET.get("product_name") or "").strip(),
        "business_type": (request.GET.get("business_type") or "").strip(),
        "policy_category": (request.GET.get("policy_category") or "").strip(),
        "sum_insured": (request.GET.get("sum_insured") or "").strip(),
        "age": (request.GET.get("age") or "").strip(),
        "pincode": (request.GET.get("pincode") or "").strip(),
        "insurance_companies": [c for c in request.GET.getlist("insurance_company") if c],
        "target_date": target_date,
        "policy_term": policy_term,
    }

    return results, total_found, has_more, has_searched, selected


def health_payout_rates(request):
    results, total_found, has_more, has_searched, selected = _run_health_payout_search(request)

    product_list = HealthRateMaster.objects.exclude(product_name__isnull=True).exclude(
        product_name=""
    ).values_list("product_name", flat=True).distinct().order_by("product_name")
    business_type_list = HealthRateMaster.objects.exclude(business_type__isnull=True).exclude(
        business_type=""
    ).values_list("business_type", flat=True).distinct().order_by("business_type")
    category_list = HealthRateMaster.objects.exclude(policy_category__isnull=True).exclude(
        policy_category=""
    ).values_list("policy_category", flat=True).distinct().order_by("policy_category")
    insurer_list = HealthRateMaster.objects.exclude(insurance_company="").values_list(
        "insurance_company", flat=True
    ).distinct().order_by("insurance_company")

    return render(request, "health_payout_rates.html", {
        "has_searched": has_searched,
        "data": results,
        "total_found": total_found,
        "has_more": has_more,
        "max_results": HEALTH_PAYOUT_MAX_RESULTS,
        "product_list": product_list,
        "business_type_list": business_type_list,
        "category_list": category_list,
        "insurer_list": insurer_list,
        "sum_insured_options": HEALTH_SUM_INSURED_OPTIONS,
        "selected": selected,
    })

# -------------------------
# MOTOR PAYOUT RATES
# -------------------------
MOTOR_PAYOUT_BATCH_SIZE = 50
MOTOR_PAYOUT_MAX_RESULTS = 300
MOTOR_PAYOUT_FIELD_NAMES = [
    "display_group_id", "status", "insurance_company", "tariff_range", "po_type",
    "po_rate", "po_flat_amount", "add_tnc"
]


def _build_motor_payout_queryset(request, target_date):
    """Filter-building logic shared by motor_payout_rates (first batch) and
    motor_payout_rates_more (subsequent batches), so both stay in sync."""
    qs = RateMaster.objects.select_related(
        "product", "sub_product", "policy_type", "fuel_type",
        "make_model_class", "is_ncb", "is_cpa", "is_zd"
    ).all()

    qs = qs.exclude(is_deleted="YES")
    qs = qs.filter(status__in=["ACTIVE", "INACTIVE"])

    product = (request.GET.get("product") or "").strip()
    make_model_class = (request.GET.get("make_model_class") or "").strip()
    sub_product = (request.GET.get("sub_product") or "").strip()
    make_names = (request.GET.get("make_names") or "").strip()
    rto_code = (request.GET.get("rto_code") or "").strip()
    cc = (request.GET.get("cc") or "").strip()
    fuel = (request.GET.get("fuel") or "").strip()
    sc = (request.GET.get("sc") or "").strip()
    mfg_year = (request.GET.get("mfg_year") or "").strip()
    is_zd = (request.GET.get("is_zd") or "").strip().upper()
    is_cpa = (request.GET.get("is_cpa") or "").strip().upper()
    is_ncb = (request.GET.get("is_ncb") or "").strip().upper()

    if target_date:
        qs = qs.filter(
            (Q(from_date__lte=target_date) | Q(from_date__isnull=True)) &
            (Q(to_date__gte=target_date) | Q(to_date__isnull=True))
        )

    if product:
        if str(product).isdigit():
            qs = qs.filter(product_id=product)
        else:
            qs = qs.filter(product__name__iexact=str(product).strip())

    qs = apply_make_model_filter(qs, product, make_model_class)

    if sub_product:
        if str(sub_product).isdigit():
            qs = qs.filter(sub_product_id=sub_product)
        else:
            qs = qs.filter(sub_product__name__iexact=str(sub_product).strip())

    if fuel:
        if str(fuel).isdigit():
            qs = qs.filter(fuel_type_id=fuel)
        else:
            qs = qs.filter(fuel_type__name__iexact=str(fuel).strip())

    matching_rto_names = []
    if rto_code:
        potential_rtos = RTOMaster.objects.filter(rto_cluster__icontains=rto_code)
        for rto_record in potential_rtos:
            if strict_match_in_cluster(rto_code, rto_record.rto_cluster):
                matching_rto_names.append(rto_record.rto_name.strip().upper())
        if matching_rto_names:
            q_rto = Q()
            for rto_name in matching_rto_names:
                q_rto |= Q(new_rto_list__icontains=rto_name)
            qs = qs.filter(q_rto)
        else:
            qs = qs.filter(new_rto_list__icontains=rto_code.strip())

    matching_make_groups = []
    if make_names:
        potential_makes = MakeModelMaster.objects.filter(make_model_cluster__icontains=make_names)
        for make_record in potential_makes:
            if strict_match_in_cluster(make_names, make_record.make_model_cluster):
                matching_make_groups.append(make_record.make_model_name.strip().upper())
        if matching_make_groups:
            q_make = Q()
            for group_name in matching_make_groups:
                if group_name:
                    q_make |= Q(new_vehicle_makes__icontains=group_name.strip())
            qs = qs.filter(q_make)
        else:
            qs = qs.filter(new_vehicle_makes__icontains=make_names.strip())

    if cc and cc.isdigit():
        qs = qs.filter(
            (Q(cc_min__lte=int(cc)) | Q(cc_min__isnull=True)) &
            (Q(cc_max__gte=int(cc)) | Q(cc_max__isnull=True))
        )

    if sc:
        try:
            qs = qs.filter(
                (Q(sc_min__lte=float(sc)) | Q(sc_min__isnull=True)) &
                (Q(sc_max__gte=float(sc)) | Q(sc_max__isnull=True))
            )
        except ValueError:
            pass

    if mfg_year and mfg_year.isdigit() and target_date:
        try:
            target_dt = datetime.strptime(target_date, "%Y-%m-%d")
            mfg_dt = datetime(int(mfg_year), 1, 1)
            exact_age = round((target_dt - mfg_dt).days / 365.25, 2)
            qs = qs.filter(
                (Q(vehicle_age_min__lte=exact_age) | Q(vehicle_age_min__isnull=True)) &
                (Q(vehicle_age_max__gte=exact_age) | Q(vehicle_age_max__isnull=True))
            )
        except ValueError:
            pass

    if is_zd:
        qs = qs.filter(Q(is_zd__code__iexact=is_zd) | Q(is_zd__code__iexact="NA"))
    if is_cpa:
        qs = qs.filter(Q(is_cpa__code__iexact=is_cpa) | Q(is_cpa__code__iexact="NA"))
    if is_ncb:
        qs = qs.filter(Q(is_ncb__code__iexact=is_ncb) | Q(is_ncb__code__iexact="NA"))

    qs = qs.order_by(
        F("po_net_rate").desc(nulls_last=True),
        F("po_od_rate").desc(nulls_last=True),
        F("po_flat_amount").desc(nulls_last=True),
        "-id"
    )

    return qs, matching_rto_names, matching_make_groups, rto_code, make_names


def _collect_motor_payout_rows(qs, matching_rto_names, matching_make_groups, rto_code, make_names, skip, limit):
    """Streams the ordered queryset and applies the same in-Python RTO/make
    cluster matching + group de-duplication as before, returning the slice
    of unique groups from `skip` up to `skip + limit` (never exceeding
    MOTOR_PAYOUT_MAX_RESULTS overall), plus whether any group exists beyond
    that slice."""
    effective_limit = min(limit, MOTOR_PAYOUT_MAX_RESULTS - skip)
    if effective_limit <= 0:
        return [], False

    results = []
    seen_groups = set()
    group_index = 0

    for row in qs.iterator(chunk_size=2000):
        if rto_code and matching_rto_names:
            if not row.new_rto_list:
                continue
            row_rtos = [x.strip().upper() for x in row.new_rto_list.split(",")]
            if not any(x in matching_rto_names for x in row_rtos):
                continue

        if make_names and matching_make_groups:
            if not row.new_vehicle_makes:
                continue
            row_makes = [x.strip().upper() for x in row.new_vehicle_makes.split(",")]
            if not any(x in matching_make_groups for x in row_makes):
                continue

        gid = row.group_id if row.group_id is not None else row.id
        if gid in seen_groups:
            continue
        seen_groups.add(gid)

        if group_index < skip:
            group_index += 1
            continue
        group_index += 1

        if len(results) >= effective_limit:
            # This row is proof at least one more group exists beyond the
            # slice we're returning — no need to keep scanning further.
            return results, True

        row.display_group_id = gid

        mmc_name = row.make_model_class.name.strip().upper() if row.make_model_class else "NA"

        if mmc_name in ["NA", ""]:
            prod_name = row.product.name if row.product else ""
            translated_name = get_translated_make_model(prod_name)
            if translated_name:
                row.make_model_class = MakeModelClassMaster(name=translated_name)
                row.display_make_model_class = translated_name
            else:
                row.display_make_model_class = "NA"
        else:
            row.display_make_model_class = row.make_model_class.name

        if row.po_net_rate and row.po_net_rate > 0:
            row.po_rate = row.po_net_rate
        elif row.po_od_rate and row.po_od_rate > 0:
            row.po_rate = row.po_od_rate
        elif row.po_tp_rate and row.po_tp_rate > 0:
            row.po_rate = row.po_tp_rate
        else:
            row.po_rate = 0.0

        results.append(row)

    return results, False


def motor_payout_rates(request):
    has_searched = bool(request.GET)

    if has_searched:
        _log_motor_points_search(request)

    today_str = datetime.today().strftime("%Y-%m-%d")
    target_date = (request.GET.get("target_date") or today_str).strip()

    product = (request.GET.get("product") or "").strip()
    make_model_class = (request.GET.get("make_model_class") or "").strip()
    sub_product = (request.GET.get("sub_product") or "").strip()
    make_names = (request.GET.get("make_names") or "").strip()
    rto_code = (request.GET.get("rto_code") or "").strip()
    cc = (request.GET.get("cc") or "").strip()
    fuel = (request.GET.get("fuel") or "").strip()
    sc = (request.GET.get("sc") or "").strip()
    mfg_year = (request.GET.get("mfg_year") or "").strip()
    is_zd = (request.GET.get("is_zd") or "").strip().upper()
    is_cpa = (request.GET.get("is_cpa") or "").strip().upper()
    is_ncb = (request.GET.get("is_ncb") or "").strip().upper()

    results = []
    has_more = False

    # Nothing is queried until the user actually submits the search form —
    # a bare GET with no params at all (the very first page load) skips the
    # database entirely.
    if has_searched:
        qs, matching_rto_names, matching_make_groups, rto_code, make_names = _build_motor_payout_queryset(request, target_date)
        results, has_more = _collect_motor_payout_rows(
            qs, matching_rto_names, matching_make_groups, rto_code, make_names,
            skip=0, limit=MOTOR_PAYOUT_BATCH_SIZE
        )

    all_makes_json, class_makes_mapping_json, all_makes = get_make_mapping_context()

    return render(request, "motor_payout_rates.html", {
        "has_searched": has_searched,
        "data": results,
        "total_found": len(results),
        "has_more": has_more,
        "next_offset": len(results),
        "field_names": MOTOR_PAYOUT_FIELD_NAMES,
        "product_list": ProductMaster.objects.all().order_by("name"),
        "sub_product_list": SubProductMaster.objects.all().order_by("name"),
        "fuel_list": FuelTypeMaster.objects.all().order_by("name"),
        "make_model_class_list": get_dynamic_make_model_class_list(product),
        "all_makes_json": all_makes_json,
        "class_makes_mapping_json": class_makes_mapping_json,

        "make_class_mapping_json": json.dumps(NA_MAKE_MODEL_MAP),
        "all_make_classes_json": json.dumps(list(MakeModelClassMaster.objects.exclude(name__iexact="NA").values('id', 'name'))),

        "selected": {
            "target_date": target_date,
            "product": product,
            "make_model_class": make_model_class,
            "sub_product": sub_product,
            "make_names": make_names,
            "rto_code": rto_code,
            "cc": cc,
            "fuel": fuel,
            "sc": sc,
            "mfg_year": mfg_year,
            "is_zd": is_zd,
            "is_cpa": is_cpa,
            "is_ncb": is_ncb,
        }
    })


def motor_payout_rates_more(request):
    """AJAX 'Load More' endpoint for motor_payout_rates — same filters
    (passed again as query params, exactly as the browser already has them
    in the URL), plus a `skip` telling it how many groups have already been
    shown. Returns a JSON payload with a pre-rendered HTML fragment of the
    next rows, rather than the whole page."""
    today_str = datetime.today().strftime("%Y-%m-%d")
    target_date = (request.GET.get("target_date") or today_str).strip()

    try:
        skip = int(request.GET.get("skip", 0))
    except (TypeError, ValueError):
        skip = 0
    skip = max(skip, 0)

    qs, matching_rto_names, matching_make_groups, rto_code, make_names = _build_motor_payout_queryset(request, target_date)
    results, has_more = _collect_motor_payout_rows(
        qs, matching_rto_names, matching_make_groups, rto_code, make_names,
        skip=skip, limit=MOTOR_PAYOUT_BATCH_SIZE
    )

    rows_html = render_to_string(
        "_motor_payout_rates_rows.html",
        {"data": results, "field_names": MOTOR_PAYOUT_FIELD_NAMES},
        request=request,
    )

    return JsonResponse({
        "html": rows_html,
        "has_more": has_more,
        "next_offset": skip + len(results),
    })

def _log_motor_points_search(request):
    """Record a MOTOR_POINTS_SEARCH audit log entry from the current request's
    GET params. Shared by every "Check Eligibility" search page (Rate Checker,
    Motor Payout Rates Checker, ...) so they all feed the same
    /points-logs/ audit trail in the same format."""
    flat_params = {}
    for key in request.GET.keys():
        val = request.GET.get(key, "").strip()
        if val and key != "csrfmiddlewaretoken":
            flat_params[key] = val

    if flat_params:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="MOTOR_POINTS_SEARCH",
            details=str(flat_params)
        )


def _log_health_points_search(request):
    """Record a HEALTH_POINTS_SEARCH audit log entry from the current
    request's GET params, so the Health Points Checker feeds the same
    /points-logs/ audit trail as the Motor search pages. Unlike
    _log_motor_points_search, Insurer (insurance_company) is a multi-select
    field on this form, so it's read with getlist() and joined rather than
    losing every value but the last."""
    flat_params = {}
    for key in request.GET.keys():
        if key == "csrfmiddlewaretoken":
            continue
        values = [v.strip() for v in request.GET.getlist(key) if v and v.strip()]
        if values:
            flat_params[key] = ", ".join(values)

    if flat_params:
        AuditLog.objects.create(
            user=request.user if request.user.is_authenticated else None,
            action="HEALTH_POINTS_SEARCH",
            details=str(flat_params)
        )

# -------------------------
# POLICY LOCK CHECKER
# -------------------------
def _run_policy_lock_checker_search(request):
    """
    Core search/filter/dedup logic shared by the policy_lock_checker HTML
    page and PolicyLockCheckerAPIView. Returns (results, has_searched, selected)
    -- everything the two callers need, minus the HTML-only dropdown lists.
    """
    has_searched = bool(request.GET)

    if has_searched:
        _log_motor_points_search(request)

    qs = RateMaster.objects.select_related(
        "product", "sub_product", "policy_type", "fuel_type",
        "make_model_class", "is_ncb", "is_cpa", "is_zd"
    ).all()

    qs = qs.exclude(is_deleted="YES")
    qs = qs.filter(status__in=["ACTIVE", "INACTIVE"])

    today_str = datetime.today().strftime("%Y-%m-%d")
    target_date = (request.GET.get("target_date") or today_str).strip()

    vehicle_no = (request.GET.get("vehicle_no") or "").strip()
    policy_holder_name = (request.GET.get("policy_holder_name") or "").strip()
    product = (request.GET.get("product") or "").strip()
    make_model_class = (request.GET.get("make_model_class") or "").strip()
    sub_product = (request.GET.get("sub_product") or "").strip()
    make_names = (request.GET.get("make_names") or "").strip()
    rto_code = (request.GET.get("rto_code") or "").strip()
    cc = (request.GET.get("cc") or "").strip()
    fuel = (request.GET.get("fuel") or "").strip()
    sc = (request.GET.get("sc") or "").strip()
    mfg_year = (request.GET.get("mfg_year") or "").strip()

    is_zd = (request.GET.get("is_zd") or "").strip().upper()
    is_cpa = (request.GET.get("is_cpa") or "").strip().upper()
    is_ncb = (request.GET.get("is_ncb") or "").strip().upper()

    if target_date:
        qs = qs.filter(
            (Q(from_date__lte=target_date) | Q(from_date__isnull=True)) &
            (Q(to_date__gte=target_date) | Q(to_date__isnull=True))
        )

    if product:
        if str(product).isdigit():
            qs = qs.filter(product_id=product)
        else:
            qs = qs.filter(product__name__iexact=str(product).strip())

    qs = apply_make_model_filter(qs, product, make_model_class)

    if sub_product:
        if str(sub_product).isdigit():
            qs = qs.filter(sub_product_id=sub_product)
        else:
            qs = qs.filter(sub_product__name__iexact=str(sub_product).strip())

    if fuel:
        if str(fuel).isdigit():
            qs = qs.filter(fuel_type_id=fuel)
        else:
            qs = qs.filter(fuel_type__name__iexact=str(fuel).strip())

    matching_rto_names = []
    if rto_code:
        potential_rtos = RTOMaster.objects.filter(rto_cluster__icontains=rto_code)
        for rto_record in potential_rtos:
            if strict_match_in_cluster(rto_code, rto_record.rto_cluster):
                matching_rto_names.append(rto_record.rto_name.strip().upper())

        if matching_rto_names:
            q_rto = Q()
            for rto_name in matching_rto_names:
                q_rto |= Q(new_rto_list__icontains=rto_name)
            qs = qs.filter(q_rto)
        else:
            qs = qs.filter(new_rto_list__icontains=rto_code.strip())

    matching_make_groups = []
    if make_names:
        potential_makes = MakeModelMaster.objects.filter(make_model_cluster__icontains=make_names)
        for make_record in potential_makes:
            if strict_match_in_cluster(make_names, make_record.make_model_cluster):
                matching_make_groups.append(make_record.make_model_name.strip().upper())

        if matching_make_groups:
            q_make = Q()
            for group_name in matching_make_groups:
                if group_name:
                    q_make |= Q(new_vehicle_makes__icontains=group_name.strip())
            qs = qs.filter(q_make)
        else:
            qs = qs.filter(new_vehicle_makes__icontains=make_names.strip())

    if cc:
        try:
            cc_val = float(cc)
            qs = qs.filter(
                (Q(cc_min__lte=cc_val) | Q(cc_min__isnull=True)) &
                (Q(cc_max__gte=cc_val) | Q(cc_max__isnull=True))
            )
        except ValueError:
            pass

    if sc:
        try:
            sc_val = float(sc)
            qs = qs.filter(
                (Q(sc_min__lte=sc_val) | Q(sc_min__isnull=True)) &
                (Q(sc_max__gte=sc_val) | Q(sc_max__isnull=True))
            )
        except ValueError:
            pass

    if mfg_year and mfg_year.isdigit() and target_date:
        try:
            target_dt = datetime.strptime(target_date, "%Y-%m-%d")
            mfg_dt = datetime(int(mfg_year), 1, 1)
            exact_age = round((target_dt - mfg_dt).days / 365.25, 2)
            qs = qs.filter(
                (Q(vehicle_age_min__lte=exact_age) | Q(vehicle_age_min__isnull=True)) &
                (Q(vehicle_age_max__gte=exact_age) | Q(vehicle_age_max__isnull=True))
            )
        except ValueError:
            pass

    if is_zd:
        qs = qs.filter(Q(is_zd__code__iexact=is_zd) | Q(is_zd__code__iexact="NA"))

    if is_cpa:
        qs = qs.filter(Q(is_cpa__code__iexact=is_cpa) | Q(is_cpa__code__iexact="NA"))

    if is_ncb:
        qs = qs.filter(Q(is_ncb__code__iexact=is_ncb) | Q(is_ncb__code__iexact="NA"))

    qs = qs.order_by(
        F("po_net_rate").desc(nulls_last=True),
        F("po_od_rate").desc(nulls_last=True),
        F("po_flat_amount").desc(nulls_last=True),
        "-id"
    )

    results = []
    seen_groups = set()

    if has_searched:
        for row in qs.iterator(chunk_size=2000):
            if rto_code and matching_rto_names:
                if not row.new_rto_list:
                    continue
                row_rtos = [x.strip().upper() for x in row.new_rto_list.split(",")]
                if not any(x in matching_rto_names for x in row_rtos):
                    continue

            if make_names and matching_make_groups:
                if not row.new_vehicle_makes:
                    continue
                row_makes = [x.strip().upper() for x in row.new_vehicle_makes.split(",")]
                if not any(x in matching_make_groups for x in row_makes):
                    continue

            gid = row.group_id if row.group_id is not None else row.id
            if gid not in seen_groups:
                row.display_group_id = gid

                mmc_name = row.make_model_class.name.strip().upper() if row.make_model_class else "NA"
                
                if mmc_name in ["NA", ""]:
                    prod_name = row.product.name if row.product else ""
                    translated_name = get_translated_make_model(prod_name)
                    if translated_name:
                        row.make_model_class = MakeModelClassMaster(name=translated_name)
                        row.display_make_model_class = translated_name
                    else:
                        row.display_make_model_class = "NA"
                else:
                    row.display_make_model_class = row.make_model_class.name

                if row.po_net_rate and row.po_net_rate > 0:
                    row.po_rate = row.po_net_rate
                elif row.po_od_rate and row.po_od_rate > 0:
                    row.po_rate = row.po_od_rate
                elif row.po_tp_rate and row.po_tp_rate > 0:
                    row.po_rate = row.po_tp_rate
                else:
                    row.po_rate = 0.0

                existing_lock = None
                if vehicle_no and policy_holder_name:
                    existing_lock = LockedPolicy.objects.filter(
                        source_rate=row,
                        vehicle_no__iexact=vehicle_no,
                        policy_holder_name__iexact=policy_holder_name
                    ).order_by("-id").first()

                row.lock_status = existing_lock.status if existing_lock else "UNLOCKED"

                results.append(row)
                seen_groups.add(gid)

            # Capped high (not at the 300 display limit) so the same-insurer
            # collapse below still sees the full candidate set before it's
            # cut down for display.
            if len(results) >= 2000:
                break

    # Different group_ids can still carry what is effectively the same rate
    # card - same insurer, tariff range, PO type, flat amount, and add TNC,
    # differing only in the payout rate (a data entry issue, not two
    # genuinely different offers). Collapse those down to the single
    # lowest-rate row. Group ID and the rate itself are deliberately left out
    # of the matching key - group_id is exactly what differs between these
    # near-duplicates, and rate is the one field allowed to differ. Status is
    # also left out of the key on purpose: the real-world case this is fixing
    # has the lower (correct) rate on the ACTIVE row and a stale, wildly
    # inflated rate sitting on an INACTIVE duplicate - those still need to
    # collapse together, with the lowest rate's own status winning.
    def effective_rate(row):
        if row.po_type == "On OD and TP":
            return (row.po_od_rate or 0) + (row.po_tp_rate or 0)
        return row.po_rate or 0

    best_by_key = {}
    for row in results:
        row.effective_rate = effective_rate(row)
        key = (
            row.insurance_company,
            row.tariff_min, row.tariff_max,
            row.po_type,
            row.po_flat_amount,
            row.add_tnc,
        )
        existing = best_by_key.get(key)
        if existing is None or row.effective_rate < existing.effective_rate:
            best_by_key[key] = row

    results = sorted(best_by_key.values(), key=lambda r: r.effective_rate, reverse=True)[:300]

    selected = {
        "target_date": target_date,
        "vehicle_no": vehicle_no,
        "policy_holder_name": policy_holder_name,
        "product": product,
        "make_model_class": make_model_class,
        "sub_product": sub_product,
        "make_names": make_names,
        "rto_code": rto_code,
        "cc": cc,
        "fuel": fuel,
        "sc": sc,
        "mfg_year": mfg_year,
        "is_zd": is_zd,
        "is_cpa": is_cpa,
        "is_ncb": is_ncb,
    }

    return results, has_searched, selected


def policy_lock_checker(request):
    results, has_searched, selected = _run_policy_lock_checker_search(request)

    make_name_list = sorted({
        item.strip()
        for value in MakeModelMaster.objects.exclude(make_model_cluster__isnull=True)
        .exclude(make_model_cluster="")
        .values_list("make_model_cluster", flat=True)
        for item in str(value).split(",")
        if item.strip()
    })

    return render(request, "policy_lock_checker.html", {
        "data": results,
        "total_found": len(results),
        "has_searched": has_searched,
        "product_list": ProductMaster.objects.all().order_by("name"),
        "sub_product_list": SubProductMaster.objects.all().order_by("name"),
        "fuel_list": FuelTypeMaster.objects.all().order_by("name"),
        "make_model_class_list": get_dynamic_make_model_class_list(selected["product"]),
        "make_name_list": make_name_list,

        "make_class_mapping_json": json.dumps(NA_MAKE_MODEL_MAP),
        "all_make_classes_json": json.dumps(list(MakeModelClassMaster.objects.exclude(name__iexact="NA").values('id', 'name'))),

        "selected": selected,
    })

# -------------------------
# LOCK POLICY
# -------------------------
def lock_unlock_policy(request, rate_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    action_type = (request.POST.get("action_type") or "").strip().upper()
    vehicle_no = (request.POST.get("vehicle_no") or "").strip()
    policy_holder_name = (request.POST.get("policy_holder_name") or "").strip()
    target_date = (request.POST.get("target_date") or "").strip()

    if not vehicle_no:
        return JsonResponse({"success": False, "message": "Vehicle No. is required."})

    if action_type != "LOCK":
        return JsonResponse({"success": False, "message": "Invalid action."})
        
    today_str = datetime.today().strftime("%Y-%m-%d")
    if target_date != today_str:
        return JsonResponse({
            "success": False, 
            "message": "Policies can only be locked for today's date. Past or future dates are restricted."
        })

    rate_obj = get_object_or_404(
        RateMaster.objects.select_related("product", "sub_product", "fuel_type"),
        id=rate_id
    )

    po_rate = 0.0
    if rate_obj.po_net_rate and rate_obj.po_net_rate > 0:
        po_rate = rate_obj.po_net_rate
    elif rate_obj.po_od_rate and rate_obj.po_od_rate > 0:
        po_rate = rate_obj.po_od_rate
    elif rate_obj.po_tp_rate and rate_obj.po_tp_rate > 0:
        po_rate = rate_obj.po_tp_rate

    obj, created = LockedPolicy.objects.get_or_create(
        source_rate=rate_obj,
        vehicle_no=vehicle_no,
        policy_holder_name=policy_holder_name,
        defaults={
            "product_name": rate_obj.product.name if rate_obj.product else "",
            "sub_product_name": rate_obj.sub_product.name if rate_obj.sub_product else "",
            "insurance_company": rate_obj.insurance_company,
            "po_type": rate_obj.po_type,
            "po_rate": po_rate,
            "po_flat_amount": rate_obj.po_flat_amount,
            "add_tnc": rate_obj.add_tnc,
            "locked_by": request.user,
            "rto_code": request.POST.get("rto_code", ""),
            "make_name": request.POST.get("make_names", ""),
            "fuel": request.POST.get("fuel", ""),
            "cc": request.POST.get("cc", ""),
            "sc": request.POST.get("sc", ""),
            "mfg_year": request.POST.get("mfg_year", ""),
            "status": "LOCKED",
        }
    )

    # Object-level check: the get_or_create lookup above is keyed on
    # (source_rate, vehicle_no, policy_holder_name) only — not locked_by — so
    # without this, a second user submitting the same vehicle/holder combo
    # would find and silently take over someone else's existing lock
    # (every field below, including locked_by, gets unconditionally
    # overwritten). ADMIN keeps the ability to reassign/fix any lock;
    # everyone else can only touch a lock that's already theirs, or brand new.
    is_admin = request.user.groups.filter(name="ADMIN").exists()
    if not created and obj.locked_by_id and obj.locked_by_id != request.user.id and not is_admin:
        return JsonResponse({
            "success": False,
            "message": "This vehicle/policy holder is already locked by another user.",
        }, status=403)

    obj.product_name = rate_obj.product.name if rate_obj.product else ""
    obj.sub_product_name = rate_obj.sub_product.name if rate_obj.sub_product else ""
    obj.insurance_company = rate_obj.insurance_company
    obj.po_type = rate_obj.po_type
    obj.po_rate = po_rate
    obj.po_flat_amount = rate_obj.po_flat_amount
    obj.add_tnc = rate_obj.add_tnc
    obj.rto_code = request.POST.get("rto_code", "")
    obj.make_name = request.POST.get("make_names", "")
    obj.fuel = request.POST.get("fuel", "")
    obj.cc = request.POST.get("cc", "")
    obj.sc = request.POST.get("sc", "")
    obj.mfg_year = request.POST.get("mfg_year", "")
    obj.status = "LOCKED"
    obj.locked_by = request.user
    obj.locked_at = timezone.now()
    obj.save()

    return JsonResponse({"success": True, "message": f"Successfully locked policy for {vehicle_no}."})

# -------------------------
# LOCKED POLICY DASHBOARD
# -------------------------
def locked_policy_dashboard(request):
    # Data-privacy scoping: everyone except ADMIN only ever sees their own
    # locked policies — matches page_access_required's existing "ADMIN sees
    # everything, per-page groups see only their own scope" convention.
    is_admin = request.user.groups.filter(name="ADMIN").exists()

    qs = LockedPolicy.objects.select_related(
        "source_rate",
        "source_rate__product",
        "source_rate__make_model_class",
        "locked_by"
    ).all()
    if not is_admin:
        qs = qs.filter(locked_by=request.user)

    vehicle_no = (request.GET.get("vehicle_no") or "").strip()
    policy_holder_name = (request.GET.get("policy_holder_name") or "").strip()
    insurance_company = (request.GET.get("insurance_company") or "").strip()
    locked_by_user = (request.GET.get("locked_by_user") or "").strip()

    if vehicle_no:
        qs = qs.filter(vehicle_no__iexact=vehicle_no)
    if policy_holder_name:
        qs = qs.filter(policy_holder_name__iexact=policy_holder_name)
    if insurance_company:
        qs = qs.filter(insurance_company__iexact=insurance_company)
    if locked_by_user:
        qs = qs.filter(locked_by__username__iexact=locked_by_user)

    # Filter dropdown option lists must carry the same scoping as qs — for a
    # non-admin, listing e.g. "root" as a possible Locked By value would leak
    # other users' activity even though they can't open those records.
    all_locked = LockedPolicy.objects.select_related("locked_by").all()
    if not is_admin:
        all_locked = all_locked.filter(locked_by=request.user)

    unique_vehicles = sorted(list(set(all_locked.exclude(vehicle_no__isnull=True).exclude(vehicle_no="").values_list("vehicle_no", flat=True))))
    unique_holders = sorted(list(set(all_locked.exclude(policy_holder_name__isnull=True).exclude(policy_holder_name="").values_list("policy_holder_name", flat=True))))
    unique_companies = sorted(list(set(all_locked.exclude(insurance_company__isnull=True).exclude(insurance_company="").values_list("insurance_company", flat=True))))
    unique_users = sorted(list(set(all_locked.exclude(locked_by__isnull=True).values_list("locked_by__username", flat=True))))

    records = list(qs.order_by("-created_at")[:300])

    fuel_name_by_id = {str(f.id): f.name for f in FuelTypeMaster.objects.all()}

    for row in records:
        row.display_fuel = fuel_name_by_id.get((row.fuel or "").strip(), row.fuel)
        if row.source_rate:
            mmc_name = row.source_rate.make_model_class.name.strip().upper() if row.source_rate.make_model_class else "NA"
            
            if mmc_name in ["NA", ""]:
                prod_name = row.source_rate.product.name if row.source_rate.product else ""
                translated_name = get_translated_make_model(prod_name)
                if translated_name:
                    row.source_rate.make_model_class = MakeModelClassMaster(name=translated_name)
                    row.display_make_model_class = translated_name
                else:
                    row.display_make_model_class = "NA"
            else:
                row.display_make_model_class = row.source_rate.make_model_class.name if row.source_rate.make_model_class else ""

    return render(request, "locked_policy_dashboard.html", {
        "records": records,
        "total_records": qs.count(),
        "unique_vehicles": unique_vehicles,
        "unique_holders": unique_holders,
        "unique_companies": unique_companies,
        "unique_users": unique_users,
        "selected": {
            "vehicle_no": vehicle_no,
            "policy_holder_name": policy_holder_name,
            "insurance_company": insurance_company,
            "locked_by_user": locked_by_user,
        }
    })

# -------------------------
# EXECUTIVE ANALYSIS DASHBOARD
# -------------------------
# Some insurer grids spell the same state's RTO prefix differently, and at
# least one cluster has a stray typo. Fold those onto the standard code so
# they land in the same heatmap bucket instead of splitting a state's data
# across two rows.
STATE_CODE_ALIASES = {
    "TG": "TS",  # Telangana - some grids use TG instead of the official TS
    "OR": "OD",  # Odisha - legacy "OR" alongside the current "OD"
    "GC": "CG",  # Chhattisgarh - one-off typo found in an SBI cluster
}

# Insurers embed the grid's upload month directly in the batch label, e.g.
# "MAGMA_APR26_CG2" or "SBI_MAY26_MH_-_M". Grids that instead use a plain,
# undated label (e.g. SBI's own "AP"/"CG", or "ALLINDIA") don't carry a month
# at all — those are treated as evergreen and stay visible regardless of
# which month is selected, rather than disappearing from the chart.
MONTH_TOKEN_RE = re.compile(r"(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)(\d{2})", re.IGNORECASE)
MONTH_NAMES = {
    "JAN": "January", "FEB": "February", "MAR": "March", "APR": "April",
    "MAY": "May", "JUN": "June", "JUL": "July", "AUG": "August",
    "SEP": "September", "OCT": "October", "NOV": "November", "DEC": "December",
}
MONTH_ORDER = list(MONTH_NAMES.keys())
EVERGREEN = "EVERGREEN"


def month_bucket_for(token):
    m = MONTH_TOKEN_RE.search(token)
    if not m:
        return EVERGREEN
    return m.group(1).upper() + m.group(2)


def month_label_for(code):
    if code == EVERGREEN:
        return "No specific month"
    return f"{MONTH_NAMES.get(code[:3], code[:3])} 20{code[3:]}"


# Premium/tariff (PI) fields, not payout (PO) - what the policy is priced
# at, not what the broker earns. pi_tp_2..pi_tp_5 are the year 2-5 legs
# of a multi-year TP schedule; no upload has ever populated them, but
# they're wired up now so they light up the moment that data arrives.
RATE_FIELDS = {
    "net": "pi_net_rate", "od": "pi_od_rate", "tp": "pi_tp_rate",
    "tp2": "pi_tp_2", "tp3": "pi_tp_3", "tp4": "pi_tp_4", "tp5": "pi_tp_5",
}
RATE_FIELD_LABELS_PY = {
    "net": "Net", "od": "OD", "tp": "TP",
    "tp2": "TP2", "tp3": "TP3", "tp4": "TP4", "tp5": "TP5",
}

# Payout (PO) side of the same OD/TP/Net split — what the brokerage earns,
# not what the policy is priced at. No po_tp_2..po_tp_5 or po_vli equivalent
# exists on RateMaster, and po_flat_amount is deliberately excluded here for
# the same reason pi_flat_amount is excluded from RATE_FIELDS above: it's a
# fixed rupee amount, not a percentage rate, so averaging it in with OD/TP/Net
# would be comparing incompatible units.
PO_RATE_FIELDS = {
    "net": "po_net_rate", "od": "po_od_rate", "tp": "po_tp_rate",
}
PO_RATE_FIELD_LABELS_PY = {
    "net": "Net", "od": "OD", "tp": "TP",
}


def _track_min_max(cell, prefix, value):
    if value is None:
        return
    cur_min = cell.get(f"{prefix}_min")
    cur_max = cell.get(f"{prefix}_max")
    cell[f"{prefix}_min"] = value if cur_min is None else min(cur_min, value)
    cell[f"{prefix}_max"] = value if cur_max is None else max(cur_max, value)


def _build_rto_state_map():
    rto_state_map = {}
    for name, cluster in RTOMaster.objects.values_list("rto_name", "rto_cluster"):
        if not name or not cluster:
            continue
        states = set()
        for item in cluster.split(","):
            item = item.strip()
            if len(item) >= 2 and item[:2].isalpha():
                code = item[:2].upper()
                states.add(STATE_CODE_ALIASES.get(code, code))
        if states:
            rto_state_map[name.strip().upper()] = states
    return rto_state_map


def _filtered_rate_master_qs(request):
    qs = RateMaster.objects.filter(is_deleted="NO", status__in=["ACTIVE", "INACTIVE"])

    # Same date-range overlap filter used on the Rate Master Matrix
    # (dashboard/edit_rate): keeps only rows whose validity window
    # (from_date/to_date) covers the selected date, not an exact-match on
    # any single date field.
    date_range = (request.GET.get("date_range") or "").strip()
    if date_range:
        dates = date_range.split(" - ")
        if len(dates) == 2:
            d_from, d_to = dates[0].strip(), dates[1].strip()
            if d_from and d_to:
                qs = qs.filter(
                    (Q(from_date__lte=d_to) | Q(from_date__isnull=True)) &
                    (Q(to_date__gte=d_from) | Q(to_date__isnull=True))
                )

    # Engine CC / seating-capacity filters: a row matches if the entered
    # value falls inside its cc_min-cc_max (or sc_min-sc_max) band, treating
    # a null bound as open-ended - same semantics as the quote simulator's
    # CC/SC matching.
    cc = (request.GET.get("cc") or "").strip()
    if cc:
        try:
            cc_val = float(cc)
        except ValueError:
            cc_val = None
        if cc_val is not None:
            qs = qs.filter(
                (Q(cc_min__lte=cc_val) | Q(cc_min__isnull=True)) &
                (Q(cc_max__gte=cc_val) | Q(cc_max__isnull=True))
            )

    sc = (request.GET.get("sc") or "").strip()
    if sc:
        try:
            sc_val = float(sc)
        except ValueError:
            sc_val = None
        if sc_val is not None:
            qs = qs.filter(
                (Q(sc_min__lte=sc_val) | Q(sc_min__isnull=True)) &
                (Q(sc_max__gte=sc_val) | Q(sc_max__isnull=True))
            )

    return qs, date_range


def _row_states_and_months(rto_list, rto_state_map):
    states, months = set(), set()
    for token in (rto_list or "").split(","):
        token = token.strip().upper()
        if not token:
            continue
        if token in rto_state_map:
            states |= rto_state_map[token]
        months.add(month_bucket_for(token))
    return states, months or {EVERGREEN}


def business_analysis(request):
    rto_state_map = _build_rto_state_map()
    qs, date_range = _filtered_rate_master_qs(request)
    auto_loaded_today = request.GET.get("auto_loaded_today") or ""

    def new_cell():
        cell = {"age_min": None, "age_max": None, "cc_min": None, "cc_max": None,
                "tariff_min": None, "tariff_max": None, "ncb": set(), "cpa": set(), "zd": set()}
        for key in RATE_FIELDS:
            cell[f"{key}_min"] = None
            cell[f"{key}_max"] = None
            cell[f"{key}_n"] = 0
        return cell

    agg = defaultdict(new_cell)
    months_seen = set()

    for row in qs.values(
        "product__name", "sub_product__name", "insurance_company", "new_rto_list",
        *RATE_FIELDS.values(),
        "vehicle_age_min", "vehicle_age_max", "cc_min", "cc_max",
        "tariff_min", "tariff_max",
        "is_ncb__code", "is_cpa__code", "is_zd__code",
    ).iterator(chunk_size=5000):
        rto_list = row["new_rto_list"]
        if not rto_list:
            continue

        states, months = _row_states_and_months(rto_list, rto_state_map)
        if not states:
            continue
        months_seen |= months

        product = row["product__name"] or "Uncategorized"
        sub_product = row["sub_product__name"] or "Uncategorized"
        insurer = row["insurance_company"]

        for state in states:
            for mo in months:
                cell = agg[(product, sub_product, insurer, state, mo)]
                for key, field in RATE_FIELDS.items():
                    val = row[field]
                    if val and val > 0:
                        cell[f"{key}_n"] += 1
                        _track_min_max(cell, key, val)

                _track_min_max(cell, "age", row["vehicle_age_min"])
                _track_min_max(cell, "age", row["vehicle_age_max"])
                _track_min_max(cell, "cc", row["cc_min"])
                _track_min_max(cell, "cc", row["cc_max"])
                _track_min_max(cell, "tariff", row["tariff_min"])
                _track_min_max(cell, "tariff", row["tariff_max"])

                if row["is_ncb__code"]:
                    cell["ncb"].add(row["is_ncb__code"])
                if row["is_cpa__code"]:
                    cell["cpa"].add(row["is_cpa__code"])
                if row["is_zd__code"]:
                    cell["zd"].add(row["is_zd__code"])

    records = []
    product_volume = defaultdict(int)
    sub_product_volume = defaultdict(int)
    insurer_set = set()
    state_set = set()

    for (product, sub_product, insurer, state, month), cell in agg.items():
        total_n = sum(cell[f"{key}_n"] for key in RATE_FIELDS)
        if not total_n:
            continue
        rec = {
            "product": product, "sub_product": sub_product, "insurer": insurer, "state": state, "month": month,
            "age_min": cell["age_min"], "age_max": cell["age_max"],
            "cc_min": cell["cc_min"], "cc_max": cell["cc_max"],
            "tariff_min": cell["tariff_min"], "tariff_max": cell["tariff_max"],
            "ncb": sorted(cell["ncb"]), "cpa": sorted(cell["cpa"]), "zd": sorted(cell["zd"]),
        }
        for key in RATE_FIELDS:
            n = cell[f"{key}_n"]
            rec[key] = {"min": round(cell[f"{key}_min"], 2), "max": round(cell[f"{key}_max"], 2), "n": n} if n else None
        records.append(rec)
        product_volume[product] += total_n
        sub_product_volume[sub_product] += total_n
        insurer_set.add(insurer)
        state_set.add(state)

    def month_sort_key(code):
        if code == EVERGREEN:
            return (9999, 99)
        mon, yy = code[:3], code[3:]
        return (int(yy), MONTH_ORDER.index(mon) if mon in MONTH_ORDER else 99)

    months_list = [
        {"code": m, "label": month_label_for(m)}
        for m in sorted(months_seen, key=month_sort_key) if m != EVERGREEN
    ]

    chart_data = {
        "records": records,
        "products": sorted(product_volume.keys(), key=lambda p: -product_volume[p]),
        "sub_products": sorted(sub_product_volume.keys(), key=lambda p: -sub_product_volume[p]),
        "insurers": sorted(insurer_set),
        "states": sorted(state_set),
        "months": months_list,
    }

    return render(request, "analysis.html", {
        "chart_data": chart_data,
        "date_range": date_range,
        "auto_loaded_today": auto_loaded_today,
        "cc_filter": request.GET.get("cc") or "",
        "sc_filter": request.GET.get("sc") or "",
    })


def analysis_pivot_data(request):
    rto_state_map = _build_rto_state_map()
    qs, _ = _filtered_rate_master_qs(request)

    # Pass 1: cheap query for the universe of *real* months in play, so an
    # evergreen row (no month tag of its own) knows which months to fold
    # into. Only needs the RTO text, not the full row.
    real_months = set()
    for rto_list in qs.values_list("new_rto_list", flat=True).iterator(chunk_size=5000):
        if not rto_list:
            continue
        for token in rto_list.split(","):
            token = token.strip().upper()
            if token:
                mo = month_bucket_for(token)
                if mo != EVERGREEN:
                    real_months.add(mo)

    # Pass 2: accumulate into (product, sub_product, insurer, state, month,
    # rate_type, is_evergreen) -> {sum, count} instead of emitting one flat
    # row per matching RateMaster row. Many raw rows share the same pivot
    # key (different vehicle-age/CC/NCB/etc. variants of the same
    # insurer/state/month/rate_type) -- pre-summing collapses those here
    # rather than shipping every one of them to the browser. This stays
    # exactly as correct for the pivot table as the finest-grain values
    # would be: the client's aggregator computes sum(sum) / sum(count)
    # under any regrouping, which equals a true average of the raw values,
    # never an average-of-averages. An evergreen row is folded into every
    # real month in play instead of its own EVERGREEN bucket, tagged
    # is_evergreen so it stays traceable rather than silently blended in.
    agg = defaultdict(lambda: [0.0, 0])
    for row in qs.values(
        "product__name", "sub_product__name", "insurance_company", "new_rto_list",
        *RATE_FIELDS.values(),
    ).iterator(chunk_size=5000):
        states, months = _row_states_and_months(row["new_rto_list"], rto_state_map)
        if not states:
            continue

        product = row["product__name"] or "Uncategorized"
        sub_product = row["sub_product__name"] or "Uncategorized"
        insurer = row["insurance_company"]

        for state in states:
            for mo in months:
                is_evergreen = mo == EVERGREEN
                target_months = real_months if is_evergreen else {mo}
                for target_mo in target_months:
                    month_label = month_label_for(target_mo)
                    for key, field in RATE_FIELDS.items():
                        val = row[field]
                        if val and val > 0:
                            cell = agg[(
                                product, sub_product, insurer, state,
                                month_label, RATE_FIELD_LABELS_PY[key], is_evergreen,
                            )]
                            cell[0] += val
                            cell[1] += 1

    flat_rows = [
        {
            "product": p, "sub_product": sp, "insurer": ins, "state": st,
            "month": mo, "rate_type": rt, "is_evergreen": eg,
            "sum": round(total, 6), "count": n,
        }
        for (p, sp, ins, st, mo, rt, eg), (total, n) in agg.items()
    ]

    return JsonResponse(flat_rows, safe=False)


def _row_months_only(rto_list):
    """
    Same month-bucket derivation _row_states_and_months uses, minus the state
    resolution. analysis_payout_data doesn't break payout figures out by
    state, so it has no reason to build/consult the RTO->state map — and
    critically must NOT skip a row just because its RTO cluster doesn't
    resolve to a recognised state (that's a gap in RTOMaster's cluster data,
    unrelated to whether the row's payout figures are valid).
    """
    months = {month_bucket_for(token.strip().upper()) for token in (rto_list or "").split(",") if token.strip()}
    return months or {EVERGREEN}


def analysis_payout_data(request):
    """
    PO (payout/commission) feed for the Leaderboard and Commission Tracking
    charts. Deliberately a separate endpoint from analysis_pivot_data, which
    the existing heatmap and pivot table call and which stays untouched —
    mirrors its filtering (_filtered_rate_master_qs) and month-bucket
    derivation (grid/RTO naming, e.g. "MAGMA_APR26_...") so all three charts
    stay in sync under the same page-level filters, but aggregates the po_*
    fields instead of pi_*, and drops the state dimension neither new chart
    needs.
    """
    qs, _ = _filtered_rate_master_qs(request)

    real_months = set()
    for rto_list in qs.values_list("new_rto_list", flat=True).iterator(chunk_size=5000):
        for mo in _row_months_only(rto_list):
            if mo != EVERGREEN:
                real_months.add(mo)

    # (product, sub_product, insurer, month, rate_type, is_evergreen) -> [sum, count],
    # same sum/count-not-average shape as analysis_pivot_data so the client
    # can compute a true weighted average under any regrouping.
    agg = defaultdict(lambda: [0.0, 0])
    for row in qs.values(
        "product__name", "sub_product__name", "insurance_company", "new_rto_list",
        *PO_RATE_FIELDS.values(),
    ).iterator(chunk_size=5000):
        product = row["product__name"] or "Uncategorized"
        sub_product = row["sub_product__name"] or "Uncategorized"
        insurer = row["insurance_company"]

        for mo in _row_months_only(row["new_rto_list"]):
            is_evergreen = mo == EVERGREEN
            target_months = real_months if is_evergreen else {mo}
            for target_mo in target_months:
                month_label = month_label_for(target_mo)
                for key, field in PO_RATE_FIELDS.items():
                    val = row[field]
                    if val and val > 0:
                        cell = agg[(
                            product, sub_product, insurer,
                            month_label, PO_RATE_FIELD_LABELS_PY[key], is_evergreen,
                        )]
                        cell[0] += val
                        cell[1] += 1

    flat_rows = [
        {
            "product": p, "sub_product": sp, "insurer": ins,
            "month": mo, "rate_type": rt, "is_evergreen": eg,
            "sum": round(total, 6), "count": n,
        }
        for (p, sp, ins, mo, rt, eg), (total, n) in agg.items()
    ]

    return JsonResponse(flat_rows, safe=False)

# -------------------------
# AUDIT TRAIL LOGS
# -------------------------
# Action types tracked on the Security Audit & History Log page - edit/update
# actions only. MOTOR_POINTS_SEARCH and the sso_ticket_* events have their own
# purposes and are never shown here. Keep this in sync with
# SECURITY_AUDIT_LOG_ACTIONS in insurance/tasks.py (the 7-day purge job).
SECURITY_AUDIT_LOG_ACTIONS = ["MANUAL EDIT", "BULK UPDATE", "HEALTH RATE EDIT", "HEALTH BULK UPDATE"]

def _filtered_audit_logs(request):
    # Bounded to the last 7 days to match the retention window enforced by
    # cleanup_security_audit_logs (insurance/tasks.py).
    retention_cutoff = timezone.now() - timedelta(days=7)
    qs = AuditLog.objects.filter(
        action__in=SECURITY_AUDIT_LOG_ACTIONS, timestamp__gte=retention_cutoff
    ).select_related("user").order_by("-timestamp")

    user_id = (request.GET.get("user") or "").strip()
    action = (request.GET.get("action") or "").strip()
    date_range = (request.GET.get("date_range") or "").strip()

    if user_id:
        qs = qs.filter(user_id=user_id)
    if action:
        qs = qs.filter(action=action)
    if date_range:
        dates = date_range.split(" - ")
        if len(dates) == 2:
            d_from, d_to = dates[0].strip(), dates[1].strip()
            if d_from and d_to:
                qs = qs.filter(timestamp__date__gte=d_from, timestamp__date__lte=d_to)

    return qs, {"user": user_id, "action": action, "date_range": date_range}

def audit_logs(request):
    qs, selected = _filtered_audit_logs(request)

    # AuditLog's default ordering (-timestamp) makes .distinct() on a single
    # column ineffective at the SQL level (timestamp isn't in the SELECT, so
    # "DISTINCT action" ends up including every differently-timestamped
    # row) - dedupe in Python instead of relying on DB-level DISTINCT here.
    # Scoped to SECURITY_AUDIT_LOG_ACTIONS to match _filtered_audit_logs's base filter.
    tracked_qs = AuditLog.objects.filter(action__in=SECURITY_AUDIT_LOG_ACTIONS)
    user_ids = set(tracked_qs.exclude(user__isnull=True).values_list("user_id", flat=True))
    user_choices = User.objects.filter(id__in=user_ids).order_by("username")
    action_choices = sorted(set(tracked_qs.values_list("action", flat=True)))

    return render(request, "audit_log.html", {
        "logs": qs[:200],
        "user_choices": user_choices,
        "action_choices": action_choices,
        "selected": selected,
    })

def export_audit_log_xlsx(request):
    qs, _ = _filtered_audit_logs(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Audit Log"
    ws.append(["ID", "Date & Time", "User", "Action Type", "Details"])

    for log in qs[:2000]:
        ws.append([
            log.id,
            log.timestamp.strftime("%Y-%m-%d %H:%M:%S") if log.timestamp else "",
            log.user.username if log.user else "System",
            log.action,
            log.details,
        ])

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="audit_log.xlsx"'
    wb.save(response)
    return response

# -------------------------
# EMAIL BACKGROUND
# -------------------------
def send_email_background(subject, message, recipient_list):
    try:
        send_mail(
            subject=subject,
            message=message,
            from_email=settings.EMAIL_HOST_USER,
            recipient_list=recipient_list,
            fail_silently=False,
        )
        print("✅ Background email sent successfully!")
    except Exception as e:
        print(f"\n❌ BACKGROUND EMAIL FAILED: {e}\n")

# -------------------------
# GRID MANAGEMENT
# -------------------------
def grid_management(request):
    if request.method == "POST":
        action = request.POST.get("action")

        if action == "update_status":
            doc_id = request.POST.get("doc_id")
            new_status = request.POST.get("status")
            if doc_id and new_status:
                doc = get_object_or_404(GridDocument, id=doc_id)
                doc.status = new_status
                doc.save()
            return redirect("grid_management")

        insurer_name = request.POST.get("insurer_name")
        remarks = request.POST.get("remarks")
        uploaded_file = request.FILES.get("uploaded_file")

        if insurer_name and uploaded_file:
            GridDocument.objects.create(
                insurer_name=insurer_name,
                remarks=remarks,
                uploaded_file=uploaded_file,
                uploaded_by=request.user,
                status="PENDING"
            )

            uploader_name = "System User"
            subject = f"🔔 New Grid Uploaded: {insurer_name}"
            message = (
                f"Hello,\n\n"
                f"A new provider grid has just been uploaded to the Insurance Portal.\n\n"
                f"• Insurer: {insurer_name}\n"
                f"• Uploaded By: {uploader_name}\n"
                f"• Remarks: {remarks or 'No remarks provided.'}\n\n"
                f"Please log in to the portal to download it."
            )

            email_thread = threading.Thread(
                target=send_email_background,
                args=(subject, message, ["harsh.t@arhamsecure.com"])
            )
            email_thread.start()

            return redirect("grid_management")

    total_all = GridDocument.objects.count()
    status_counts = dict(
        GridDocument.objects.values("status").annotate(n=Count("id")).values_list("status", "n")
    )

    # "all" (or anything unrecognized) clears the filter, matching the Clear
    # Filters convention used elsewhere — only a real STATUS_CHOICES code
    # narrows the queryset and highlights its card.
    status_filter = (request.GET.get("status") or "").strip()
    valid_status_codes = {code for code, _ in GridDocument.STATUS_CHOICES}
    selected_status = status_filter if status_filter in valid_status_codes else ""

    documents = GridDocument.objects.all().order_by("-uploaded_date")
    if selected_status:
        documents = documents.filter(status=selected_status)

    return render(request, "grid_management.html", {
        "documents": documents,
        "status_choices": GridDocument.STATUS_CHOICES,
        "status_counts": status_counts,
        "total_all": total_all,
        "selected_status": selected_status,
    })

# -------------------------
# POINTS AUDIT LOGS (Motor + Health "Check Eligibility" search trail)
# -------------------------
# Kept in sync with POINTS_SEARCH_ACTIONS in insurance/tasks.py (the 7-day
# purge job).
POINTS_SEARCH_ACTIONS = ["MOTOR_POINTS_SEARCH", "HEALTH_POINTS_SEARCH"]

def points_audit_logs(request):
    retention_cutoff = timezone.now() - timedelta(days=7)

    type_filter = (request.GET.get("type") or "").strip().lower()
    action_filter = {
        "motor": ["MOTOR_POINTS_SEARCH"],
        "health": ["HEALTH_POINTS_SEARCH"],
    }.get(type_filter, POINTS_SEARCH_ACTIONS)

    # Other active filters (vehicle_no, username, ...), preserved when a
    # Motor/Health/All toggle link is clicked so switching type doesn't
    # silently drop them. Trailing "&" so the template can just append
    # "type=...", even when this is empty.
    _qs_without_type = request.GET.copy()
    _qs_without_type.pop("type", None)
    filter_base_qs = _qs_without_type.urlencode() + "&" if _qs_without_type else ""

    qs = AuditLog.objects.filter(
        action__in=action_filter, timestamp__gte=retention_cutoff
    ).select_related("user").order_by("-timestamp")

    vehicle_no_filter = (request.GET.get("vehicle_no") or "").strip()
    policy_holder_name_filter = (request.GET.get("policy_holder_name") or "").strip()
    insurance_company_filter = (request.GET.get("insurance_company") or "").strip()
    username_filter = (request.GET.get("username") or "").strip()
    # Health-tab equivalents of the three Motor-only fields above (Vehicle
    # No./Policy Holder/Make-Model Name don't apply to Health searches) --
    # same field names the Health Points Checker itself submits, so they
    # match what's actually stored in a HEALTH_POINTS_SEARCH log's details.
    product_name_filter = (request.GET.get("product_name") or "").strip()
    business_type_filter = (request.GET.get("business_type") or "").strip()
    policy_category_filter = (request.GET.get("policy_category") or "").strip()

    if vehicle_no_filter:
        qs = qs.filter(details__icontains=vehicle_no_filter)
    if policy_holder_name_filter:
        qs = qs.filter(details__icontains=policy_holder_name_filter)
    if insurance_company_filter:
        qs = qs.filter(details__icontains=insurance_company_filter)
    if username_filter:
        qs = qs.filter(user__username__icontains=username_filter)
    if product_name_filter:
        qs = qs.filter(details__icontains=product_name_filter)
    if business_type_filter:
        qs = qs.filter(details__icontains=business_type_filter)
    if policy_category_filter:
        qs = qs.filter(details__icontains=policy_category_filter)

    logs = qs[:500]

    all_logs_for_dropdowns = AuditLog.objects.filter(
        action__in=action_filter, timestamp__gte=retention_cutoff
    ).select_related("user").order_by("-timestamp")[:1000]
    unique_vehicles = set()
    unique_holders = set()
    unique_companies = set()
    unique_users = set()
    unique_health_products = set()
    unique_health_sub_products = set()
    unique_health_plan_types = set()

    for log in all_logs_for_dropdowns:
        if log.user and log.user.username:
            unique_users.add(log.user.username)

        try:
            clean_str = log.details.replace("Eligibility Check Parameters: ", "")
            params_dict = ast.literal_eval(clean_str)

            flat_params = {}
            if isinstance(params_dict, dict):
                for k, v in params_dict.items():
                    if isinstance(v, list) and len(v) > 0:
                        flat_params[k] = str(v[0]).strip()
                    else:
                        flat_params[k] = str(v).strip()

            if flat_params.get("vehicle_no"):
                unique_vehicles.add(flat_params["vehicle_no"])
            if flat_params.get("policy_holder_name"):
                unique_holders.add(flat_params["policy_holder_name"])
            if flat_params.get("make_names"):
                unique_companies.add(flat_params["make_names"])
            if flat_params.get("product_name"):
                unique_health_products.add(flat_params["product_name"])
            if flat_params.get("business_type"):
                unique_health_sub_products.add(flat_params["business_type"])
            if flat_params.get("policy_category"):
                unique_health_plan_types.add(flat_params["policy_category"])
        except:
            pass

    # Loaded once and reused for every row below, instead of running up to 4
    # separate ProductMaster/SubProductMaster/MakeModelClassMaster/FuelTypeMaster
    # queries per log row (which, at 500 rows, meant up to ~2000 queries on
    # this one page load).
    product_names = dict(ProductMaster.objects.values_list("id", "name"))
    sub_product_names = dict(SubProductMaster.objects.values_list("id", "name"))
    make_model_class_names = dict(MakeModelClassMaster.objects.values_list("id", "name"))
    fuel_type_names = dict(FuelTypeMaster.objects.values_list("id", "name"))

    for log in logs:
        try:
            clean_str = log.details.replace("Eligibility Check Parameters: ", "")
            params_dict = ast.literal_eval(clean_str)

            flat_params = {}
            if isinstance(params_dict, dict):
                for k, v in params_dict.items():
                    if isinstance(v, list) and len(v) > 0:
                        flat_params[k] = v[0]
                    else:
                        flat_params[k] = v

            if flat_params.get("product") and str(flat_params["product"]).isdigit():
                name = product_names.get(int(flat_params["product"]))
                if name:
                    flat_params["product"] = name

            if flat_params.get("sub_product") and str(flat_params["sub_product"]).isdigit():
                name = sub_product_names.get(int(flat_params["sub_product"]))
                if name:
                    flat_params["sub_product"] = name

            if flat_params.get("make_model_class") and str(flat_params["make_model_class"]).isdigit():
                name = make_model_class_names.get(int(flat_params["make_model_class"]))
                if name:
                    flat_params["make_model_class"] = name

            if flat_params.get("fuel") and str(flat_params["fuel"]).isdigit():
                name = fuel_type_names.get(int(flat_params["fuel"]))
                if name:
                    flat_params["fuel"] = name

            log.params = flat_params
        except Exception:
            log.params = {}

    return render(request, "points_audit_logs.html", {
        "logs": logs,
        "unique_vehicles": sorted(list(unique_vehicles)),
        "unique_holders": sorted(list(unique_holders)),
        "unique_companies": sorted(list(unique_companies)),
        "unique_users": sorted(list(unique_users)),
        "unique_health_products": sorted(list(unique_health_products)),
        "unique_health_sub_products": sorted(list(unique_health_sub_products)),
        "unique_health_plan_types": sorted(list(unique_health_plan_types)),
        "filter_base_qs": filter_base_qs,
        "selected": {
            "vehicle_no": vehicle_no_filter,
            "policy_holder_name": policy_holder_name_filter,
            "insurance_company": insurance_company_filter,
            "username": username_filter,
            "product_name": product_name_filter,
            "business_type": business_type_filter,
            "policy_category": policy_category_filter,
            "type": type_filter,
        }
    })

# =========================================================
# REST API ENDPOINTS
# =========================================================
class ExportRatesAPIView(generics.ListAPIView):
    permission_classes = [HasAPIKey]
    queryset = RateMaster.objects.filter(is_deleted="NO")
    serializer_class = RateMasterSerializer


class PolicyLockCheckerAPIView(APIView):
    """
    Server-to-server JSON equivalent of policy_lock_checker (the Motor
    Points page): same search/filter/dedup logic via
    _run_policy_lock_checker_search, secured with a rest_framework_api_key
    key instead of a session login -- same pattern as IssueSSOTicketAPIView.
    """
    permission_classes = [HasAPIKey]

    def get(self, request):
        results, has_searched, selected = _run_policy_lock_checker_search(request)

        def serialize_row(row):
            # Mirrors the "Eligibility Results" table in policy_lock_checker.html
            # (see the {% for row in data %} block there) so the API's numbers
            # render exactly like the web page's, string-for-string.
            if row.po_type == "On OD and TP":
                po_rate = f"{row.po_od_rate or 0}+{row.po_tp_rate or 0}%"
            else:
                po_rate = f"{row.po_rate or 0}%"

            return {
                "group_id": row.display_group_id,
                "insurance_company": row.insurance_company,
                "tariff_range": f"{row.tariff_min or 0}-{row.tariff_max or 0}",
                "po_type": row.po_type,
                "po_rate": po_rate,
                "add_tnc": row.add_tnc or "-",
            }

        return Response({
            "has_searched": has_searched,
            "total_found": len(results),
            "filters": selected,
            "results": [serialize_row(row) for row in results],
        })


class HealthPayoutRatesAPIView(APIView):
    """
    Server-to-server JSON equivalent of health_payout_rates (the Health
    Points Checker's "Check Eligibility" search): same eligibility-matching
    logic via _run_health_payout_search / _build_health_payout_queryset,
    secured with a rest_framework_api_key key instead of a session login --
    same pattern as PolicyLockCheckerAPIView.
    """
    permission_classes = [HasAPIKey]

    def get(self, request):
        results, total_found, has_more, has_searched, selected = _run_health_payout_search(request)

        def serialize_row(row):
            # Mirrors the "Valid Matches Found" table in health_payout_rates.html
            # (see the {% for row in data %} block there) so the API's values
            # render exactly like the web page's, string-for-string -- payouts
            # are always on a NET 1 Year basis regardless of the policy term
            # searched, so this is intentionally not derived from policy_term.
            rate = row.applicable_rate
            applicable_rate = f"{rate:.2f}%" if rate is not None else "-%"

            return {
                "id": f"h-{row.id}",
                "insurer": row.insurance_company,
                "applicable_rate": applicable_rate,
                "payout_type": "NET 1 Year",
                "plans": row.plan_names or "-",
            }

        return Response({
            "has_searched": has_searched,
            "total_found": total_found,
            "has_more": has_more,
            "max_results": HEALTH_PAYOUT_MAX_RESULTS,
            "filters": selected,
            "results": [serialize_row(row) for row in results],
        })


class MakeModelMasterAPIView(APIView):
    """
    Reference-data helper for external API consumers (e.g. the tech
    manager's system calling PolicyLockCheckerAPIView): returns the exact
    strings valid for that endpoint's make_names, make_model_class, and
    product query params, so callers don't have to guess spellings. Purely
    read-only reference data -- secured the same way as the other
    server-to-server endpoints.
    """
    permission_classes = [HasAPIKey]

    def get(self, request):
        # Same cluster-token extraction as make_name_list in
        # policy_lock_checker -- these are the individual tokens
        # strict_match_in_cluster actually matches against, not the raw
        # make_model_name column.
        make_names = sorted({
            item.strip()
            for value in MakeModelMaster.objects.exclude(make_model_cluster__isnull=True)
            .exclude(make_model_cluster="")
            .values_list("make_model_cluster", flat=True)
            for item in str(value).split(",")
            if item.strip()
        })

        make_model_class = list(
            MakeModelClassMaster.objects.exclude(name__iexact="NA")
            .values_list("name", flat=True)
            .distinct()
            .order_by("name")
        )

        product = list(
            ProductMaster.objects.values_list("name", flat=True)
            .distinct()
            .order_by("name")
        )

        return Response({
            "make_names": make_names,
            "make_model_class": make_model_class,
            "product": product,
        })

# -------------------------
# MIS REVIEW
# -------------------------
def mis_review(request, pk):
    record = get_object_or_404(PolicyMISRecord.objects.select_related('source_document'), pk=pk)
    
    rules = ExtractionField.objects.filter(is_active=True).order_by('category', 'order_index')
    
    if request.method == 'POST':
        updated_json = record.raw_ai_json or {}
        
        for rule in rules:
            field_key = rule.field_name.strip().lower().replace(" ", "_")
            new_val = request.POST.get(field_key)
            if new_val is not None:
                updated_json[field_key] = new_val
                
        record.raw_ai_json = updated_json
        
        record.policy_number = updated_json.get("policy_number", record.policy_number)
        record.insurer_name = updated_json.get("insurance_company", record.insurer_name)
        record.insured_name = updated_json.get("insured_name", record.insured_name)
        record.save()
        
        return redirect('my_mis')

    field_data = []
    parsed = record.raw_ai_json or {}
    
    for rule in rules:
        field_key = rule.field_name.strip().lower().replace(" ", "_")
        val = parsed.get(field_key, "")
        
        is_missing_mandatory = rule.is_mandatory and not val
        
        dropdown_choices = []
        if getattr(rule, 'has_dropdown', False) and getattr(rule, 'dropdown_options', None):
            dropdown_choices = [x.strip() for x in rule.dropdown_options.split(',') if x.strip()]
            
        field_data.append({
            'rule': rule,
            'key': field_key,
            'value': val,
            'is_missing_mandatory': is_missing_mandatory,
            'dropdown_choices': dropdown_choices
        })
        
    return render(request, 'mis_review.html', {
        'record': record,
        'field_data': field_data
    })

# =========================================================
# SPECIAL RATE REQUESTS
# =========================================================
def special_rate_requests(request):
    mgbg_form = MgBgRateRequestForm(prefix="mgbg")
    bg_form = BgRateRequestForm(prefix="bg")
    active_section = "MG-BG"

    if request.method == "POST" and "submit_mgbg" in request.POST:
        active_section = "MG-BG"
        mgbg_form = MgBgRateRequestForm(request.POST, request.FILES, prefix="mgbg")
        if mgbg_form.is_valid():
            req_obj = mgbg_form.save(commit=False)
            req_obj.rate_type = "MG-BG"
            req_obj.requested_by = request.user
            req_obj.save()
            messages.success(request, f"✅ Your Special MG-BG request (Entry No {req_obj.entry_no}) has been submitted.")
            return redirect("special_rate_requests")
        messages.error(request, "⚠️ Please correct the errors below and resubmit.")

    elif request.method == "POST" and "submit_bg" in request.POST:
        active_section = "BG"
        bg_form = BgRateRequestForm(request.POST, request.FILES, prefix="bg")
        if bg_form.is_valid():
            req_obj = bg_form.save(commit=False)
            req_obj.rate_type = "BG"
            req_obj.requested_by = request.user
            req_obj.save()
            messages.success(request, f"✅ Your Special BG request (Entry No {req_obj.entry_no}) has been submitted.")
            return redirect("special_rate_requests")
        messages.error(request, "⚠️ Please correct the errors below and resubmit.")

    my_requests = SpecialRateRequest.objects.filter(requested_by=request.user).order_by("-created_at")

    return render(request, "special_rate_requests.html", {
        "mgbg_form": mgbg_form,
        "bg_form": bg_form,
        "active_section": active_section,
        "my_requests": my_requests,
    })


def special_rate_requests_review(request):
    requests_qs = SpecialRateRequest.objects.select_related("requested_by", "reviewed_by").all()

    status_counts = {"PENDING": 0, "APPROVED": 0, "REJECTED": 0}
    for row in requests_qs.values("status").annotate(n=Count("id")):
        if row["status"] in status_counts:
            status_counts[row["status"]] = row["n"]

    # Keyed without the hyphen ("MG-BG" -> MGBG) since Django template variable
    # lookups can't contain a "-" (`rate_type_counts.MG-BG` is a
    # TemplateSyntaxError, not a lookup miss).
    raw_type_counts = {"MG-BG": 0, "BG": 0}
    for row in requests_qs.values("rate_type").annotate(n=Count("id")):
        if row["rate_type"] in raw_type_counts:
            raw_type_counts[row["rate_type"]] = row["n"]
    rate_type_counts = {
        "MGBG": raw_type_counts["MG-BG"],
        "BG": raw_type_counts["BG"],
    }

    # Every request's approver set is one of exactly two fixed lists (Motor
    # vs. everything else), so resolve both once instead of querying per row.
    approvers_by_product = {
        "MOTOR": list(SpecialRateRequest.assigned_approvers_for_product("MOTOR")),
        "DEFAULT": list(SpecialRateRequest.assigned_approvers_for_product("LIFE")),
    }
    is_full_admin = request.user.groups.filter(name="ADMIN").exists()

    requests_list = list(requests_qs)
    for req in requests_list:
        approvers = approvers_by_product["MOTOR" if req.product == "MOTOR" else "DEFAULT"]
        req.assigned_approver_names = ", ".join(u.get_full_name() or u.username for u in approvers)
        req.can_review = is_full_admin or any(u.pk == request.user.pk for u in approvers)

    return render(request, "special_rate_requests_review.html", {
        "requests": requests_list,
        "status_counts": status_counts,
        "rate_type_counts": rate_type_counts,
        "total_requests": requests_qs.count(),
    })


def update_special_rate_request_status(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            request_id = data.get("request_id")
            new_status = (data.get("status") or "").upper()

            if not request_id or not new_status:
                return JsonResponse({"success": False, "message": "Request ID and status are required."})

            valid_statuses = ["PENDING", "APPROVED", "REJECTED"]
            if new_status not in valid_statuses:
                return JsonResponse({"success": False, "message": "Invalid status."})

            req_obj = SpecialRateRequest.objects.get(id=request_id)
            if not req_obj.can_be_reviewed_by(request.user):
                return JsonResponse({"success": False, "message": "You are not an assigned approver for this request."}, status=403)

            req_obj.status = new_status
            req_obj.reviewed_by = request.user
            req_obj.reviewed_at = timezone.now()
            req_obj.save()

            return JsonResponse({"success": True})

        except SpecialRateRequest.DoesNotExist:
            return JsonResponse({"success": False, "message": "Request not found."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})

# =========================================================
# TICKETING SYSTEM VIEWS
# =========================================================
# Maps the human-readable labels the ticket-raising JS saves (see
# formatKey() in policy_lock_checker.html / motor_payout_rates.html) back to
# the GET param names /motor-payout-rates/ actually reads. Fields with no
# equivalent there (Vehicle No., Policy Holder Name) are simply dropped.
TICKET_TO_MOTOR_PAYOUT_PARAMS = {
    "Target Date": "target_date",
    "Product": "product",
    "Sub Product": "sub_product",
    "Make Model Class": "make_model_class",
    "Make Name (Code)": "make_names",
    "RTO Code": "rto_code",
    "CC / GVW": "cc",
    "Fuel": "fuel",
    "Exact SC": "sc",
    "Manufacturing Year": "mfg_year",
    "Zero Dep (ZD)": "is_zd",
    "CPA": "is_cpa",
    "NCB": "is_ncb",
}
MOTOR_PAYOUT_PARAM_NAMES = set(TICKET_TO_MOTOR_PAYOUT_PARAMS.values())


def _build_motor_payout_rates_url(ticket):
    """_build_motor_payout_queryset (the search/filter backend) accepts either
    a numeric id or the plain text name for product/sub_product/fuel/
    make_model_class, so it finds the right rows either way. But the
    motor_payout_rates.html <select> dropdowns pick their "selected" option
    by comparing against option value="{{ id }}" — a ticket's saved
    human-readable name (e.g. "Private Car") never matches that, so the
    dropdown silently falls back to "Select..." even though the search
    itself worked. Resolve name -> id below so the destination page's
    dropdowns actually show the right option, not just the right results.

    Only MOTOR tickets: Health/Life payloads reuse labels like "Product" and
    "Sub Product" for their own (non-vehicle) fields, so applying this
    mapping to them would build a nonsense motor-rates search."""
    if ticket.category != "MOTOR" or not ticket.form_payload:
        return None

    params = {}
    for label, value in ticket.form_payload.items():
        if not value:
            continue
        param_name = TICKET_TO_MOTOR_PAYOUT_PARAMS.get(label)
        if not param_name and label in MOTOR_PAYOUT_PARAM_NAMES:
            # Older tickets saved the raw field name instead of the
            # human-readable label — already a valid param, use as-is.
            param_name = label
        if param_name:
            params[param_name] = value

    if not params:
        return None

    # Resolve the product's real name regardless of whether the ticket saved
    # an id or a name for it — NA_MAKE_MODEL_MAP below is keyed by name, and
    # older tickets may have already saved a raw id here (see the "already a
    # valid param" branch above).
    product_param = str(params.get("product") or "").strip()
    if product_param.isdigit():
        prod = ProductMaster.objects.filter(id=product_param).first()
        resolved_product_name = prod.name if prod else ""
    else:
        resolved_product_name = product_param
        if product_param:
            prod = ProductMaster.objects.filter(name__iexact=product_param).first()
            if prod:
                params["product"] = prod.id

    if params.get("sub_product") and not str(params["sub_product"]).isdigit():
        sub_prod = SubProductMaster.objects.filter(name__iexact=str(params["sub_product"]).strip()).first()
        if sub_prod:
            params["sub_product"] = sub_prod.id

    if params.get("fuel") and not str(params["fuel"]).isdigit():
        fuel_obj = FuelTypeMaster.objects.filter(name__iexact=str(params["fuel"]).strip()).first()
        if fuel_obj:
            params["fuel"] = fuel_obj.id

    # Make Model Class is the one exception: motor_payout_rates.html's JS
    # swaps in string-valued options (value == class name, e.g. "Car")
    # for products in NA_MAKE_MODEL_MAP instead of the usual database-id
    # options, so for those products the saved name is already the correct
    # value and must NOT be converted to an id.
    if params.get("make_model_class") and not str(params["make_model_class"]).isdigit():
        if resolved_product_name not in NA_MAKE_MODEL_MAP:
            mmc = MakeModelClassMaster.objects.filter(name__iexact=str(params["make_model_class"]).strip()).first()
            if mmc:
                params["make_model_class"] = mmc.id

    return reverse("motor_payout_rates") + "?" + urlencode(params)


def ticket_dashboard(request):
    qs = SupportTicket.objects.select_related('user').all()

    # Keyed without hyphens ("FOLLOW-UP" -> FOLLOWUP) since Django template
    # variable lookups can't contain a "-" (`status_counts.FOLLOW-UP` is a
    # TemplateSyntaxError, not a lookup miss).
    raw_counts = {"OPEN": 0, "FOLLOW-UP": 0, "CLOSED": 0}
    for row in qs.values('status').annotate(n=Count('id')):
        if row['status'] in raw_counts:
            raw_counts[row['status']] = row['n']
    status_counts = {
        "OPEN": raw_counts["OPEN"],
        "FOLLOWUP": raw_counts["FOLLOW-UP"],
        "CLOSED": raw_counts["CLOSED"],
    }

    category_counts = {"MOTOR": 0, "HEALTH": 0, "LIFE": 0}
    for row in qs.values('category').annotate(n=Count('id')):
        if row['category'] in category_counts:
            category_counts[row['category']] = row['n']

    total_tickets = qs.count()

    tickets = list(qs)
    for ticket in tickets:
        ticket.motor_payout_rates_url = _build_motor_payout_rates_url(ticket)

    return render(request, "ticket_dashboard.html", {
        "tickets": tickets,
        "status_counts": status_counts,
        "category_counts": category_counts,
        "total_tickets": total_tickets,
    })

def create_ticket_api(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            remarks = data.get("remarks", "").strip()
            form_payload = data.get("form_payload", {})
            category = (data.get("category") or "MOTOR").strip().upper()

            if not remarks:
                return JsonResponse({"success": False, "message": "Remarks are required."})

            if category not in dict(SupportTicket.CATEGORY_CHOICES):
                category = "MOTOR"

            # 1. Translate Product ID to Name
            if form_payload.get("Product") and str(form_payload["Product"]).isdigit():
                obj = ProductMaster.objects.filter(id=int(form_payload["Product"])).first()
                if obj: form_payload["Product"] = obj.name
                
            # 2. Translate Sub Product ID to Name
            if form_payload.get("Sub Product") and str(form_payload["Sub Product"]).isdigit():
                obj = SubProductMaster.objects.filter(id=int(form_payload["Sub Product"])).first()
                if obj: form_payload["Sub Product"] = obj.name
                
            # 3. Translate Make Model Class ID to Name
            if form_payload.get("Make Model Class") and str(form_payload["Make Model Class"]).isdigit():
                obj = MakeModelClassMaster.objects.filter(id=int(form_payload["Make Model Class"])).first()
                if obj: form_payload["Make Model Class"] = obj.name
                
            # 4. Translate Fuel ID to Name
            if form_payload.get("Fuel") and str(form_payload["Fuel"]).isdigit():
                obj = FuelTypeMaster.objects.filter(id=int(form_payload["Fuel"])).first()
                if obj: form_payload["Fuel"] = obj.name

            ticket = SupportTicket.objects.create(
                user=request.user,
                remarks=remarks,
                form_payload=form_payload,
                category=category,
            )
            return JsonResponse({"success": True, "ticket_id": ticket.id})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
            
    return JsonResponse({"success": False, "message": "Invalid request method."})

def update_ticket_status(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ticket_id = data.get("ticket_id")
            new_status = data.get("status")

            if not ticket_id or not new_status:
                return JsonResponse({"success": False, "message": "Ticket ID and Status are required."})

            ticket = SupportTicket.objects.get(id=ticket_id)

            valid_statuses = ["OPEN", "FOLLOW-UP", "CLOSED"]
            if new_status.upper() not in valid_statuses:
                return JsonResponse({"success": False, "message": "Invalid status."})

            ticket.status = new_status.upper()
            ticket.save()
            
            return JsonResponse({"success": True})
            
        except SupportTicket.DoesNotExist:
             return JsonResponse({"success": False, "message": "Ticket not found."})
        except Exception as e:
            return JsonResponse({"success": False, "message": str(e)})
            
    return JsonResponse({"success": False, "message": "Invalid request method."})

# =========================================================
# AUTOMATED MIS PAYOUT CALCULATION VIEWS
# =========================================================

def mis_payout_automation(request):
    msg = ""
    error = ""
    if request.method == 'POST':
        form = MISUploadForm(request.POST, request.FILES)
        if form.is_valid():
            mis_obj = form.save(commit=False)
            mis_obj.uploaded_by = request.user
            mis_obj.save()

            # Handed off to Celery so the UI doesn't block while Pandas does the heavy lifting.
            # Task id is saved so a later Cancel action can revoke this exact job.
            async_result = process_mis_mapping_task.delay(mis_obj.id)
            mis_obj.celery_task_id = async_result.id
            mis_obj.save(update_fields=['celery_task_id'])

            msg = "File uploaded successfully. The mapping engine has started processing in the background."
            return redirect('mis_payout_automation')
        else:
            error = "Please upload a valid CSV or Excel file."
    else:
        form = MISUploadForm()
    
    files = MISFile.objects.all().order_by('-created_at')[:50]
    
    return render(request, 'insurance/mis_payout_automation.html', {
        'form': form,
        'files': files,
        'msg': msg,
        'error': error
    })

def download_processed_mis(request, file_id):
    mis_obj = get_object_or_404(MISFile, id=file_id)
    if not mis_obj.processed_file:
        return HttpResponse("Processed file is not available yet. Please check back when status is COMPLETED.", status=404)

    try:
        content_type, _ = mimetypes.guess_type(mis_obj.processed_file.name)
        response = HttpResponse(mis_obj.processed_file, content_type=content_type or 'application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{mis_obj.processed_file.name.split("/")[-1]}"'
        return response
    except Exception:
        # The DB row can outlive the actual object in storage (deleted/moved
        # out from under it, a storage outage, etc.) - surface a clean
        # message instead of a raw 500 from whatever the storage backend
        # raised trying to open it.
        logger.exception("Could not read processed MIS file %s (id=%s) from storage", mis_obj.processed_file.name, mis_obj.id)
        return HttpResponse("The processed file could not be retrieved from storage. Please contact support.", status=404)

def cancel_mis_processing(request, file_id):
    if request.method != "POST":
        return JsonResponse({"success": False, "message": "Invalid request."})

    mis_obj = get_object_or_404(MISFile, id=file_id)

    if mis_obj.status not in ('PENDING', 'PROCESSING'):
        return JsonResponse({
            "success": False,
            "message": f"File #{file_id} is already {mis_obj.status.title()} — nothing to cancel."
        })

    # Best-effort OS-level termination — not the source of truth for the UI.
    # A missing task id or an unreachable broker must never block the DB
    # update below, which is what the dashboard actually reflects.
    if mis_obj.celery_task_id:
        try:
            from project.celery import app as celery_app
            celery_app.control.revoke(mis_obj.celery_task_id, terminate=True, signal='SIGTERM')
        except Exception:
            logger.exception("Failed to revoke Celery task %s for MISFile %s", mis_obj.celery_task_id, file_id)

    # Conditional update guards the race where the task completes/fails right
    # as Cancel is clicked — only flip status if it's still PENDING/PROCESSING,
    # never overwrite a real COMPLETED/FAILED result.
    updated = MISFile.objects.filter(id=file_id, status__in=['PENDING', 'PROCESSING']).update(
        status='CANCELLED',
        processed_at=timezone.now(),
        error_message=f"Cancelled by {request.user.username}.",
    )

    if not updated:
        mis_obj.refresh_from_db()
        return JsonResponse({
            "success": False,
            "message": f"File #{file_id} finished as {mis_obj.status.title()} before it could be cancelled."
        })

    return JsonResponse({"success": True, "message": f"File #{file_id} cancelled."})

def mis_mapping_dashboard(request):
    mappings = MappingConfiguration.objects.all()
    return render(request, 'insurance/mis_mapping_dashboard.html', {'mappings': mappings})


# --- JSON Schema injected dynamically into the Dropdowns for cross-table builder ---
def get_rule_schema_context():
    return {
        'MIS_File': [], 
        'RateMaster': [
            'insurance_company', 'product', 'sub_product', 'policy_type', 
            'fuel_type', 'make_model_class', 'new_vehicle_makes', 'new_rto_list', 
            'cc_range', 'sc_range', 'age_range', 'date_range', 
            'is_ncb', 'is_cpa', 'is_zd'
        ]
    }

def add_mis_mapping(request):
    if request.method == 'POST':
        form = MappingConfigurationForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('mis_mapping_dashboard')
    else:
        form = MappingConfigurationForm()
        
    return render(request, 'insurance/mis_mapping_form.html', {
        'form': form, 
        'title': 'Add Multi-Table Mapping Rule',
        'schema_json': json.dumps(get_rule_schema_context())
    })

def edit_mis_mapping(request, pk):
    mapping = get_object_or_404(MappingConfiguration, pk=pk)
    if request.method == 'POST':
        form = MappingConfigurationForm(request.POST, instance=mapping)
        if form.is_valid():
            form.save()
            return redirect('mis_mapping_dashboard')
    else:
        form = MappingConfigurationForm(instance=mapping)
        
    return render(request, 'insurance/mis_mapping_form.html', {
        'form': form, 
        'title': 'Edit Multi-Table Mapping Rule',
        'schema_json': json.dumps(get_rule_schema_context())
    })

def delete_mis_mapping(request, pk):
    mapping = get_object_or_404(MappingConfiguration, pk=pk)
    mapping.delete()
    return redirect('mis_mapping_dashboard')